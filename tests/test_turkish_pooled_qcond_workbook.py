from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

import scripts.build_clean_workbook as builder
from scripts.build_clean_workbook import build_turkish_pooled_qcond, build_turkish_pooled_qcond_provenance


def _report(path: Path) -> None:
    cells = [
        (f"Q{i:02d}", "Qwen", "audio_only", "not_applicable") for i in range(1, 6)
    ] + [
        (f"G{i:02d}", "Gemma 4", "audio_only", "not_applicable") for i in range(1, 6)
    ]
    cells += [
        (f"Q{i:02d}", "Qwen", "text_only", "native") for i in (2, 3)
    ] + [
        (f"G{i:02d}", "Gemma 4", "text_only", "native") for i in (2, 3)
    ]
    # The workbook renderer only needs the report cardinalities and summary
    # fields; the production report test enforces the full matrix semantics.
    while len(cells) < 10:
        cells.append((f"X{len(cells):02d}", "Qwen", "audio_text", "native"))
    summary = []
    seed_results = []
    fold_results = []
    provenance = {}
    for index, (cell_id, model, modality, transcript) in enumerate(cells[:10]):
        for route in ("teacher_forced", "logreg", "xgb_optuna100"):
            base = {
                "cell_id": cell_id, "model": model, "modality": modality,
                "transcript_condition": transcript, "route": route,
                "backend": "test", "provenance_keys": ",".join(f"prov-{index}-{seed}-{fold}" for seed in (7, 1337, 2024) for fold in range(5)),
                "baseline_positive_status": "baseline unavailable or not reportable",
                "baseline_negative_status": "baseline unavailable or not reportable",
                "pooled_oof_status": "not reported",
            }
            summary.append({
                **base,
                **{name: value for name, value in {
                    "positive_question_macro_f1_seed_mean": 0.51,
                    "positive_question_macro_f1_seed_sample_sd": 0.01,
                    "negative_question_macro_f1_seed_mean": 0.52,
                    "negative_question_macro_f1_seed_sample_sd": 0.02,
                    "combined_macro_f1_seed_mean": 0.53,
                    "combined_macro_f1_seed_sample_sd": 0.03,
                    "positive_question_positive_f1_seed_mean": 0.54,
                    "negative_question_positive_f1_seed_mean": 0.55,
                    "combined_positive_f1_seed_mean": 0.56,
                    "baseline_positive_macro_f1": None,
                    "baseline_negative_macro_f1": None,
                    "pooled_minus_positive_baseline_macro_f1": None,
                    "pooled_minus_negative_baseline_macro_f1": None,
                }.items()},
            })
            for seed in (7, 1337, 2024):
                seed_results.append({
                    **base, "seed": seed, "fold_count": 5,
                    "positive_question_macro_f1_fold_mean": 0.51,
                    "negative_question_macro_f1_fold_mean": 0.52,
                    "combined_macro_f1_fold_mean": 0.53,
                    "positive_question_positive_f1_fold_mean": 0.54,
                    "negative_question_positive_f1_fold_mean": 0.55,
                    "combined_positive_f1_fold_mean": 0.56,
                })
            for seed in (7, 1337, 2024):
                for fold in range(5):
                    key = f"prov-{index}-{route}-{seed}-{fold}"
                    provenance[key] = {"attempt_id": key, "locally_verified": True, "reportable": True}
                    fold_results.append({"cell_id": cell_id, "route": route, "seed": seed, "fold": fold, "provenance_key": key})
    payload = {
        "schema_version": "audiollm.turkish_pooled_qcond_report.v1",
        "group_id": "turkish-pooled-qcond-clean-v1-20260903", "deployment_id": "deployment-test", "source_git_sha": "a" * 40,
        "validation": {"status": "passed"},
        "tables": {"summary": summary, "seed_results": seed_results[:90], "fold_results": fold_results[:450]},
        "provenance_index": {key: value for key, value in list(provenance.items())[:450]},
    }
    path.write_text(json.dumps(payload), encoding="utf-8")


def test_pooled_workbook_sheet_and_provenance_are_rendered_from_report(tmp_path: Path) -> None:
    report_path = tmp_path / "report.json"
    _report(report_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    build_turkish_pooled_qcond(workbook, report_path=report_path)
    ws = workbook["Turkish Pooled QCond"]
    assert ws["A1"].value == "Turkish pooled question-conditioned training"
    assert ws["E6"].value == 0.51
    assert ws["P6"].value is None
    assert "baseline unavailable" in str(ws["T6"].value)
    output = tmp_path / "pooled.xlsx"
    workbook.save(output)
    reopened = load_workbook(output, data_only=False)
    assert "Turkish Pooled QCond" in reopened.sheetnames
    assert reopened["Turkish Pooled QCond"]["I6"].value == 0.53

    provenance_book = Workbook()
    provenance_book.remove(provenance_book.active)
    provenance_ws = provenance_book.create_sheet("Provenance")
    rows = []
    def put(*values):
        rows.append(values)
    build_turkish_pooled_qcond_provenance(provenance_ws, put, report_path=report_path)
    assert len(rows) == 300
    assert all(value[3] is None for value in rows if "baseline" in str(value[2]).lower() or "pooled minus" in str(value[2]).lower())


def test_workbook_without_pooled_report_keeps_historical_sheet_values(tmp_path: Path) -> None:
    report_path = tmp_path / "report.json"
    _report(report_path)

    def build(include_pooled: bool) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)
        builder.build_summary(workbook, detailed=False)
        builder.build_gemma_vs_qwen(workbook)
        builder.build_native_vs_english(workbook)
        if include_pooled:
            builder.build_turkish_pooled_qcond(workbook, report_path=report_path)
        builder.build_packed30(workbook)
        builder.build_provenance(
            workbook,
            detailed=False,
            turkish_pooled_qcond_report_path=report_path if include_pooled else None,
        )
        return workbook

    without = build(False)
    with_pooled = build(True)
    assert "Turkish Pooled QCond" not in without.sheetnames
    assert "Turkish Pooled QCond" in with_pooled.sheetnames
    for sheet in without.sheetnames:
        left = without[sheet]
        right = with_pooled[sheet]
        assert left.max_row <= right.max_row, sheet
        assert left.max_column == right.max_column, sheet
        for row in range(1, left.max_row + 1):
            for column in range(1, left.max_column + 1):
                assert left.cell(row, column).value == right.cell(row, column).value, (
                    sheet,
                    row,
                    column,
                )
