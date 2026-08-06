from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from scripts.update_androids_hidden_workbook import (
    HEADS,
    MODALITIES,
    _update_combined,
    _update_graph,
    _update_summary,
    _update_xgboost_head,
    _validate_workbook,
)


def _acceptance() -> dict[str, object]:
    pooled: dict[str, object] = {}
    for modality_index, modality in enumerate(MODALITIES):
        for head_index, head in enumerate(HEADS):
            value = (modality_index * len(HEADS) + head_index + 1) / 100.0
            pooled[f"{modality}/{head}"] = {
                "metrics": {
                    "accuracy": value,
                    "positive_f1": value + 0.001,
                    "precision": value + 0.002,
                    "recall": value + 0.003,
                    "macro_f1": value + 0.004,
                    "negative_f1": value + 0.005,
                    "auroc": value + 0.006,
                    "confusion_matrix": [[1, 2], [3, 4]],
                }
            }
    return {
        "status": "passed",
        "mode": "production",
        "counts": {"pooled_subjects_per_result": 116},
        "pooled_results": pooled,
    }


def test_androids_workbook_rows_tables_and_chart_are_updated(tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    source = repo_root / "docs" / "archive" / "results_20260806" / "depression_results_combined_with_posf1_graphs.xlsx"
    workbook_path = tmp_path / source.name
    shutil.copy2(source, workbook_path)
    wb = load_workbook(workbook_path)
    payload = _acceptance()
    _update_summary(wb["Summary"], payload)
    _update_xgboost_head(wb["XGBoost Head"], payload)
    _update_combined(wb["Combined Results"], payload)
    _update_graph(wb["Androids PosF1 Graph"], payload)
    wb.save(workbook_path)

    _validate_workbook(workbook_path)
    checked = load_workbook(workbook_path, data_only=False)
    assert checked["XGBoost Head"].tables["Table1"].ref == "A1:J46"
    assert checked["Combined Results"].tables["Table_1"].ref == "A4:M64"
    assert checked["Androids PosF1 Graph"]._charts[0].ser.__len__() == 3
    assert checked["Combined Results"]["L56"].value == 128.0 / 180.0
