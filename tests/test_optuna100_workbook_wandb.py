"""Task 5 tests: Optuna-100 workbook structure, explicit selections, and
safe W&B resolver entries."""

from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from src.experiment_tracking import registry as registry_mod

ROOT = Path(__file__).resolve().parents[1]
SELECTION_PATH = ROOT / "experiments/definitions/workbook_optuna100_selection.yaml"
WANDB_SELECTION_PATH = ROOT / "experiments/definitions/workbook_wandb_selection.yaml"
WORKBOOK_PATH = ROOT / "depression_results_clean.xlsx"
BUILDER_PATH = ROOT / "scripts/build_clean_workbook.py"

EXPECTED_COUNTS = {
    "Native": 126,
    "English": 80,
    "Symmetric merged": 36,
    "DAIC official development": 6,
}


def test_optuna100_selection_definitions_exact_counts_and_unique() -> None:
    payload = yaml.safe_load(SELECTION_PATH.read_text(encoding="utf-8"))
    assert payload["schema_version"] == "audiollm.selected_results.v1"
    selections = payload["selections"]
    assert len(selections) == 248
    families = Counter(selection["family"] for selection in selections)
    assert dict(families) == EXPECTED_COUNTS
    keys = [
        (selection["cell"], selection["fold"], selection.get("stage"))
        for selection in selections
    ]
    assert len(keys) == len(set(keys))
    for selection in selections:
        assert selection["metric"] == "macro_f1"
        assert selection["namespace"] == "headline/binary_strict"
        assert selection["view"] == "harmonized_all_windows_full_coverage"
        assert selection["aggregation"] == "subject_level"
        assert selection["attempt_id"] is None
        assert selection["protocol_profile"] == "harmonized_optuna100_v1"
    backends = Counter(selection["backend"] for selection in selections)
    assert backends["qwen_hidden_xgb_optuna100"] == 106
    assert backends["gemma4_hidden_xgb_optuna100"] == 106
    assert backends["qwen_hidden_xgb_optuna100_symmetric_merged"] == 18
    assert backends["gemma4_hidden_xgb_optuna100_symmetric_merged"] == 18


def test_wandb_selection_entries_are_safe_and_unique() -> None:
    payload = yaml.safe_load(WANDB_SELECTION_PATH.read_text(encoding="utf-8"))
    entries = payload["entries"]
    ids = [entry["selection_id"] for entry in entries]
    assert len(ids) == len(set(ids))
    optuna_entries = [entry for entry in entries if "Optuna100|" in entry["selection_id"]]
    assert len(optuna_entries) == 248
    assert all(entry["wandb_policy"] == "skip_not_run" for entry in optuna_entries)
    assert all(entry["source_type"] == "hidden_classifier" for entry in optuna_entries)
    for entry in optuna_entries:
        assert "study not run yet" in entry["blocking_reasons"]
    # The existing Gemma legacy reconciliation entries are untouched.
    legacy = [entry for entry in entries if entry["wandb_policy"] == "pending_wandb_reconciliation"]
    assert len(legacy) == 15
    merged_final_tf = [
        entry for entry in entries
        if entry["selection_id"].startswith("Gemma merged|Final (DAIC test)")
        and entry["selection_id"].endswith("Teacher-forced")
    ]
    assert len(merged_final_tf) == 3
    assert all(
        entry["required_evaluations"][0]["dataset"] == "daic"
        for entry in merged_final_tf
    )


def test_workbook_regeneration_keeps_compact_sheet_set_with_qwen_vs_gemma() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "Qwen vs Gemma",
        "Native vs EN",
        "DAIC Packed30 Family",
        "Provenance",
    ]
    sheet = workbook["Qwen vs Gemma"]
    value_cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column == 5 and cell.row > 4:
                value_cells.append(cell.value)
    assert value_cells and any(isinstance(value, (int, float)) for value in value_cells)
    assert not any(isinstance(value, (int, float)) and (value < 0 or value > 1) for value in value_cells)
    methods = {sheet.cell(r, 4).value for r in range(5, sheet.max_row + 1)}
    assert {"Teacher-forced", "LogReg head", "XGBoost"} <= methods

    merged_final_tf = {
        sheet.cell(row, 3).value: sheet.cell(row, 6).value
        for row in range(5, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "Merged — Final (DAIC test)"
        and sheet.cell(row, 4).value == "Teacher-forced"
    }
    assert merged_final_tf == {
        "Audio + Text": pytest.approx(0.7257294429708223),
        "Audio only": pytest.approx(0.5190058479532164),
        "Text only": pytest.approx(0.7755968169761273),
    }


def test_workbook_and_builder_hashes_recorded_in_wandb_selection() -> None:
    import hashlib

    payload = yaml.safe_load(WANDB_SELECTION_PATH.read_text(encoding="utf-8"))
    block = payload["workbook"]

    def digest(path: Path) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1 << 20), b""):
                h.update(chunk)
        return h.hexdigest()

    assert block["sha256"] == digest(ROOT / block["path"])
    assert block["builder_sha256"] == digest(ROOT / block["builder_path"])


def test_wandb_resolver_never_exports_optuna100_entries_without_evidence(
    tmp_path: Path,
) -> None:
    from src.experiment_tracking.workbook_selection import (
        build_dependency_inventory,
        load_selection,
        resolve_manifest,
    )

    empty_scan = tmp_path / "empty_scan"
    empty_scan.mkdir()
    db_path = tmp_path / "registry.sqlite"
    registry_mod.rebuild_registry(empty_scan, db_path)

    selection = load_selection(WANDB_SELECTION_PATH)
    inventory = build_dependency_inventory(WORKBOOK_PATH, builder_path=BUILDER_PATH)
    manifest = resolve_manifest(
        selection, inventory, db_path=db_path, selection_path=WANDB_SELECTION_PATH
    )
    optuna_statuses = [
        entry["status"]
        for entry in manifest["entries"]
        if entry["selection_id"].startswith("Optuna100|")
    ]
    assert len(optuna_statuses) == 248
    assert all(
        status in ("skip_not_run", "stale_not_in_workbook") for status in optuna_statuses
    )
    assert not any(
        status in ("sync", "resolved", "blocked") for status in optuna_statuses
    )
    units = manifest.get("units") or []
    optuna_unit_ids = {
        unit["wandb_run_id"]
        for unit in units
        if any(selection_id.startswith("Optuna100|") for selection_id in unit["selection_ids"])
    }
    assert optuna_unit_ids == set()


def test_generator_is_idempotent() -> None:
    import subprocess

    before = WANDB_SELECTION_PATH.read_text(encoding="utf-8")
    result = subprocess.run(
        ["python", str(ROOT / "tools/build_optuna100_wandb_selection.py")],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr
    after = WANDB_SELECTION_PATH.read_text(encoding="utf-8")
    assert before == after
