import json

from openpyxl import Workbook

from scripts import build_clean_workbook as builder


def test_native_vs_en_sheet_has_locked_summary_and_seed_detail_cardinalities(tmp_path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    builder.build_native_vs_english(workbook, report_path=tmp_path / "missing-report.json")

    sheet = workbook["Native vs EN"]
    assert sheet.max_column == 19
    assert [sheet.cell(5, column).value for column in range(1, 5)] == [
        "Endpoint",
        "Dataset / target",
        "Backbone",
        "Head",
    ]
    assert [sheet.cell(5, column).value for column in range(5, 17)] == [
        "Native macro mean",
        "Native macro SD",
        "English macro mean",
        "English macro SD",
        "Δ macro mean",
        "Δ macro SD",
        "Native positive mean",
        "Native positive SD",
        "English positive mean",
        "English positive SD",
        "Δ positive mean",
        "Δ positive SD",
    ]
    assert sheet.cell(6, 1).value == "Standalone"
    assert sheet.cell(6, 2).value == "D3TEC"
    assert sheet.cell(6, 3).value == "Qwen"
    assert sheet.cell(6, 4).value == "LogReg"
    assert sheet.cell(6, 5).value is None

    summary_rows = [
        row for row in range(6, 30)
        if sheet.cell(row, 1).value in {"Standalone", "Merged CV", "Final merged DAIC"}
    ]
    assert len(summary_rows) == 24
    detail_header = next(
        row for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "Endpoint"
        and sheet.cell(row, 5).value == "Training seed"
    )
    detail_rows = [
        row for row in range(detail_header + 1, sheet.max_row + 1)
        if sheet.cell(row, 5).value in {7, 1337, 2024}
    ]
    assert len(detail_rows) == 72


def test_native_vs_en_sheet_marks_missing_evidence_without_inventing_values(tmp_path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    builder.build_native_vs_english(workbook, report_path=tmp_path / "missing-report.json")
    sheet = workbook["Native vs EN"]

    assert sheet.cell(6, 5).value is None
    assert sheet.cell(6, 19).value.endswith("[pending_local_evidence]")
    detail_header = next(
        row for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "Endpoint"
        and sheet.cell(row, 5).value == "Training seed"
    )
    assert sheet.cell(detail_header + 1, 6).value is None
    assert sheet.cell(detail_header + 1, 19).value == "pending_local_evidence"


def test_native_vs_en_selected_results_validate_against_report_key(tmp_path) -> None:
    cell = "Native vs EN v2|D3TEC|Standalone / Qwen|LogReg — Native macro mean"
    selected = tmp_path / "selected.json"
    selected.write_text(
        json.dumps(
            {
                "selections": [
                    {
                        "cell": cell,
                        "sheet": "Native vs EN",
                        "source_type": "native_en_report",
                        "status": "selected",
                        "value": 0.5,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    ok, report = builder.validate_selected_results(selected, {cell: 0.5})

    assert ok
    assert report == []
