from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import yaml

from .canonical import canonical_sha256, sha256_file
from .registry import RegistryError, connect, show_attempt
from .wandb_export import legacy_wandb_id, wandb_run_id_for_attempt

SELECTION_SCHEMA_VERSION = "audiollm.workbook_wandb_selection.v1"
DEPENDENCY_INVENTORY_SCHEMA_VERSION = "audiollm.workbook_dependency_inventory.v1"
MANIFEST_SCHEMA_VERSION = "audiollm.workbook_wandb_manifest.v1"

PROVENANCE_SHEET_NAME = "Provenance"
PROVENANCE_COLUMNS = (
    "experiment",
    "dataset",
    "modality",
    "method",
    "macro_f1",
    "source_run",
    "aggregation_view",
    "local_artifact",
    "verification",
)
KEY_COLUMNS = ("experiment", "dataset", "modality", "method")

SOURCE_TYPES = (
    "ordinary_qwen",
    "hidden_classifier",
    "merged",
    "audit_derived",
    "mn5_only",
    "not_run",
)

WANDB_POLICIES = (
    "sync",
    "pending_local_evidence",
    "pending_importer_support",
    "pending_wandb_reconciliation",
    "quarantine_ambiguous",
    "skip_not_run",
    "skip_derived_only",
)

SOURCE_TYPE_LABELS = {
    "ordinary_qwen": "ordinary Qwen execution",
    "hidden_classifier": "hidden-classifier execution",
    "merged": "merged execution",
    "audit_derived": "audit-derived value (no execution run of its own)",
    "mn5_only": "MN5-only execution, not locally synced",
    "not_run": "not run / intentionally blank",
}

_FAMILY_RULES: tuple[tuple[str, str | None, str], ...] = (
    ("Standalone heads", None, "hidden_classifier"),
    ("Standalone", "Fine-tuned Qwen", "ordinary_qwen"),
    ("Merged DAIC official", None, "merged"),
    ("Merged CV", None, "merged"),
    ("EN heads", None, "hidden_classifier"),
    ("DAIC packed30 family", "Qwen TF", "ordinary_qwen"),
    ("DAIC packed30 family", "LogReg raw", "hidden_classifier"),
    ("DAIC packed30 family", "XGBoost raw", "hidden_classifier"),
    ("Harmonized standalone", "Fine-tuned Qwen", "ordinary_qwen"),
    ("Harmonized heads", None, "hidden_classifier"),
    ("Harmonized merged final", None, "merged"),
    ("Harmonized merged CV", None, "merged"),
    ("Gemma 4 DAIC", "Gemma 4 macro-F1", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 positive-F1", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 accuracy", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 confusion matrix", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 invalid count", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 teacher-forced head macro-F1", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 teacher-forced head positive-F1", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 teacher-forced head accuracy", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 teacher-forced head precision", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 teacher-forced head recall", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 teacher-forced head confusion matrix", "ordinary_qwen"),
    ("Gemma 4 DAIC", "Gemma 4 LogReg raw hidden head", "hidden_classifier"),
    ("Gemma 4 DAIC", "Gemma 4 XGBoost raw hidden head", "hidden_classifier"),
    ("DAIC official development", "Qwen teacher-forced", "ordinary_qwen"),
    ("DAIC official development", "Gemma 4 teacher-forced", "ordinary_qwen"),
    ("DAIC official development heads", None, "hidden_classifier"),
    ("DAIC Head Ablation", None, "hidden_classifier"),
)

_NOT_RUN_HINTS = ("not run", "optuna not run", "not_run", "never run")
_MN5_HINTS = ("mn5", "output_model_en", "not synced")

DEFAULT_WORKBOOK_PATH = "depression_results_clean.xlsx"
DEFAULT_BUILDER_PATH = "scripts/build_clean_workbook.py"


class WorkbookSelectionError(Exception):
    pass


# --------------------------------------------------------------------------- workbook
def read_provenance_sheet(workbook_path: str | Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    target = Path(workbook_path)
    if not target.is_file():
        raise WorkbookSelectionError(f"workbook does not exist: {target}")
    try:
        wb = load_workbook(target, read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises a variety of parse errors
        raise WorkbookSelectionError(f"workbook unreadable: {target}: {error}") from error
    try:
        if PROVENANCE_SHEET_NAME not in wb.sheetnames:
            raise WorkbookSelectionError(
                f"workbook has no {PROVENANCE_SHEET_NAME} sheet: {target}"
            )
        ws = wb[PROVENANCE_SHEET_NAME]
        header_row: int | None = None
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    for index, row in enumerate(rows):
        if not row:
            continue
        first = str(row[0] or "")
        if "Experiment" in first:
            header_row = index
            break
    if header_row is None:
        raise WorkbookSelectionError(
            f"workbook {target} has no {PROVENANCE_SHEET_NAME} header row"
        )
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        cells = list(row) + [None] * (len(PROVENANCE_COLUMNS) - len(list(row)))
        record = {
            "row_number": row_number,
            "sheet": PROVENANCE_SHEET_NAME,
            **{
                name: (value if isinstance(value, str) else value)
                for name, value in zip(PROVENANCE_COLUMNS, cells)
            },
        }
        for name in PROVENANCE_COLUMNS:
            value = record[name]
            if isinstance(value, str):
                record[name] = value.strip()
        if not any(record[name] not in (None, "") for name in PROVENANCE_COLUMNS):
            continue
        records.append(record)
    return records


def provenance_key_of(record: dict[str, Any]) -> str:
    parts = [record.get(name) for name in KEY_COLUMNS]
    if any(part in (None, "") for part in parts):
        raise WorkbookSelectionError(
            f"provenance row {record.get('row_number')} is missing a key column"
        )
    return "|".join(str(part) for part in parts)


def provenance_key_dict(key: str) -> dict[str, str]:
    experiment, dataset, modality, method = key.split("|", 3)
    return {
        "experiment": experiment,
        "dataset": dataset,
        "modality": modality,
        "method": method,
    }


def classify_source_type(experiment: str, method: str) -> str | None:
    for experiment_rule, method_rule, source_type in _FAMILY_RULES:
        if experiment != experiment_rule:
            continue
        if method_rule is None or method == method_rule:
            return source_type
    if experiment == "Gemma 4 DAIC":
        if method.startswith("Gemma 4 LogReg raw hidden head") or method.startswith(
            "Gemma 4 XGBoost raw hidden head"
        ):
            return "hidden_classifier"
        if method.startswith("Gemma 4 teacher-forced head"):
            return "ordinary_qwen"
    if experiment == "EN translation":
        if method.startswith("Native"):
            return "ordinary_qwen"
        if method.startswith("Translated EN"):
            return "mn5_only"
    if experiment == "DAIC packed30 family":
        return "ordinary_qwen"
    return None


def is_blank_or_not_run(record: dict[str, Any]) -> tuple[bool, str | None]:
    value = record.get("macro_f1")
    verification = str(record.get("verification") or "").lower()
    source = str(record.get("source_run") or "").lower()
    if value in (None, ""):
        if any(hint in verification or hint in source for hint in _NOT_RUN_HINTS):
            return True, "macro blank and source/verification indicates not run"
        return True, "macro blank"
    if any(hint in verification for hint in _NOT_RUN_HINTS):
        return True, "verification indicates not run"
    return False, None


def build_dependency_inventory(
    workbook_path: str | Path,
    *,
    builder_path: str | Path = DEFAULT_BUILDER_PATH,
) -> dict[str, Any]:
    records = read_provenance_sheet(workbook_path)
    rows: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    duplicates: list[dict[str, Any]] = []
    malformed: list[dict[str, Any]] = []
    family_counts: dict[str, int] = {}
    for record in records:
        row = {
            "row_number": record["row_number"],
            "sheet": PROVENANCE_SHEET_NAME,
            "raw": {name: record[name] for name in PROVENANCE_COLUMNS},
        }
        missing = [name for name in KEY_COLUMNS if record.get(name) in (None, "")]
        if missing:
            malformed.append(
                {
                    "row_number": record["row_number"],
                    "reason": f"missing key column(s): {','.join(missing)}",
                    "raw": row["raw"],
                }
            )
            rows.append({**row, "provenance_key": None, "source_type": None})
            continue
        key = provenance_key_of(record)
        if key in seen:
            duplicates.append(
                {
                    "row_number": record["row_number"],
                    "first_row_number": seen[key],
                    "provenance_key": key,
                    "raw": row["raw"],
                }
            )
        else:
            seen[key] = record["row_number"]
        source_type = classify_source_type(str(record["experiment"]), str(record["method"]))
        family_counts[source_type or "unknown"] = family_counts.get(source_type or "unknown", 0) + 1
        blank, blank_reason = is_blank_or_not_run(record)
        row.update(
            {
                "provenance_key": key,
                "source_type": source_type,
                "blank": blank,
                "blank_reason": blank_reason,
                "mn5_only": _mn5_hinted(record),
            }
        )
        rows.append(row)
    return {
        "schema_version": DEPENDENCY_INVENTORY_SCHEMA_VERSION,
        "workbook": {
            "path": str(Path(workbook_path)),
            "sha256": sha256_file(workbook_path),
            "builder_path": str(Path(builder_path)),
            "builder_sha256": (
                sha256_file(builder_path) if Path(builder_path).is_file() else None
            ),
            "provenance_sheet": PROVENANCE_SHEET_NAME,
            "provenance_data_rows": len(rows),
        },
        "rows": rows,
        "duplicates": duplicates,
        "malformed": malformed,
        "summary": {
            "data_rows": len(rows),
            "malformed_rows": len(malformed),
            "duplicate_keys": len(duplicates),
            "families": family_counts,
        },
    }


def _mn5_hinted(record: dict[str, Any]) -> bool:
    text = " ".join(
        str(record.get(name) or "") for name in ("source_run", "aggregation_view", "local_artifact", "verification")
    ).lower()
    return any(hint in text for hint in _MN5_HINTS)


# --------------------------------------------------------------------------- selection YAML
def load_selection(path: str | Path) -> dict[str, Any]:
    target = Path(path)
    if not target.is_file():
        raise WorkbookSelectionError(f"selection file does not exist: {target}")
    try:
        payload = yaml.safe_load(target.read_text(encoding="utf-8"))
    except yaml.YAMLError as error:
        raise WorkbookSelectionError(f"selection YAML unreadable: {target}: {error}") from error
    if not isinstance(payload, dict):
        raise WorkbookSelectionError(f"selection YAML is not an object: {target}")
    if payload.get("schema_version") != SELECTION_SCHEMA_VERSION:
        raise WorkbookSelectionError(
            f"selection schema_version is {payload.get('schema_version')!r}, "
            f"expected {SELECTION_SCHEMA_VERSION!r}"
        )
    workbook = payload.get("workbook")
    if not isinstance(workbook, dict):
        raise WorkbookSelectionError("selection YAML is missing the workbook block")
    for key in ("path", "sha256", "builder_path", "builder_sha256", "provenance_sheet"):
        if key not in workbook:
            raise WorkbookSelectionError(f"selection workbook block is missing {key}")
    entries = payload.get("entries")
    if not isinstance(entries, list) or not entries:
        raise WorkbookSelectionError("selection YAML has no entries")
    seen_ids: set[str] = set()
    for entry in entries:
        _validate_entry(entry)
        selection_id = entry["selection_id"]
        if selection_id in seen_ids:
            raise WorkbookSelectionError(f"duplicate selection_id: {selection_id}")
        seen_ids.add(selection_id)
    return payload


def _validate_entry(entry: Any) -> None:
    if not isinstance(entry, dict):
        raise WorkbookSelectionError("selection entries must be objects")
    for key in ("selection_id", "provenance_key", "source_type", "reason", "dependency_policy", "wandb_policy"):
        if key not in entry:
            raise WorkbookSelectionError(f"selection entry missing {key}: {entry.get('selection_id')}")
    pkey = entry["provenance_key"]
    if not isinstance(pkey, dict) or any(key not in pkey for key in KEY_COLUMNS):
        raise WorkbookSelectionError(
            f"selection entry provenance_key must contain {KEY_COLUMNS}: {entry['selection_id']}"
        )
    if entry["source_type"] not in SOURCE_TYPES:
        raise WorkbookSelectionError(
            f"selection entry has invalid source_type {entry['source_type']!r}: {entry['selection_id']}"
        )
    if entry["wandb_policy"] not in WANDB_POLICIES:
        raise WorkbookSelectionError(
            f"selection entry has invalid wandb_policy {entry['wandb_policy']!r}: {entry['selection_id']}"
        )
    required = entry.get("required_evaluations", [])
    if not isinstance(required, list):
        raise WorkbookSelectionError(f"required_evaluations must be a list: {entry['selection_id']}")
    for spec in required:
        if not isinstance(spec, dict):
            raise WorkbookSelectionError(f"required evaluation must be an object: {entry['selection_id']}")
        for key in ("dataset", "namespace", "backend", "aggregation", "checkpoint_role"):
            if key not in spec:
                raise WorkbookSelectionError(
                    f"required evaluation missing {key}: {entry['selection_id']}"
                )
    for key in ("attempt_ids", "expected_folds"):
        value = entry.get(key, [])
        if not isinstance(value, list):
            raise WorkbookSelectionError(f"{key} must be a list: {entry['selection_id']}")


def selection_sha256(payload: dict[str, Any]) -> str:
    return canonical_sha256(payload)


def payload_hash(plan: dict[str, Any]) -> str:
    return canonical_sha256(plan)


def verify_selection_hashes(
    selection: dict[str, Any],
    *,
    workbook_path: str | Path,
    builder_path: str | Path,
) -> list[str]:
    failures: list[str] = []
    workbook = selection["workbook"]
    expected_workbook = workbook["sha256"]
    actual_workbook = sha256_file(workbook_path)
    if actual_workbook != expected_workbook:
        failures.append(
            f"workbook sha256 mismatch: selection recorded {expected_workbook}, "
            f"file at {workbook_path} hashes to {actual_workbook}"
        )
    builder = Path(builder_path)
    expected_builder = workbook["builder_sha256"]
    if not builder.is_file():
        failures.append(f"builder file missing: {builder_path}")
    elif expected_builder is None or sha256_file(builder) != expected_builder:
        failures.append(
            f"builder sha256 mismatch: selection recorded {expected_builder}, "
            f"file at {builder_path} hashes to {sha256_file(builder)}"
        )
    return failures


# --------------------------------------------------------------------------- manifest resolution
def registry_evidence_hash(connection: Any, attempt_id: str, fold: int) -> str:
    payload = show_attempt(connection, attempt_id, fold)
    return canonical_sha256(
        {
            "attempt_id": attempt_id,
            "fold": fold,
            "attempt": {
                "resolved_config_sha256": payload["attempt"].get("resolved_config_sha256"),
                "manifest_sha256": payload["attempt"].get("manifest_sha256"),
                "split_sha256": payload["attempt"].get("split_sha256"),
                "current_state": payload["attempt"].get("current_state"),
                "git_commit": payload["attempt"].get("git_commit"),
                "git_branch": payload["attempt"].get("git_branch"),
                "git_dirty": payload["attempt"].get("git_dirty"),
            },
            "fold": {
                "fold": payload["folds"][0]["fold"] if payload["folds"] else None,
                "locally_verified": payload["folds"][0]["locally_verified"] if payload["folds"] else None,
                "run_dir": payload["folds"][0]["run_dir"] if payload["folds"] else None,
            },
            "evaluations": [
                {
                    "evaluation_id": entry["evaluation"]["evaluation_id"],
                    "dataset": entry["evaluation"].get("dataset"),
                    "split_name": entry["evaluation"].get("split_name"),
                    "split_protocol": entry["evaluation"].get("split_protocol"),
                    "checkpoint_role": entry["evaluation"].get("checkpoint_role"),
                    "backend": entry["evaluation"].get("backend"),
                    "evaluation_view": entry["evaluation"].get("evaluation_view"),
                    "aggregation": entry["evaluation"].get("aggregation"),
                    "metric_namespace": entry["evaluation"].get("metric_namespace"),
                    "locally_verified": entry["evaluation"].get("locally_verified"),
                    "reportable": entry["evaluation"].get("reportable"),
                    "metrics": [
                        {"metric_name": metric["metric_name"], "metric_value": metric["metric_value"]}
                        for metric in entry["metrics"]
                    ],
                }
                for entry in payload["evaluations"]
            ],
            "artifacts": [
                {
                    "artifact_type": artifact.get("artifact_type"),
                    "role": artifact.get("role"),
                    "path": artifact.get("path"),
                    "sha256": artifact.get("sha256"),
                }
                for artifact in payload["artifacts"]
            ],
        }
    )


def wandb_run_id_for(connection: Any, attempt_id: str, fold: int) -> str:
    payload = show_attempt(connection, attempt_id, fold)
    first_evaluation_id = None
    for entry in sorted(payload["evaluations"], key=lambda item: item["evaluation"]["evaluation_id"]):
        if first_evaluation_id is None:
            first_evaluation_id = entry["evaluation"]["evaluation_id"]
    if str(attempt_id).startswith("legacy-"):
        return legacy_wandb_id(attempt_id, fold, first_evaluation_id or "")
    return wandb_run_id_for_attempt(attempt_id, fold)


def _evaluation_matches(evaluation: dict[str, Any], spec: dict[str, Any]) -> bool:
    for key in ("dataset", "namespace", "backend", "aggregation", "checkpoint_role"):
        if evaluation.get("metric_namespace" if key == "namespace" else key) != spec.get(key):
            return False
    view = spec.get("view")
    if view is not None and evaluation.get("evaluation_view") != view:
        return False
    return True


def _attempts_for_entry(
    connection: Any,
    entry: dict[str, Any],
) -> tuple[list[str], list[str]]:
    reasons: list[str] = []
    attempt_ids = [str(value) for value in entry.get("attempt_ids") or []]
    if attempt_ids:
        existing = [
            connection.execute(
                "SELECT 1 FROM run_attempts WHERE attempt_id = ?", (attempt_id,)
            ).fetchone()
            for attempt_id in attempt_ids
        ]
        missing = [
            attempt_id for attempt_id, row in zip(attempt_ids, existing) if row is None
        ]
        if missing:
            reasons.append(f"attempt_id(s) not in registry: {', '.join(missing)}")
        return [attempt_id for attempt_id, row in zip(attempt_ids, existing) if row is not None], reasons
    name = entry.get("logical_run_name")
    if not isinstance(name, str) or not name:
        return [], ["no attempt_ids and no logical_run_name to resolve dependencies"]
    runs = connection.execute(
        "SELECT lr.logical_run_id, COUNT(a.attempt_id) AS attempt_count "
        "FROM logical_runs lr "
        "JOIN run_attempts a ON a.logical_run_id = lr.logical_run_id "
        "WHERE lr.logical_run_name = ? GROUP BY lr.logical_run_id "
        "ORDER BY lr.logical_run_id",
        (name,),
    ).fetchall()
    if len(runs) > 1:
        return [], [
            f"multiple logical runs match name {name!r}; explicit attempt_ids required"
        ]
    if len(runs) == 1:
        rows = connection.execute(
            "SELECT attempt_id FROM run_attempts WHERE logical_run_id = ? "
            "ORDER BY attempt_id",
            (runs[0]["logical_run_id"],),
        ).fetchall()
        return [row["attempt_id"] for row in rows], []
    reasons.append(f"logical run not in registry: {name}")
    return [], reasons


def _folds_of_attempt(
    connection: Any, attempt_id: str, expected_folds: list[int]
) -> tuple[list[int], list[str]]:
    rows = connection.execute(
        "SELECT fold FROM folds WHERE attempt_id = ? ORDER BY fold", (attempt_id,)
    ).fetchall()
    actual = [row["fold"] for row in rows]
    reasons: list[str] = []
    if expected_folds:
        missing = [fold for fold in expected_folds if fold not in actual]
        if missing:
            reasons.append(
                f"expected folds {missing} missing for attempt {attempt_id} (have {actual})"
            )
    return actual, reasons


def local_evidence_checks(
    connection: Any,
    attempt_id: str,
    fold: int,
    evaluation_ids: list[str],
    *,
    source_type: str | None = None,
) -> dict[str, Any]:
    payload = show_attempt(connection, attempt_id, fold)
    run_dir = Path(payload["folds"][0]["run_dir"]) if payload["folds"] else None
    checks: dict[str, Any] = {"run_dir": None}
    if run_dir is None or not run_dir.is_dir():
        checks["run_dir"] = f"missing run dir: {run_dir}"
        return checks
    checks["run_dir"] = str(run_dir)
    checks["run_config"] = "ok" if (run_dir / "run_config.yaml").is_file() else "missing run_config.yaml"
    if source_type != "hidden_classifier":
        checks["best_model"] = "ok" if (run_dir / "best_model").is_dir() else "missing best_model dir"
    else:
        # Post-hoc hidden-classifier attempts have no best_model of their own;
        # the parent checkpoint is referenced and hashed in provenance.
        checks["best_model"] = "ok"
    artifacts = {
        artifact["artifact_id"]: artifact for artifact in payload["artifacts"]
    }
    for evaluation in payload["evaluations"]:
        if evaluation["evaluation"]["evaluation_id"] not in evaluation_ids:
            continue
        for key in ("metrics_artifact_id", "predictions_artifact_id"):
            artifact_id = evaluation["evaluation"].get(key)
            if not artifact_id:
                checks[key] = "not recorded"
                continue
            artifact = artifacts.get(artifact_id)
            if artifact is None:
                checks[key] = "artifact row missing"
                continue
            path = run_dir / artifact["path"]
            if not path.is_file():
                checks[key] = f"missing on disk: {artifact['path']}"
                continue
            try:
                actual = sha256_file(path)
            except OSError as error:
                checks[key] = f"unreadable: {artifact['path']}: {error}"
                continue
            if artifact.get("sha256") is not None and actual != artifact["sha256"]:
                checks[key] = f"sha256 mismatch after import: {artifact['path']}"
                continue
            checks[key] = "ok"
    return checks


def _unit_evidence_checks(
    connection: Any,
    attempt_id: str,
    fold: int,
    required: list[dict[str, Any]],
    *,
    source_type: str | None = None,
) -> tuple[dict[str, Any], list[str], list[str], bool]:
    reasons: list[str] = []
    payload = show_attempt(connection, attempt_id, fold)
    evaluations = [item["evaluation"] for item in payload["evaluations"]]
    matched_ids: list[str] = []
    if not required:
        reasons.append("no required evaluation qualifiers specified")
    for spec in required:
        matches = [
            evaluation["evaluation_id"]
            for evaluation in evaluations
            if _evaluation_matches(evaluation, spec)
        ]
        if not matches:
            reasons.append(
                f"no evaluation matching required qualifiers {json.dumps(spec, sort_keys=True)}"
            )
        else:
            matched_ids.extend(matches)
    matched_ids = sorted(set(matched_ids))
    reportable_ids = [
        evaluation["evaluation_id"]
        for evaluation in evaluations
        if evaluation["evaluation_id"] in matched_ids and evaluation.get("reportable") == 1
    ]
    if matched_ids and len(reportable_ids) != len(matched_ids):
        reasons.append(
            "matched evaluation(s) not reportable: "
            + ", ".join(sorted(set(matched_ids) - set(reportable_ids)))
        )
    evidence = local_evidence_checks(
        connection, attempt_id, fold, matched_ids, source_type=source_type
    )
    evidence_problems = [
        f"{key}: {value}" for key, value in evidence.items() if key != "run_dir" and value != "ok"
    ]
    if evidence_problems:
        reasons.extend(f"local evidence: {problem}" for problem in evidence_problems)
    reportable = bool(matched_ids) and len(reportable_ids) == len(matched_ids)
    return evidence, reasons, matched_ids, reportable


def _unit_skeleton(
    connection: Any, attempt_id: str, fold: int, source_type: str
) -> dict[str, Any]:
    payload = show_attempt(connection, attempt_id, fold)
    logical = payload["logical_run"] or {}
    return {
        "selection_ids": [],
        "provenance_keys": [],
        "attempt_id": attempt_id,
        "fold": fold,
        "logical_run_name": logical.get("logical_run_name"),
        "wandb_run_id": wandb_run_id_for(connection, attempt_id, fold),
        "source_type": source_type,
        "evaluation_ids": [],
        "tags": [
            "workbook-selected",
            "modern" if not str(attempt_id).startswith("legacy-") else "legacy",
            f"dataset:{logical.get('dataset')}" if logical.get("dataset") else None,
            f"modality:{logical.get('modality')}" if logical.get("modality") else None,
            f"source-type:{source_type}",
        ],
        "group": f"family:{source_type}",
        "local_evidence": {},
        "lifecycle_state": payload["attempt"].get("current_state"),
        "reportable": False,
        "registry_evidence_sha256": registry_evidence_hash(connection, attempt_id, fold),
        "export_decision": "blocked",
        "blocking_reasons": [],
        "contributions": [],
    }


def resolve_manifest(
    selection: dict[str, Any],
    inventory: dict[str, Any],
    *,
    db_path: str | Path,
    selection_path: str | Path | None = None,
) -> dict[str, Any]:
    workbook_keys = {
        row["provenance_key"]
        for row in inventory["rows"]
        if row["provenance_key"] is not None
    }
    entry_keys = {entry_key_of(entry) for entry in selection["entries"]}
    connection = connect(db_path)
    try:
        entries_out: list[dict[str, Any]] = []
        units: dict[tuple[str, int], dict[str, Any]] = {}
        unresolved: list[dict[str, Any]] = []
        for row in inventory["rows"]:
            if row["provenance_key"] is not None and row["provenance_key"] not in entry_keys:
                unresolved.append(
                    {
                        "row_number": row["row_number"],
                        "provenance_key": row["provenance_key"],
                        "reason": "no reviewed selection entry for this provenance key",
                    }
                )
        for entry in selection["entries"]:
            key = entry_key_of(entry)
            status: dict[str, Any] = {
                "selection_id": entry["selection_id"],
                "provenance_key": provenance_key_dict(key),
                "source_type": entry["source_type"],
                "wandb_policy": entry["wandb_policy"],
                "status": entry["wandb_policy"],
                "reasons": list(entry.get("blocking_reasons") or []),
                "unit_run_ids": [],
            }
            if key not in workbook_keys:
                status["status"] = "stale_not_in_workbook"
                status["reasons"].append("selection entry not present in current workbook")
                entries_out.append(status)
                continue
            if entry["wandb_policy"] != "sync":
                entries_out.append(status)
                continue
            attempt_ids, resolve_reasons = _attempts_for_entry(connection, entry)
            status["reasons"].extend(resolve_reasons)
            if not attempt_ids:
                status["status"] = "blocked"
                entries_out.append(status)
                continue
            structural_ok = not resolve_reasons
            folds_seen: set[int] = set()
            expected_folds = [int(value) for value in entry.get("expected_folds") or []]
            for attempt_id in attempt_ids:
                folds, missing_folds = _folds_of_attempt(connection, attempt_id, [])
                fold_reasons = list(missing_folds)
                if not folds:
                    fold_reasons.append(f"no folds in registry for attempt {attempt_id}")
                    structural_ok = False
                for fold in folds:
                    folds_seen.add(fold)
                    unit_key = (attempt_id, fold)
                    if unit_key not in units:
                        units[unit_key] = _unit_skeleton(
                            connection, attempt_id, fold, entry["source_type"]
                        )
                    unit = units[unit_key]
                    evidence, evidence_reasons, matched_ids, reportable = _unit_evidence_checks(
                        connection, attempt_id, fold, entry.get("required_evaluations", []),
                        source_type=entry["source_type"],
                    )
                    unit["local_evidence"] = evidence
                    unit["evaluation_ids"] = sorted(set(unit["evaluation_ids"]) | set(matched_ids))
                    unit["reportable"] = unit["reportable"] or reportable
                    unit["contributions"].append(
                        {
                            "selection_id": entry["selection_id"],
                            "provenance_key": key,
                            "group": entry.get("group"),
                            "structural_ok": structural_ok,
                            "fold_reasons": fold_reasons,
                            "evidence_reasons": evidence_reasons,
                        }
                    )
                    status["unit_run_ids"].append(unit["wandb_run_id"])
            if expected_folds:
                missing = sorted(set(expected_folds) - folds_seen)
                if missing:
                    structural_ok = False
                    status["reasons"].append(
                        f"expected folds {missing} missing across contributing attempts "
                        f"(have {sorted(folds_seen)})"
                    )
            status["status"] = "resolved" if structural_ok and folds_seen else "blocked"
            entries_out.append(status)
        units_out: list[dict[str, Any]] = []
        for unit_key, unit in units.items():
            unit = dict(unit)
            unit["selection_ids"] = sorted(
                {contribution["selection_id"] for contribution in unit["contributions"]}
            )
            unit["provenance_keys"] = sorted(
                {contribution["provenance_key"] for contribution in unit["contributions"]}
            )
            unit["group"] = sorted(
                {
                    contribution["group"]
                    for contribution in unit["contributions"]
                    if contribution.get("group")
                }
            )[0] if any(
                contribution.get("group") for contribution in unit["contributions"]
            ) else unit["group"]
            unit["tags"] = [tag for tag in unit["tags"] if tag is not None]
            blocking_reasons: list[str] = []
            all_ok = True
            for contribution in unit["contributions"]:
                contribution_reasons = [
                    *contribution["fold_reasons"],
                    *contribution["evidence_reasons"],
                ]
                if not contribution["structural_ok"] or contribution_reasons:
                    all_ok = False
                    blocking_reasons.extend(contribution_reasons)
            unit["export_decision"] = (
                "sync" if all_ok and unit["reportable"] and unit["evaluation_ids"] else "blocked"
            )
            unit["blocking_reasons"] = sorted(set(blocking_reasons))
            unit.pop("contributions")
            units_out.append(unit)
        sync_units = [unit for unit in units_out if unit["export_decision"] == "sync"]
        blocked_units = [unit for unit in units_out if unit["export_decision"] != "sync"]
        return {
            "schema_version": MANIFEST_SCHEMA_VERSION,
            "selection": {
                "path": str(selection_path) if selection_path else None,
                "sha256": selection_sha256(selection),
                "schema_version": SELECTION_SCHEMA_VERSION,
            },
            "workbook": {
                "path": inventory["workbook"]["path"],
                "sha256": inventory["workbook"]["sha256"],
                "builder_path": inventory["workbook"]["builder_path"],
                "builder_sha256": inventory["workbook"]["builder_sha256"],
                "provenance_sheet": inventory["workbook"]["provenance_sheet"],
                "provenance_data_rows": inventory["workbook"]["provenance_data_rows"],
            },
            "db_path": str(db_path),
            "entries": entries_out,
            "export_units": sync_units,
            "blocked_units": blocked_units,
            "unresolved_entries": unresolved,
            "stale_entries": [
                entry for entry in entries_out if entry["status"] == "stale_not_in_workbook"
            ],
            "summary": {
                "selection_entries": len(selection["entries"]),
                "workbook_rows": len(workbook_keys),
                "unresolved_rows": len(unresolved),
                "stale_entries": sum(1 for entry in entries_out if entry["status"] == "stale_not_in_workbook"),
                "sync_units": len(sync_units),
                "blocked_units": len(blocked_units),
                "deduplicated_units": sum(
                    len(unit["selection_ids"]) - 1 for unit in units_out
                ),
            },
        }
    finally:
        connection.close()


def entry_key_of(entry: dict[str, Any]) -> str:
    pkey = entry["provenance_key"]
    return "|".join(str(pkey[name]) for name in KEY_COLUMNS)
