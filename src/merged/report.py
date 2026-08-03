from __future__ import annotations

import argparse
import csv
import json
import math
import os
import shutil
import statistics
import subprocess
import sys
from datetime import datetime, timezone
from collections import defaultdict
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.merged.protocol import DATASETS, METHODS
from src.merged.runtime import load_merged_config
from src.metrics import binary_auroc, classification_metrics
from src.utils import ensure_dir, read_json, read_jsonl, save_json


def _read_prediction_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix == ".jsonl":
        return read_jsonl(path)
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _prediction_value(row: dict[str, Any]) -> int:
    for field in ("prediction", "predicted_class", "teacher_forced_prediction", "parsed_prediction"):
        value = row.get(field)
        if value not in (None, "", "INVALID"):
            try:
                parsed = int(value)
                if parsed in (0, 1):
                    return parsed
            except (TypeError, ValueError):
                pass
    return 1 - int(row["label"])


def _probability_value(row: dict[str, Any], prediction: int) -> float:
    for field in ("probability", "dep_score", "score_margin", "teacher_forced_margin"):
        value = row.get(field)
        if value not in (None, ""):
            try:
                number = float(value)
                if field in {"score_margin", "teacher_forced_margin", "dep_score"}:
                    return 1.0 / (1.0 + math.exp(-number))
                return number
            except (TypeError, ValueError, OverflowError):
                pass
    return float(prediction)


def _metrics_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    y_true = [int(row["label"]) for row in rows]
    y_pred = [_prediction_value(row) for row in rows]
    scores = [_probability_value(row, pred) for row, pred in zip(rows, y_pred)]
    metrics = classification_metrics(y_true, y_pred)
    tn, fp = metrics["confusion_matrix"][0]
    fn, _ = metrics["confusion_matrix"][1]
    neg_precision = tn / (tn + fn) if tn + fn else 0.0
    neg_recall = tn / (tn + fp) if tn + fp else 0.0
    metrics.update(
        {
            "negative_f1": 2 * neg_precision * neg_recall / (neg_precision + neg_recall)
            if neg_precision + neg_recall
            else 0.0,
            "auroc": binary_auroc(y_true, scores),
            "invalid_qwen_outputs": sum(
                1 for row in rows if any(row.get(field) == "INVALID" for field in ("prediction_text", "teacher_forced_prediction_text", "generation_prediction_text"))
            ),
            "class_support_negative": int(sum(value == 0 for value in y_true)),
            "class_support_positive": int(sum(value == 1 for value in y_true)),
            "sample_count": len(rows),
        }
    )
    return metrics


def _metric_row(
    *,
    dataset: str,
    modality: str,
    method: str,
    stage: str,
    fold: int | str,
    metrics: dict[str, Any],
    protocol_label: str,
    fold_coverage: int,
) -> dict[str, Any]:
    return {
        "Dataset": dataset,
        "Modality": modality,
        "Method": method,
        "Stage": stage,
        "Fold": fold,
        "Accuracy": metrics.get("accuracy", 0.0),
        "Positive F1": metrics.get("positive_f1", 0.0),
        "Precision": metrics.get("precision", 0.0),
        "Recall": metrics.get("recall", 0.0),
        "Macro-F1": metrics.get("macro_f1", 0.0),
        "Negative F1": metrics.get("negative_f1", 0.0),
        "AUROC": metrics.get("auroc", 0.0),
        "Support Negative": metrics.get("class_support_negative", metrics.get("support_negative", 0)),
        "Support Positive": metrics.get("class_support_positive", metrics.get("support_positive", 0)),
        "Confusion Matrix": json.dumps(metrics.get("confusion_matrix", [[0, 0], [0, 0]]), sort_keys=True),
        "Invalid Qwen Outputs": metrics.get("invalid_qwen_outputs", 0),
        "Fold Coverage": fold_coverage,
        "Protocol": protocol_label,
    }


def _prediction_rows_for_fold(fold_root: Path) -> dict[tuple[str, str], list[dict[str, Any]]]:
    """Read subject-level predictions without collapsing fold boundaries."""

    result: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for dataset in DATASETS:
        path = fold_root / "qwen" / dataset / "predictions_subject_level.csv"
        if path.is_file():
            result[(dataset, "qwen")] = _read_prediction_rows(path)
    for method in ("logreg", "xgb_fixed", "xgb_optuna"):
        path = fold_root / "heads" / method / "predictions_subject_level.jsonl"
        if not path.is_file():
            continue
        for item in _read_prediction_rows(path):
            dataset = str(item.get("dataset", "")).lower()
            if dataset not in DATASETS:
                raise ValueError(f"Unknown dataset in {path}: {dataset!r}")
            result.setdefault((dataset, method), []).append(item)
    return result


def _assert_unique_prediction_ids(rows: list[dict[str, Any]], *, dataset: str, method: str, fold: int | str) -> None:
    identities = [str(row.get("subject_id") or row.get("sample_id")) for row in rows]
    if any(value in {"", "None"} for value in identities):
        raise ValueError(f"Missing prediction identity for {dataset}/{method}/fold={fold}")
    if len(identities) != len(set(identities)):
        raise ValueError(f"Duplicate subject predictions for {dataset}/{method}/fold={fold}")


def collect_stage_rows(config_path: str | Path, *, run_id: str, stage: str) -> list[dict[str, Any]]:
    config = load_merged_config(config_path)
    modality = str(config["modality"])
    root = Path(config["output_dirs"]["merged_root"]) / run_id / stage
    folds = [0] if stage == "final" else list(range(5))
    expected_datasets = {"daic"} if stage == "final" else set(DATASETS)
    expected_keys = {
        (dataset, method) for dataset in expected_datasets for method in METHODS
    }
    rows: list[dict[str, Any]] = []
    for fold in folds:
        predictions = _prediction_rows_for_fold(root / f"fold_{fold}")
        observed_keys = set(predictions)
        missing_keys = sorted(expected_keys - observed_keys)
        unexpected_keys = sorted(observed_keys - expected_keys)
        if missing_keys or unexpected_keys:
            raise ValueError(
                f"Incomplete {stage} prediction coverage for {modality}/fold={fold}: "
                f"missing={missing_keys} unexpected={unexpected_keys}"
            )
        for dataset in DATASETS:
            if stage == "final" and dataset != "daic":
                continue
            for method in METHODS:
                prediction_rows = predictions.get((dataset, method))
                if not prediction_rows:
                    raise ValueError(
                        f"Empty {stage} prediction file for "
                        f"{modality}/{dataset}/{method}/fold={fold}"
                    )
                _assert_unique_prediction_ids(prediction_rows, dataset=dataset, method=method, fold=fold)
                rows.append(
                    _metric_row(
                        dataset=dataset,
                        modality=modality,
                        method=method,
                        stage=stage,
                        fold=fold,
                        metrics=_metrics_from_rows(prediction_rows),
                        protocol_label="symmetric_merged_cv" if stage == "cv" else "symmetric_merged_daic_official_test",
                        fold_coverage=1,
                    )
                )
    return rows


def collect_pooled_stage_rows(config_path: str | Path, *, run_id: str, stage: str = "cv") -> list[dict[str, Any]]:
    """Pool the five non-overlapping outer-fold prediction files by dataset."""

    if stage != "cv":
        raise ValueError("Pooled outer-fold rows are defined only for stage=cv.")
    config = load_merged_config(config_path)
    modality = str(config["modality"])
    root = Path(config["output_dirs"]["merged_root"]) / run_id / stage
    grouped: dict[tuple[str, str], list[tuple[int, list[dict[str, Any]]]]] = defaultdict(list)
    expected_keys = {(dataset, method) for dataset in DATASETS for method in METHODS}
    for fold in range(5):
        predictions = _prediction_rows_for_fold(root / f"fold_{fold}")
        for key, prediction_rows in predictions.items():
            dataset, method = key
            if not prediction_rows:
                raise ValueError(
                    f"Empty pooled CV prediction file for "
                    f"{modality}/{dataset}/{method}/fold={fold}"
                )
            _assert_unique_prediction_ids(prediction_rows, dataset=dataset, method=method, fold=fold)
            grouped[key].append((fold, prediction_rows))
    observed_keys = set(grouped)
    missing_keys = sorted(expected_keys - observed_keys)
    unexpected_keys = sorted(observed_keys - expected_keys)
    if missing_keys or unexpected_keys:
        raise ValueError(
            f"Incomplete pooled CV prediction coverage for {modality}: "
            f"missing={missing_keys} unexpected={unexpected_keys}"
        )
    rows: list[dict[str, Any]] = []
    for (dataset, method), fold_values in sorted(grouped.items()):
        observed_folds = sorted(fold for fold, _ in fold_values)
        if observed_folds != list(range(5)):
            raise ValueError(
                f"Incomplete pooled CV prediction coverage for "
                f"{modality}/{dataset}/{method} "
                f"is {observed_folds}, expected five outer folds"
            )
        combined = [item for _, values in sorted(fold_values) for item in values]
        _assert_unique_prediction_ids(combined, dataset=dataset, method=method, fold="pooled")
        rows.append(
            _metric_row(
                dataset=dataset,
                modality=modality,
                method=method,
                stage=stage,
                fold="pooled",
                metrics=_metrics_from_rows(combined),
                protocol_label="symmetric_merged_cv_pooled",
                fold_coverage=len(fold_values),
            )
        )
    return rows


def _write_rows(rows: list[dict[str, Any]], path: Path) -> None:
    ensure_dir(path.parent)
    fields = list(rows[0]) if rows else ["Dataset", "Modality", "Method"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def _aggregate_fold_rows(rows: list[dict[str, Any]], *, stage: str) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["Dataset"], row["Modality"], row["Method"])].append(row)
    result: list[dict[str, Any]] = []
    for (dataset, modality, method), values in sorted(grouped.items()):
        out = {
            "Dataset": dataset,
            "Modality": modality,
            "Method": method,
            "Stage": stage,
            "Fold": "mean±std",
            "Fold Coverage": len(values),
            "Protocol": values[0]["Protocol"],
        }
        for metric in ("Accuracy", "Positive F1", "Precision", "Recall", "Macro-F1", "Negative F1", "AUROC"):
            numbers = [float(value[metric]) for value in values]
            out[metric] = f"{statistics.mean(numbers):.6f}±{statistics.pstdev(numbers):.6f}"
        out["Support Negative"] = sum(int(value["Support Negative"]) for value in values)
        out["Support Positive"] = sum(int(value["Support Positive"]) for value in values)
        out["Confusion Matrix"] = "pooled_in_pooled_table"
        out["Invalid Qwen Outputs"] = sum(int(value["Invalid Qwen Outputs"]) for value in values)
        result.append(out)
    return result


def _aggregate_summary(rows: list[dict[str, Any]], *, stage: str) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["Modality"], row["Method"])].append(row)
    result: list[dict[str, Any]] = []
    for (modality, method), values in sorted(grouped.items()):
        by_dataset = {row["Dataset"]: float(row["Macro-F1"]) for row in values}
        if stage == "cv" and set(by_dataset) != set(DATASETS):
            raise ValueError(
                f"Pooled CV summary for {modality}/{method} does not cover all datasets: "
                f"{sorted(by_dataset)}"
            )
        result.append(
            {
                "Dataset": "five_dataset_mean",
                "Modality": modality,
                "Method": method,
                "Stage": stage,
                "Fold": "aggregate",
                "Accuracy": "",
                "Positive F1": "",
                "Precision": "",
                "Recall": "",
                "Macro-F1": statistics.mean(by_dataset.values()) if by_dataset else 0.0,
                "Negative F1": "",
                "AUROC": "",
                "Support Negative": "",
                "Support Positive": "",
                "Confusion Matrix": "",
                "Invalid Qwen Outputs": sum(int(row["Invalid Qwen Outputs"]) for row in values),
                "Fold Coverage": len(values),
                "Protocol": values[0]["Protocol"],
            }
        )
        result.append(
            {
                "Dataset": "worst_dataset",
                "Modality": modality,
                "Method": method,
                "Stage": stage,
                "Fold": "aggregate",
                "Accuracy": "",
                "Positive F1": "",
                "Precision": "",
                "Recall": "",
                "Macro-F1": min(by_dataset.values()) if by_dataset else 0.0,
                "Negative F1": "",
                "AUROC": "",
                "Support Negative": "",
                "Support Positive": "",
                "Confusion Matrix": "",
                "Invalid Qwen Outputs": sum(int(row["Invalid Qwen Outputs"]) for row in values),
                "Fold Coverage": len(values),
                "Protocol": values[0]["Protocol"],
            }
        )
    return result


def _write_symmetric_summary_sheet(
    workbook: Any,
    *,
    cv_pooled_rows: list[dict[str, Any]] | None,
    final_rows: list[dict[str, Any]],
    run_id: str | None,
) -> None:
    """Add a compact, presentation-ready summary tab for the merged run."""

    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    title = "Merged Symmetric Summary"
    if title in workbook.sheetnames:
        del workbook[title]
    index = workbook.sheetnames.index("Summary") + 1 if "Summary" in workbook.sheetnames else 0
    sheet = workbook.create_sheet(title, index=index)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A6"

    dark_fill = PatternFill("solid", fgColor="203864")
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    body_fill = PatternFill("solid", fgColor="EAF2F8")
    note_fill = PatternFill("solid", fgColor="FFF4E5")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color="1F1F1F")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for column, width in {"A": 31, "B": 18, "C": 18, "D": 18, "E": 18, "F": 4}.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["G"].width = 3

    sheet.merge_cells("A1:E1")
    sheet["A1"] = "Symmetric Merged — Macro-F1 Summary"
    sheet["A1"].font = title_font
    sheet["A1"].fill = dark_fill
    sheet["A1"].alignment = center
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells("A2:E3")
    sheet["A2"] = (
        f"Run: {run_id or 'not specified'}. Macro-F1 only; higher is better. "
        "CV values are pooled subject-level predictions across the five datasets; "
        "final values are from the protected DAIC official holdout. The dataset-level "
        "CV table below shows every dataset and modality."
    )
    sheet["A2"].font = body_font
    sheet["A2"].fill = blue_fill
    sheet["A2"].alignment = wrap_left
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20

    methods = (
        ("qwen", "Fine-tuned Qwen"),
        ("logreg", "LogReg head"),
        ("xgb_fixed", "XGBoost fixed"),
        ("xgb_optuna", "XGBoost Optuna"),
    )
    modalities = (
        ("audio_text", "Audio + Text"),
        ("audio_only", "Audio only"),
        ("text_only", "Text only"),
    )
    headers = ["Evaluation / Modality", *(label for _, label in methods)]

    final_lookup = {
        (str(row.get("Modality")), str(row.get("Method"))): row.get("Macro-F1")
        for row in final_rows
    }
    cv_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    cv_dataset_values: dict[tuple[str, str, str], float] = {}
    for row in cv_pooled_rows or []:
        if row.get("Dataset") not in DATASETS:
            continue
        value = row.get("Macro-F1")
        if value not in (None, ""):
            numeric_value = float(value)
            modality = str(row.get("Modality"))
            method = str(row.get("Method"))
            dataset = str(row.get("Dataset"))
            cv_values[(modality, method)].append(numeric_value)
            cv_dataset_values[(dataset, modality, method)] = numeric_value

    def section(row: int, label: str) -> None:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = sheet.cell(row, 1, label)
        cell.font = white_font
        cell.fill = dark_fill
        cell.alignment = left
        sheet.row_dimensions[row].height = 22

    def table(row: int, values: dict[tuple[str, str], Any]) -> None:
        for column, value in enumerate(headers, start=1):
            cell = sheet.cell(row, column, value)
            cell.font = white_font
            cell.fill = dark_fill
            cell.alignment = center
            cell.border = border
        for offset, (modality, modality_label) in enumerate(modalities, start=1):
            current = row + offset
            label_cell = sheet.cell(current, 1, modality_label)
            label_cell.font = body_font
            label_cell.fill = body_fill
            label_cell.alignment = left
            label_cell.border = border
            for column, (method, _) in enumerate(methods, start=2):
                value = values.get((modality, method), "")
                cell = sheet.cell(current, column, value if value is not None else "")
                cell.font = body_font
                cell.fill = body_fill
                cell.alignment = center
                cell.border = border
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = "0.000"
            sheet.row_dimensions[current].height = 21

    section(5, "DAIC Official Holdout — Final Macro-F1")
    table(6, final_lookup)

    section(11, "Five-Dataset CV — Pooled Mean Macro-F1")
    table(
        12,
        {
            key: statistics.mean(values)
            for key, values in cv_values.items()
            if values
        },
    )

    section(17, "Five-Dataset CV — Worst-Dataset Macro-F1")
    table(18, {key: min(values) for key, values in cv_values.items() if values})

    section(27, "Five-Dataset CV — Macro-F1 by Dataset and Modality")
    dataset_header_row = 28
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(dataset_header_row, column, value)
        cell.font = white_font
        cell.fill = dark_fill
        cell.alignment = center
        cell.border = border
    dataset_labels = {
        "daic": "DAIC",
        "cmdc": "CMDC",
        "turkish": "Turkish",
        "d3tec": "D3TEC",
        "androids_interview": "Androids Interview",
    }
    dataset_row = dataset_header_row + 1
    for dataset in DATASETS:
        for modality, modality_label in modalities:
            label_cell = sheet.cell(dataset_row, 1, f"{dataset_labels[dataset]} — {modality_label}")
            label_cell.font = body_font
            label_cell.fill = body_fill
            label_cell.alignment = left
            label_cell.border = border
            for column, (method, _) in enumerate(methods, start=2):
                value = cv_dataset_values.get((dataset, modality, method), "")
                cell = sheet.cell(dataset_row, column, value)
                cell.font = body_font
                cell.fill = body_fill
                cell.alignment = center
                cell.border = border
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = "0.000"
            sheet.row_dimensions[dataset_row].height = 21
            dataset_row += 1
    dataset_ref = f"A{dataset_header_row}:E{dataset_row - 1}"
    dataset_table = Table(displayName="MergedCVByDataset", ref=dataset_ref)
    dataset_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(dataset_table)
    sheet.auto_filter.ref = dataset_ref

    sheet.merge_cells("A45:E46")
    sheet["A45"] = (
        "Methods: fine-tuned Qwen, standardized Logistic Regression, fixed XGBoost, "
        "and grouped 150-trial Optuna XGBoost. Blank cells indicate that the corresponding "
        "compact artifact was unavailable."
    )
    sheet["A45"].font = body_font
    sheet["A45"].fill = note_fill
    sheet["A45"].alignment = wrap_left
    sheet.row_dimensions[45].height = 21
    sheet.row_dimensions[46].height = 21

    if final_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Final DAIC Macro-F1"
        chart.x_axis.title = "Macro-F1"
        chart.y_axis.title = "Modality"
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 1
        chart.height = 7
        chart.width = 14
        data = Reference(sheet, min_col=2, max_col=5, min_row=6, max_row=9)
        categories = Reference(sheet, min_col=1, min_row=7, max_row=9)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend.position = "b"
        sheet.add_chart(chart, "G5")


def update_workbook(
    workbook_path: Path,
    cv_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
    *,
    cv_pooled_rows: list[dict[str, Any]] | None = None,
    run_id: str | None = None,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = load_workbook(workbook_path) if workbook_path.is_file() else None
    if workbook is None:
        from openpyxl import Workbook

        workbook = Workbook()
        workbook.remove(workbook.active)
    for title, rows in (("Merged Symmetric CV", cv_rows), ("Merged DAIC Official", final_rows)):
        if title in workbook.sheetnames:
            del workbook[title]
        sheet = workbook.create_sheet(title)
        if not rows:
            rows = [{"Status": "No completed artifacts"}]
        headers = list(rows[0])
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        if sheet.max_row >= 2:
            ref = f"A1:{chr(64 + min(sheet.max_column, 26))}{sheet.max_row}"
            table = Table(displayName="Merged" + ("CV" if "CV" in title else "DAICOfficial"), ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
            sheet.auto_filter.ref = ref
        sheet.freeze_panes = "A2"
    _write_symmetric_summary_sheet(
        workbook,
        cv_pooled_rows=cv_pooled_rows,
        final_rows=final_rows,
        run_id=run_id,
    )
    workbook.save(workbook_path)


def validate_workbook(
    workbook_path: str | Path,
    *,
    cv_rows: list[dict[str, Any]],
    final_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Validate the two generated sheets, tables, filters, and cell values.

    The report is a tracked deliverable, so a successful save is not enough:
    verify that the workbook still contains the expected rows and that each
    dedicated sheet has a usable table/filter range.
    """

    from openpyxl import load_workbook

    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"Generated workbook is missing: {path}")
    workbook = load_workbook(path, data_only=False)
    expected = {
        "Merged Symmetric CV": cv_rows,
        "Merged DAIC Official": final_rows,
    }
    sheet_results: dict[str, Any] = {}
    for title, rows in expected.items():
        if title not in workbook.sheetnames:
            raise ValueError(f"Generated workbook is missing sheet {title!r}: {path}")
        sheet = workbook[title]
        display_rows = rows if rows else [{"Status": "No completed artifacts"}]
        headers = list(display_rows[0])
        if [sheet.cell(1, column).value for column in range(1, len(headers) + 1)] != headers:
            raise ValueError(f"Workbook headers do not match report rows on {title!r}.")
        if sheet.max_row != len(display_rows) + 1:
            raise ValueError(
                f"Workbook row count mismatch on {title!r}: "
                f"found={sheet.max_row - 1} expected={len(display_rows)}"
            )
        for row_number, row in enumerate(display_rows, start=2):
            for column, header in enumerate(headers, start=1):
                actual = sheet.cell(row_number, column).value
                expected_value = row.get(header, "")
                blank_equivalent = actual in (None, "") and expected_value in (None, "")
                numeric_equivalent = (
                    isinstance(actual, (int, float))
                    and not isinstance(actual, bool)
                    and isinstance(expected_value, (int, float))
                    and not isinstance(expected_value, bool)
                    and math.isclose(
                        float(actual), float(expected_value), rel_tol=1e-12, abs_tol=1e-12
                    )
                )
                if not blank_equivalent and not numeric_equivalent and actual != expected_value:
                    raise ValueError(
                        f"Workbook value mismatch on {title!r}, row={row_number}, "
                        f"column={header!r}."
                    )
        if not sheet.tables:
            raise ValueError(f"Workbook sheet {title!r} has no table.")
        table_refs = {table.ref for table in sheet.tables.values()}
        if not sheet.auto_filter.ref or sheet.auto_filter.ref not in table_refs:
            raise ValueError(f"Workbook sheet {title!r} has an invalid table/filter range.")
        sheet_results[title] = {
            "rows": len(display_rows),
            "headers": headers,
            "table_refs": sorted(table_refs),
            "auto_filter": sheet.auto_filter.ref,
        }
    summary_title = "Merged Symmetric Summary"
    if summary_title not in workbook.sheetnames:
        raise ValueError(f"Generated workbook is missing sheet {summary_title!r}: {path}")
    summary = workbook[summary_title]
    if summary["A1"].value != "Symmetric Merged — Macro-F1 Summary":
        raise ValueError(f"Generated workbook has an invalid {summary_title!r} title.")
    expected_summary_headers = [
        "Evaluation / Modality",
        "Fine-tuned Qwen",
        "LogReg head",
        "XGBoost fixed",
        "XGBoost Optuna",
    ]
    for header_row in (6, 12, 18):
        actual_headers = [summary.cell(header_row, column).value for column in range(1, 6)]
        if actual_headers != expected_summary_headers:
            raise ValueError(
                f"Generated workbook has invalid summary headers on row {header_row}: {path}"
            )
    dataset_header_row = 28
    actual_dataset_headers = [summary.cell(dataset_header_row, column).value for column in range(1, 6)]
    if actual_dataset_headers != expected_summary_headers:
        raise ValueError(f"Generated workbook has invalid dataset summary headers: {path}")
    if summary.max_row < 43 or "MergedCVByDataset" not in summary.tables:
        raise ValueError(f"Generated workbook is missing the dataset-level CV summary table: {path}")
    if not summary._charts:
        raise ValueError(f"Generated workbook summary has no comparison chart: {path}")
    sheet_results[summary_title] = {
        "rows": 43,
        "headers": expected_summary_headers,
        "charts": len(summary._charts),
        "dataset_table": "A28:E43",
    }
    return {"status": "passed", "workbook": str(path), "sheets": sheet_results}


def _execution_metadata(config_paths: list[str | Path], *, run_id: str, source_commit: str) -> dict[str, Any]:
    configs = [load_merged_config(path) for path in config_paths]
    registry_path = Path(configs[0]["output_dirs"]["merged_root"]).parents[1] / "symmetric_merged_jobs" / f"{run_id}.json"
    registry = read_json(registry_path) if registry_path.is_file() else None
    jobs = (registry or {}).get("jobs", [])
    audits: dict[str, dict[str, Any]] = {}
    for config in configs:
        modality = str(config["modality"])
        audits[modality] = {}
        for stage in ("smoke", "cv", "final"):
            path = Path(config["output_dirs"]["merged_root"]) / run_id / stage / "acceptance_audit.json"
            if path.is_file():
                audits[modality][stage] = read_json(path).get("status", "unknown")
    job_accounting = [
        {
            "job_id": str(job.get("job_id")),
            "job_key": job.get("job_key"),
            "stage": job.get("stage"),
            "modality": job.get("modality"),
            "fold": job.get("fold"),
            "kind": job.get("kind"),
            "state": job.get("observed_state", job.get("state")),
            "exit_code": job.get("exit_code", ""),
            "elapsed": job.get("elapsed", ""),
            "max_rss": job.get("max_rss", ""),
            "allocated_cpus": job.get("allocated_cpus", ""),
            "allocated_tres": job.get("allocated_tres", ""),
            "node_list": job.get("node_list", ""),
        }
        for job in jobs
        if job.get("job_id")
    ]
    return {
        "schema_version": "symmetric_merged_execution_metadata.v1",
        "run_id": run_id,
        "source_commit": source_commit,
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "configs": [str(path) for path in config_paths],
        "job_registry": str(registry_path) if registry is not None else None,
        "registry_status": (registry or {}).get("registry_status"),
        "job_ids": sorted(str(job["job_id"]) for job in jobs if job.get("job_id")),
        "job_count": len(jobs),
        "job_states": {
            state: sum(1 for job in jobs if str(job.get("observed_state", job.get("state", ""))).upper() == state)
            for state in sorted({str(job.get("observed_state", job.get("state", ""))).upper() for job in jobs})
        },
        "job_accounting": job_accounting,
        "runtime_storage": {
            "accounted_job_count": sum(
                bool(row.get("elapsed") or row.get("max_rss") or row.get("allocated_tres"))
                for row in job_accounting
            ),
            "job_accounting_fields": [
                "elapsed",
                "max_rss",
                "allocated_cpus",
                "allocated_tres",
                "node_list",
            ],
            "project_disk_path": str(PROJECT_ROOT),
            "project_disk_free_bytes": shutil.disk_usage(PROJECT_ROOT).free,
            "project_disk_used_bytes": shutil.disk_usage(PROJECT_ROOT).used,
        },
        "acceptance_audits": audits,
        "project_disk": {
            "path": str(PROJECT_ROOT),
            "free_bytes_at_report": shutil.disk_usage(PROJECT_ROOT).free,
            "used_bytes_at_report": shutil.disk_usage(PROJECT_ROOT).used,
        },
        "limitations": [
            "Best-model checkpoints, hidden feature arrays, classifier joblibs, and Optuna SQLite databases remain on MN5 GPFS.",
            "Reported CV headline metrics are pooled from non-overlapping subject-level outer-fold predictions by dataset.",
        ],
    }


def generate_reports(
    config_paths: list[str | Path], *, run_id: str, output_dir: str | Path, workbook_path: str | Path | None = None
) -> dict[str, Any]:
    output = ensure_dir(output_dir)
    cv_rows: list[dict[str, Any]] = []
    cv_pooled_rows: list[dict[str, Any]] = []
    final_rows: list[dict[str, Any]] = []
    for config_path in config_paths:
        cv_rows.extend(collect_stage_rows(config_path, run_id=run_id, stage="cv"))
        cv_pooled_rows.extend(collect_pooled_stage_rows(config_path, run_id=run_id, stage="cv"))
        final_rows.extend(collect_stage_rows(config_path, run_id=run_id, stage="final"))
    cv_fold_path = output / "symmetric_merged_cv_fold_level.csv"
    cv_summary_path = output / "symmetric_merged_cv_fold_mean_std.csv"
    cv_pooled_path = output / "symmetric_merged_cv_pooled.csv"
    cv_aggregate_path = output / "symmetric_merged_cv_aggregate.csv"
    final_path = output / "symmetric_merged_daic_official_test.csv"
    _write_rows(cv_rows, cv_fold_path)
    _write_rows(_aggregate_fold_rows(cv_rows, stage="cv"), cv_summary_path)
    _write_rows(cv_pooled_rows, cv_pooled_path)
    cv_aggregate_rows = cv_pooled_rows + _aggregate_summary(cv_pooled_rows, stage="cv")
    _write_rows(cv_aggregate_rows, cv_aggregate_path)
    _write_rows(final_rows, final_path)
    workbook = Path(workbook_path) if workbook_path else Path("depression_results_combined_with_posf1_graphs.xlsx")
    workbook_validation: dict[str, Any] | None = None
    if cv_rows or final_rows:
        workbook_cv_rows = cv_rows + _aggregate_fold_rows(cv_rows, stage="cv") + cv_aggregate_rows
        update_workbook(
            workbook,
            workbook_cv_rows,
            final_rows,
            cv_pooled_rows=cv_pooled_rows,
            run_id=run_id,
        )
        workbook_validation = validate_workbook(
            workbook,
            cv_rows=workbook_cv_rows,
            final_rows=final_rows,
        )
    report_path = output / "symmetric_merged_execution_results.md"
    commit = os.environ.get("SYMMETRIC_MERGED_SOURCE_COMMIT")
    if not commit:
        commit = subprocess.check_output(["git", "rev-parse", "HEAD"], text=True).strip()
    execution_metadata = _execution_metadata(config_paths, run_id=run_id, source_commit=commit)
    execution_metadata_path = output / "symmetric_merged_execution_metadata.json"
    save_json(execution_metadata, execution_metadata_path)
    audit_lines = [
        f"- `{modality}`: " + ", ".join(f"{stage}={status}" for stage, status in stages.items())
        for modality, stages in execution_metadata["acceptance_audits"].items()
    ]
    report_path.write_text(
        "# Symmetric merged execution/results\n\n"
        f"- Run ID: `{run_id}`\n- Git commit: `{commit}`\n"
        f"- CV fold rows: {len(cv_rows)}\n- CV pooled rows: {len(cv_pooled_rows)}\n"
        f"- DAIC official-test rows: {len(final_rows)}\n"
        f"- CV pooled CSV: `{cv_pooled_path}`\n"
        f"- Job registry: `{execution_metadata.get('job_registry')}`\n"
        f"- Job IDs: {', '.join(execution_metadata['job_ids']) or 'none'}\n"
        f"- Registry status: `{execution_metadata.get('registry_status')}`\n\n"
        f"- Slurm accounting rows: {len(execution_metadata['job_accounting'])}; "
        f"accounted runtime/resource rows: {execution_metadata['runtime_storage']['accounted_job_count']}\n"
        f"- Runtime/storage metadata: `{execution_metadata_path}`\n\n"
        "## Acceptance audits\n\n"
        + ("\n".join(audit_lines) if audit_lines else "- No acceptance audits found.")
        + "\n\n## Protocol\n\n"
        "Five datasets, three modalities, Qwen + standardized Logistic Regression + fixed XGBoost + 150-trial grouped Optuna XGBoost.\n\n"
        "## Limitations\n\n"
        + "\n".join(f"- {item}" for item in execution_metadata["limitations"])
        + "\n",
        encoding="utf-8",
    )
    return {
        "cv_rows": len(cv_rows),
        "cv_pooled_rows": len(cv_pooled_rows),
        "final_rows": len(final_rows),
        "csv_paths": [
            str(cv_fold_path),
            str(cv_summary_path),
            str(cv_pooled_path),
            str(cv_aggregate_path),
            str(final_path),
        ],
        "workbook": str(workbook),
        "workbook_validation": workbook_validation,
        "report": str(report_path),
        "execution_metadata": str(execution_metadata_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build compact symmetric merged CSV/report/workbook artifacts.")
    parser.add_argument("--config", action="append", required=True, type=Path)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--workbook", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = generate_reports(args.config, run_id=args.run_id, output_dir=args.output_dir, workbook_path=args.workbook)
    print(json.dumps(result, indent=2), flush=True)


if __name__ == "__main__":
    main()
