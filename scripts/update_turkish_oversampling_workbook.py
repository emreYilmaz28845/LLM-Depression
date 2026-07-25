#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import csv
import json
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.chart import Reference
from openpyxl.styles import Alignment, Font, PatternFill


MODALITY_LABELS = {
    "audio_text": "Audio + Text",
    "audio_only": "Audio only",
    "text_only": "Text only",
}
SEED_ORDER = {1337: 0, 2024: 1, 7: 2}


def _copy_style(source, target) -> None:
    target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy.copy(source.alignment)


def _read_aurocs(path: Path) -> dict[tuple[str, str], float]:
    with path.open(encoding="utf-8", newline="") as handle:
        rows = csv.DictReader(handle)
        return {
            (row["modality"], row["classifier_variant"]): float(row["auroc"])
            for row in rows
            if row["dataset"] == "turkish"
        }


def _fold_selected_macro(
    matrix: dict, project_root: Path
) -> dict[tuple[str, int], float]:
    values = {}
    for job in matrix["jobs"]:
        config = yaml.safe_load(
            (project_root / job["config"]).read_text(encoding="utf-8")
        )
        run_root = Path(
            str(config["output_dirs"]["run_root"]).replace(
                "${PROJECT_ROOT}", str(project_root.resolve())
            )
        )
        metrics_path = (
            run_root
            / job["run_name"]
            / f"fold_{job['fold']}"
            / "eval/best_validation/metrics_original_teacher_forced.json"
        )
        metrics = json.loads(metrics_path.read_text(encoding="utf-8"))
        values[(job["profile"], int(job["fold"]))] = float(metrics["macro_f1"])
    return values


def _style_sheet(sheet) -> None:
    dark = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    white_bold = Font(color="FFFFFF", bold=True)
    for row in (4, 20, 29):
        for cell in sheet[row]:
            if cell.value is not None:
                cell.fill = dark
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in (18, 27):
        for cell in sheet[row]:
            if cell.value is not None:
                cell.fill = light
                cell.font = Font(bold=True)
    sheet.freeze_panes = "A5"
    widths = {
        "A": 18,
        "B": 22,
        "C": 13,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 13,
        "H": 15,
        "I": 15,
        "J": 16,
        "K": 12,
        "L": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=5, max_row=40, min_col=4, max_col=11):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"


def update_workbook(
    workbook_path: Path,
    detailed_csv: Path,
    stage3_summary_path: Path,
    stage4_matrix_path: Path,
    stage4_summary_path: Path,
    project_root: Path,
) -> None:
    stage3 = json.loads(stage3_summary_path.read_text(encoding="utf-8"))
    stage4 = json.loads(stage4_summary_path.read_text(encoding="utf-8"))
    matrix = yaml.safe_load(stage4_matrix_path.read_text(encoding="utf-8"))
    aurocs = _read_aurocs(detailed_csv)
    fold_macro = _fold_selected_macro(matrix, project_root)

    workbook = load_workbook(workbook_path)
    summary = workbook["Summary"]
    if "A2:E2" in {str(item) for item in summary.merged_cells.ranges}:
        summary.unmerge_cells("A2:E2")
    summary.merge_cells("A2:F2")
    summary["A2"] = (
        "Macro-F1 only; higher is better. Paper scores from different evaluation "
        "sets are excluded. Blank cells mean that a comparable classifier result "
        "is unavailable. Turkish hidden-state heads use BDI>=17 table-aligned "
        "outer validation. Subject-OS values are three-seed means; the matched "
        "Qwen pilot did not clear the Stage-5 gate."
    )
    summary["F5"] = "XGBoost Subject OS\n(3-seed mean)"
    for row in range(5, 15):
        _copy_style(summary.cell(row=row, column=5), summary.cell(row=row, column=6))
    mean_by_modality = {
        item["condition"]: item["mean_sampled_macro_f1"]
        for item in stage3["decisions"]
    }
    summary["F12"] = mean_by_modality["audio_text"]
    summary["F13"] = mean_by_modality["audio_only"]
    summary["F14"] = mean_by_modality["text_only"]
    summary.column_dimensions["F"].width = 24
    chart = summary._charts[0]
    os_series = [
        index
        for index, series in enumerate(chart.ser)
        if series.tx
        and series.tx.strRef
        and series.tx.strRef.f.replace("$", "") == "'Summary'!F5"
    ]
    for index in reversed(os_series[1:]):
        del chart.ser[index]
    if not os_series:
        chart.add_data(
            Reference(summary, min_col=6, min_row=5, max_row=14),
            titles_from_data=True,
        )

    if "Turkish Oversampling" in workbook.sheetnames:
        del workbook["Turkish Oversampling"]
    sheet = workbook.create_sheet("Turkish Oversampling", 4)
    sheet.merge_cells("A1:L1")
    sheet["A1"] = "Turkish BDI>=17 Subject Oversampling"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:L2")
    sheet["A2"] = (
        "Five-fold hidden-state results use table-aligned outer validation. "
        "Qwen results are the predeclared folds 0-1 pilot, not a full-panel result."
    )
    sheet["A2"].alignment = Alignment(wrap_text=True)

    stage3_headers = (
        "Modality",
        "Profile",
        "OS seed",
        "Accuracy",
        "Positive F1",
        "Macro F1",
        "Negative F1",
        "Negative recall",
        "Positive recall",
        "Balanced accuracy",
        "AUROC",
        "Confusion matrix",
    )
    for column, value in enumerate(stage3_headers, 1):
        sheet.cell(row=4, column=column, value=value)
    results = sorted(
        stage3["results"],
        key=lambda item: (
            list(MODALITY_LABELS).index(item["condition"]),
            item["sampling_mode"] != "none",
            SEED_ORDER.get(item["oversampling_seed"], -1),
        ),
    )
    for row_index, item in enumerate(results, 5):
        variant = item["experiment_id"]
        values = (
            MODALITY_LABELS[item["condition"]],
            "Control" if item["sampling_mode"] == "none" else "Subject OS 0.75",
            item["oversampling_seed"] if item["sampling_mode"] != "none" else None,
            item["accuracy"],
            item["positive_f1"],
            item["macro_f1"],
            item["negative_f1"],
            item["negative_recall"],
            item["positive_recall"],
            item["balanced_accuracy"],
            aurocs[(item["condition"], variant)],
            json.dumps(item["confusion_matrix"]),
        )
        for column, value in enumerate(values, 1):
            sheet.cell(row=row_index, column=column, value=value)

    sheet.merge_cells("A18:L18")
    sheet["A18"] = "Stage 3 gate"
    gate_headers = (
        "Modality",
        "Control Macro F1",
        "OS mean Macro F1",
        "Mean gain",
        "Seeds beating",
        "Mean negative-recall gain",
        "Positive-recall loss",
        "Qualifies",
    )
    for column, value in enumerate(gate_headers, 1):
        sheet.cell(row=20, column=column, value=value)
    for row_index, item in enumerate(stage3["decisions"], 21):
        values = (
            MODALITY_LABELS[item["condition"]],
            item["control_macro_f1"],
            item["mean_sampled_macro_f1"],
            item["mean_macro_f1_gain"],
            item["seeds_beating_control"],
            item["mean_negative_recall_gain"],
            item["mean_positive_recall_loss"],
            "PASS" if item["qualifies_modality_gate"] else "FAIL",
        )
        for column, value in enumerate(values, 1):
            sheet.cell(row=row_index, column=column, value=value)

    sheet.merge_cells("A27:L27")
    sheet["A27"] = "Stage 4 matched Qwen pilot"
    pilot_headers = (
        "Profile",
        "Fold 0 selected Macro F1",
        "Fold 1 selected Macro F1",
        "Fold mean",
        "Pooled Macro F1",
        "Pooled negative recall",
        "Pooled positive recall",
        "Stage 5",
    )
    for column, value in enumerate(pilot_headers, 1):
        sheet.cell(row=29, column=column, value=value)
    for row_index, profile in enumerate(("weighted", "oversampled"), 30):
        pooled = stage4[
            "pooled_control" if profile == "weighted" else "pooled_oversampled"
        ]
        fold_values = [fold_macro[(profile, fold)] for fold in (0, 1)]
        values = (
            "Weighted control" if profile == "weighted" else "Subject OS 0.75",
            fold_values[0],
            fold_values[1],
            sum(fold_values) / len(fold_values),
            pooled["macro_f1"],
            pooled["negative_recall"],
            pooled["positive_recall"],
            "STOP" if not stage4["proceed_to_full"] else "PROCEED",
        )
        for column, value in enumerate(values, 1):
            sheet.cell(row=row_index, column=column, value=value)
    sheet["A33"] = "Mean selected-fold gain"
    sheet["B33"] = stage4["mean_selected_validation_macro_f1_gain"]
    sheet["A34"] = "Minimum selected-fold gain"
    sheet["B34"] = stage4["minimum_fold_selected_validation_macro_f1_gain"]
    sheet["A35"] = "Pooled Macro F1 non-inferiority"
    sheet["B35"] = stage4["pooled_macro_f1_not_below_control"]
    sheet["A36"] = "Pooled negative-recall non-inferiority"
    sheet["B36"] = stage4["pooled_negative_recall_not_below_control"]
    sheet["A38"] = "Execution"
    sheet["B38"] = (
        "86 Slurm jobs accounted: 85 completed 0:0; job 43809957 is the "
        "documented smoke configuration failure, corrected by 43810107."
    )
    sheet["A39"] = "Acceptance"
    sheet["B39"] = (
        "Stages 2-4 audits pass; zero unexpected log signatures; Stage 5 not "
        "submitted because the predeclared pilot gate failed."
    )
    sheet.merge_cells("B38:L38")
    sheet.merge_cells("B39:L39")
    sheet["B38"].alignment = Alignment(wrap_text=True)
    sheet["B39"].alignment = Alignment(wrap_text=True)
    _style_sheet(sheet)

    workbook.save(workbook_path)
    verified = load_workbook(workbook_path, read_only=False)
    if "Turkish Oversampling" not in verified.sheetnames:
        raise RuntimeError("Turkish Oversampling sheet was not saved")
    if len(verified["Summary"]._charts) != 1:
        raise RuntimeError("Macro-F1 Summary chart was not preserved")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--detailed-csv", type=Path, required=True)
    parser.add_argument("--stage3-summary", type=Path, required=True)
    parser.add_argument("--stage4-matrix", type=Path, required=True)
    parser.add_argument("--stage4-summary", type=Path, required=True)
    parser.add_argument("--project-root", type=Path, default=Path("."))
    args = parser.parse_args()
    update_workbook(
        args.workbook,
        args.detailed_csv,
        args.stage3_summary,
        args.stage4_matrix,
        args.stage4_summary,
        args.project_root,
    )
    print(f"Updated {args.workbook}")


if __name__ == "__main__":
    main()
