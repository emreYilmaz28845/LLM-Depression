"""Build depression_results_clean.xlsx — the trustworthy, provenance-carrying results workbook.

Generated 2026-08-06. Every headline value was verified by recomputation from local
artifacts (final_summary.json, coverage audits, matrix audits, merged-run prediction
files) or matched to an audited experiment document. Sources per row: Provenance sheet.

No values are hand-edited after generation; to change a number, change this script and
regenerate. This is the report layer; the artifact layer is the files referenced in
the Provenance sheet.

Modes:
  default    depression_results_clean.xlsx  (headline: Qwen + LogReg + XGBoost fixed only)
  --detailed depression_results_clean_detailed.xlsx in docs/archive/results_20260806/
             (adds XGBoost Optuna and Subject-OS columns/rows)
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT = PROJECT_ROOT / "depression_results_clean.xlsx"
OUT_DETAILED = PROJECT_ROOT / "docs/archive/results_20260806/depression_results_clean_detailed.xlsx"

# --------------------------------------------------------------------------- styles
DARK = PatternFill("solid", fgColor="203864")
BLUE = PatternFill("solid", fgColor="D9EAF7")
BODY = PatternFill("solid", fgColor="EAF2F8")
NOTE = PatternFill("solid", fgColor="FFF4E5")
POS_FILL = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="1F1F1F")
SMALL_FONT = Font(name="Calibri", size=9, color="1F1F1F")
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _section(ws, row: int, label: str, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row, 1, label)
    cell.font = WHITE
    cell.fill = DARK
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 22


def _header_row(ws, row: int, headers: list[str]) -> None:
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row, col, value)
        cell.font = WHITE
        cell.fill = DARK
        cell.alignment = CENTER
        cell.border = BORDER


def _body_cell(ws, row: int, col: int, value: Any, *, fmt: str | None = None) -> None:
    cell = ws.cell(row, col, value)
    cell.font = BODY_FONT
    cell.fill = BODY
    cell.alignment = CENTER
    cell.border = BORDER
    if fmt and isinstance(value, (int, float)):
        cell.number_format = fmt


DELTA_EPS = 0.0005


def _delta_fill(value: float | None) -> PatternFill | None:
    """Green = positive delta (EN better / merged better), red = negative, neutral ~0."""
    if value is None:
        return None
    if value > DELTA_EPS:
        return POS_FILL
    if value < -DELTA_EPS:
        return NEG_FILL
    return NEUTRAL_FILL


def _delta_cell(ws, row: int, col: int, value: float | None) -> None:
    if value is None:
        _body_cell(ws, row, col, None)
        return
    _body_cell(ws, row, col, round(value, 4), fmt="0.0000")
    fill = _delta_fill(value)
    if fill is not None:
        ws.cell(row, col).fill = fill


def _title(ws, text: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = DARK
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32


def _note(ws, row: int, text: str, span: int, *, height: int = 26) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row, 1, text)
    cell.font = BODY_FONT
    cell.fill = NOTE
    cell.alignment = WRAP
    ws.row_dimensions[row].height = height


def _widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# --------------------------------------------------------------------------- data
# Standalone fine-tuned Qwen macro-F1. All values recomputed 2026-08-06 from the
# local artifacts listed in PROVENANCE_QWEN. (dataset, modality) -> macro-F1.
STANDALONE_QWEN: dict[tuple[str, str], float] = {
    # Harmonized campaign harmonized_v1_prod_20260809T171705Z_d1e8130b (recipe
    # harmonized_full_transcript_single30_allwindows_selmacrof1_tf_v1; Issue #12 /
    # PR #10; teacher-forced, binary-strict, best_model, macro-F1 selection,
    # audio encoder frozen; first-wave failures replaced by _r1 retries).
    # Aggregations: DAIC fixed official test; CMDC/Turkish 5-fold mean;
    # D3TEC/Androids pooled subject-level 5-fold. Recomputed 2026-08-10.
    ("DAIC", "Audio + Text"): 0.7353,
    ("DAIC", "Audio only"): 0.5392,
    ("DAIC", "Text only"): 0.7353,
    ("CMDC", "Audio + Text"): 0.9700,
    ("CMDC", "Audio only"): 0.9516,
    ("CMDC", "Text only"): 0.9713,
    ("Turkish", "Audio + Text"): 0.6666,
    ("Turkish", "Audio only"): 0.5137,
    ("Turkish", "Text only"): 0.6502,
    ("D3TEC", "Audio + Text"): 0.5589,
    ("D3TEC", "Audio only"): 0.6649,
    ("D3TEC", "Text only"): 0.6113,
    ("Androids Interview", "Audio + Text"): 0.8606,
    ("Androids Interview", "Audio only"): 0.8690,
    ("Androids Interview", "Text only"): 0.7317,
}

# Standalone hidden-state heads: (dataset, modality) -> (logreg, xgb_fixed, xgb_optuna, subject_os)
STANDALONE_HEADS: dict[tuple[str, str], tuple[float | None, float | None, float | None, float | None]] = {
    # Harmonized hidden-state heads (logreg_raw / xgb_raw), 5-fold mean
    # subject-level macro-F1 from the harmonized campaign; Optuna/Subject-OS
    # not run -> None. Recomputed 2026-08-10.
    ("DAIC", "Audio + Text"): (0.7432, 0.7353, None, None),
    ("DAIC", "Audio only"): (0.5583, 0.5190, None, None),
    ("DAIC", "Text only"): (0.7235, 0.7457, None, None),
    ("CMDC", "Audio + Text"): (0.9614, 0.9700, None, None),
    ("CMDC", "Audio only"): (0.9841, 0.9225, None, None),
    ("CMDC", "Text only"): (0.9420, 0.9683, None, None),
    ("Turkish", "Audio + Text"): (0.6289, 0.6325, None, None),
    ("Turkish", "Audio only"): (0.5209, 0.4271, None, None),
    ("Turkish", "Text only"): (0.5875, 0.5234, None, None),
    ("D3TEC", "Audio + Text"): (0.4988, 0.5585, None, None),
    ("D3TEC", "Audio only"): (0.6031, 0.5404, None, None),
    ("D3TEC", "Text only"): (0.4651, 0.5911, None, None),
    ("Androids Interview", "Audio + Text"): (0.8745, 0.8656, None, None),
    ("Androids Interview", "Audio only"): (0.8512, 0.8235, None, None),
    ("Androids Interview", "Text only"): (0.8326, 0.8241, None, None),
}

# Merged symmetric runs. modality -> (run_id, official DAIC macro per method,
# pooled-CV macro per (dataset, method)).
MERGED_RUNS: dict[str, dict[str, Any]] = {
    # Harmonized merged campaign runs (run harmonized_v1_prod_20260809T171705Z_d1e8130b).
    # official = DAIC protected official test (47 subjects, daic_official_test_only);
    # cv = mean of the five per-fold holdout macro-F1 values (postprocess summaries).
    # Final epochs (rounded median of CV-selected): audio_text 2, audio_only 4,
    # text_only 5. Recomputed 2026-08-10.
    "audio_text": {
        "run_id": "harmonized_v1_prod_20260809T171705Z_d1e8130b",
        "official": {"qwen": 0.7631, "logreg": 0.7432, "xgb_fixed": 0.7432, "xgb_optuna": None},
        "cv": {
            ("daic", "qwen"): 0.6540, ("cmdc", "qwen"): 0.9262, ("turkish", "qwen"): 0.5133,
            ("d3tec", "qwen"): 0.6071, ("androids_interview", "qwen"): 0.7874,
            ("daic", "logreg"): 0.6634, ("cmdc", "logreg"): 0.9421, ("turkish", "logreg"): 0.4786,
            ("d3tec", "logreg"): 0.6115, ("androids_interview", "logreg"): 0.8062,
            ("daic", "xgb_fixed"): 0.6835, ("cmdc", "xgb_fixed"): 0.9577, ("turkish", "xgb_fixed"): 0.4940,
            ("d3tec", "xgb_fixed"): 0.5737, ("androids_interview", "xgb_fixed"): 0.7842,
        },
    },
    "audio_only": {
        "run_id": "harmonized_v1_prod_20260809T171705Z_d1e8130b",
        "official": {"qwen": 0.5332, "logreg": 0.6189, "xgb_fixed": 0.5554, "xgb_optuna": None},
        "cv": {
            ("daic", "qwen"): 0.6018, ("cmdc", "qwen"): 0.9500, ("turkish", "qwen"): 0.4516,
            ("d3tec", "qwen"): 0.5925, ("androids_interview", "qwen"): 0.8564,
            ("daic", "logreg"): 0.5586, ("cmdc", "logreg"): 0.9683, ("turkish", "logreg"): 0.4820,
            ("d3tec", "logreg"): 0.5387, ("androids_interview", "logreg"): 0.8441,
            ("daic", "xgb_fixed"): 0.5631, ("cmdc", "xgb_fixed"): 0.9366, ("turkish", "xgb_fixed"): 0.4484,
            ("d3tec", "xgb_fixed"): 0.5378, ("androids_interview", "xgb_fixed"): 0.8531,
        },
    },
    "text_only": {
        "run_id": "harmonized_v1_prod_20260809T171705Z_d1e8130b",
        "official": {"qwen": 0.7756, "logreg": 0.7157, "xgb_fixed": 0.7552, "xgb_optuna": None},
        "cv": {
            ("daic", "qwen"): 0.7219, ("cmdc", "qwen"): 0.9539, ("turkish", "qwen"): 0.6515,
            ("d3tec", "qwen"): 0.6528, ("androids_interview", "qwen"): 0.7903,
            ("daic", "logreg"): 0.7168, ("cmdc", "logreg"): 0.9698, ("turkish", "logreg"): 0.6684,
            ("d3tec", "logreg"): 0.5733, ("androids_interview", "logreg"): 0.7714,
            ("daic", "xgb_fixed"): 0.7093, ("cmdc", "xgb_fixed"): 0.9405, ("turkish", "xgb_fixed"): 0.6288,
            ("d3tec", "xgb_fixed"): 0.5622, ("androids_interview", "xgb_fixed"): 0.8079,
        },
    },
}

DATASET_LABELS = {
    "daic": "DAIC",
    "cmdc": "CMDC",
    "turkish": "Turkish",
    "d3tec": "D3TEC",
    "androids_interview": "Androids Interview",
}
MODALITY_LABELS = {"audio_text": "Audio + Text", "audio_only": "Audio only", "text_only": "Text only"}
METHOD_LABELS = {"qwen": "Fine-tuned Qwen", "logreg": "LogReg head", "xgb_fixed": "XGBoost fixed", "xgb_optuna": "XGBoost Optuna"}
METHOD_LABELS_SHORT = {"qwen": "qwen", "logreg": "logreg", "xgb_fixed": "xgb_fixed", "xgb_optuna": "xgb_optuna"}

# --------------------------------------------------------------------------- harmonized EN
# Harmonized English-translation campaign (Issue #20 / PR #21; recipe
# harmonized_full_transcript_single30_allwindows_selmacrof1_tf_en_v1; campaign
# harmonized_en_v1_prod_20260810T152759_5dc499c4; source 5dc499c4, clean main).
# Only transcript language differs from the native harmonized control. Values
# recomputed 2026-08-10 from locally synced artifacts (predictions recomputed,
# metrics matched); native macro-F1 reproduces the native harmonized campaign
# values exactly. Audio-only cells reuse the shared native control: no new
# training (delta zero by definition).
HARMONIZED_EN_CAMPAIGN = {
    "campaign_id": "harmonized_en_v1_prod_20260810T152759_5dc499c4",
    "source_sha": "5dc499c4e0e604901072be660f3fb68a9fecec0b",
    "github_issue": 20,
    "github_pr": 21,
    "recipe_id": "harmonized_full_transcript_single30_allwindows_selmacrof1_tf_en_v1",
    "config_sha256": {
        "D3TEC": {"Audio + Text": "8fa341856fef769e", "Text only": "5fc4231da4615142"},
        "Androids Interview": {"Audio + Text": "b1b6dafd27219c72", "Text only": "c93f40b36d995a55"},
        "CMDC": {"Audio + Text": "467e0b668228c96f", "Text only": "22769cd4c22e16f6"},
        "Turkish": {"Audio + Text": "1c43453ef2fa880c", "Text only": "3b0bb378d0d9b505"},
    },
    "manifest_sha256": {
        "D3TEC": "4c2c78f97c7ece41", "Androids Interview": "e9f9ffac747aad93",
        "CMDC": "17a9fe5ea428baf3", "Turkish": "68f6c237175a1aea",
    },
    "split_sha256": {
        "D3TEC": "6c3cffd04152260d", "Androids Interview": "53f4e80aceedb22e",
        "CMDC": "09cfb5925ec50362", "Turkish": "f1ca03d18f5ad1c9",
    },
    "translation_cache_sha256": {
        "D3TEC": "6d11f8f303206dcc", "Androids Interview": "261005d04e462f2c",
        "CMDC": "3e141f6f341e2103", "Turkish": "f9cd2748f7718e6f",
    },
    "translation_cache_root": "harmonized_en_complete_v1 (7,827 accepted, 0 rejected)",
}

# (dataset, modality) -> (native_macro, native_posf1, en_macro, en_posf1, aggregation, shared)
HARMONIZED_EN_QWEN: dict[tuple[str, str], tuple[float, float, float, float, str, bool]] = {
    ("D3TEC", "Audio + Text"): (0.5589, 0.5091, 0.6125, 0.6000, "pooled 5-fold subject-level", False),
    ("D3TEC", "Audio only"): (0.6649, 0.6000, 0.6649, 0.6000, "pooled 5-fold subject-level", True),
    ("D3TEC", "Text only"): (0.6113, 0.5862, 0.5465, 0.5758, "pooled 5-fold subject-level", False),
    ("Androids Interview", "Audio + Text"): (0.8606, 0.8750, 0.8873, 0.8960, "pooled 5-fold subject-level", False),
    ("Androids Interview", "Audio only"): (0.8690, 0.8837, 0.8690, 0.8837, "pooled 5-fold subject-level", True),
    ("Androids Interview", "Text only"): (0.7317, 0.7826, 0.7921, 0.8065, "pooled 5-fold subject-level", False),
    ("CMDC", "Audio + Text"): (0.9700, 0.9600, 0.9856, 0.9818, "5-fold mean", False),
    ("CMDC", "Audio only"): (0.9516, 0.9318, 0.9516, 0.9318, "5-fold mean", True),
    ("CMDC", "Text only"): (0.9713, 0.9636, 0.9713, 0.9636, "5-fold mean", False),
    ("Turkish", "Audio + Text"): (0.6666, 0.7822, 0.6295, 0.7925, "5-fold mean", False),
    ("Turkish", "Audio only"): (0.5137, 0.7719, 0.5137, 0.7719, "5-fold mean", True),
    ("Turkish", "Text only"): (0.6502, 0.7817, 0.6641, 0.7957, "5-fold mean", False),
}

# (dataset, modality) -> (native_logreg, native_xgb, en_logreg, en_xgb) — 5-fold mean
# subject-level macro-F1 (HARMONIZED_STANDALONE_HEADS convention). Native values
# from STANDALONE_HEADS (verified 2026-08-10).
HARMONIZED_EN_HEADS: dict[tuple[str, str], tuple[float, float, float, float]] = {
    ("D3TEC", "Audio + Text"): (0.4988, 0.5585, 0.5464, 0.6134),
    ("D3TEC", "Audio only"): (0.6031, 0.5404, 0.6031, 0.5404),
    ("D3TEC", "Text only"): (0.4651, 0.5911, 0.5226, 0.5800),
    ("Androids Interview", "Audio + Text"): (0.8745, 0.8656, 0.8802, 0.8562),
    ("Androids Interview", "Audio only"): (0.8512, 0.8235, 0.8512, 0.8235),
    ("Androids Interview", "Text only"): (0.8326, 0.8241, 0.8298, 0.8210),
    ("CMDC", "Audio + Text"): (0.9614, 0.9700, 0.9856, 0.9134),
    ("CMDC", "Audio only"): (0.9841, 0.9225, 0.9841, 0.9225),
    ("CMDC", "Text only"): (0.9420, 0.9683, 0.9698, 0.9528),
    ("Turkish", "Audio + Text"): (0.6289, 0.6325, 0.6076, 0.5953),
    ("Turkish", "Audio only"): (0.5209, 0.4271, 0.5209, 0.4271),
    ("Turkish", "Text only"): (0.5875, 0.5234, 0.5944, 0.5653),
}

HARMONIZED_EN_EVIDENCE = {
    "D3TEC": "output_model/harmonized_v1_en/<modality>/d3tec/<run>/fold_<n>/best_model/standalone_eval/",
    "Androids Interview": "output_model/harmonized_v1_en/<modality>/androids/<run>/fold_<n>/best_model/standalone_eval/",
    "CMDC": "output_model/harmonized_v1_en/<modality>/cmdc/<run>/fold_<n>/eval/best_validation/",
    "Turkish": "output_model/harmonized_v1_en/<modality>/turkish_t17_qwen3asr/<run>/fold_<n>/eval/best_validation/",
}

STANDALONE_QWEN_SOURCE = {
    "DAIC": (
        "harmonized_v1_harmonized_v1_prod_20260809T171705Z_d1e8130b_daic_{audio_text,audio_only,text_only}_r1"
        " (campaign harmonized_v1_prod_20260809T171705Z_d1e8130b; retry registry retry_r1_jobs.tsv)",
        "Official test, 47 subjects, teacher-forced, binary-strict, harmonized_all_windows_full_coverage",
        "output_model/harmonized_v1/*/daic/*/fold_0/best_model/standalone_eval(_r1)/metrics_original_teacher_forced.json",
    ),
    "CMDC": (
        "harmonized_v1_harmonized_v1_prod_20260809T171705Z_d1e8130b_cmdc_{audio_text,audio_only,text_only}_r1",
        "5-fold mean, teacher-forced, binary-strict, train_val protocol",
        "output_model/harmonized_v1/*/cmdc/*/fold_<n>/eval/best_validation/metrics_original_teacher_forced.json",
    ),
    "Turkish": (
        "harmonized_v1_harmonized_v1_prod_20260809T171705Z_d1e8130b_turkish_{audio_text,audio_only,text_only}_r1",
        "5-fold mean, teacher-forced, binary-strict, train_val protocol",
        "output_model/harmonized_v1/*/turkish_t17_qwen3asr/*/fold_<n>/eval/best_validation/metrics_original_teacher_forced.json",
    ),
    "D3TEC": (
        "harmonized_v1_harmonized_v1_prod_20260809T171705Z_d1e8130b_d3tec_{audio_text,audio_only,text_only}[_r1]",
        "Pooled 5-fold subject-level (62 subjects), teacher-forced, binary-strict",
        "output_model/harmonized_v1/*/d3tec/*/fold_<n>/best_model/standalone_eval(_r1)/metrics_original_teacher_forced.json",
    ),
    "Androids Interview": (
        "harmonized_v1_harmonized_v1_prod_20260809T171705Z_d1e8130b_androids_interview_{audio_text,audio_only,text_only}[_r1]",
        "Pooled 5-fold subject-level (116 subjects), teacher-forced, binary-strict",
        "output_model/harmonized_v1/*/androids/*/fold_<n>/best_model/standalone_eval(_r1)/metrics_original_teacher_forced.json",
    ),
}

DATASETS = ["DAIC", "CMDC", "Turkish", "D3TEC", "Androids Interview"]
MODALITIES = ["Audio + Text", "Audio only", "Text only"]
HEAD_METHODS = [("logreg", "LogReg head"), ("xgb_fixed", "XGBoost fixed"), ("xgb_optuna", "XGBoost Optuna")]


# --------------------------------------------------------------------------- Gemma 4 DAIC
# First Gemma 4 DAIC backbone comparison (runbook docs/GEMMA4_DAIC_IMPLEMENTATION_RUNBOOK.md).
# Backend gemma4 on google/gemma-4-12B-it revision 707f0a3b8a3c7ad586ed01e27eafbad8a27dd0f7;
# BF16 LoRA (rank 16/alpha 32/dropout 0.05, exact 288-module decoder regex), SDPA, gradient
# checkpointing, per-device batch 1, grad accumulation 32, 4xH100 DDP, seed 1337, fold 0.
# Campaign gemma4_v1_prod_20260812T020449Z_cca3f4ae; source a6749b0 (clean main; allocator
# fix), PR #31; recipe and data identical to the Qwen harmonized campaign (same manifest
# 72e2dd20…/split 441333e0… hashes). Evaluation: original_teacher_forced,
# harmonized_all_windows_full_coverage, subject mean-score aggregation, official 47-subject
# test, headline/binary_strict (INVALID counts as wrong; INVALID=0 everywhere).
# Qwen column = the canonical harmonized _r1 DAIC rows (STANDALONE_QWEN).
# All values recomputed locally from predictions_subject_level.csv (match metrics JSON).
GEMMA4_CAMPAIGN = {
    "campaign_id": "gemma4_v1_prod_20260812T020449Z_cca3f4ae",
    "group_id": "gemma4-daic-v1-cca3f4ae",
    "source_sha": "a6749b05146fd44e7c164a4f0495c72dd72bc4d4",
    "github_issue": None,
    "github_pr": 31,
    "model": "google/gemma-4-12B-it",
    "revision": "707f0a3b8a3c7ad586ed01e27eafbad8a27dd0f7",
    "manifest_sha256": "72e2dd204b915ccba3ebf922f030531fe5678b3ea8c9c52b81b41242fe9dda17",
    "split_sha256": "441333e0c88845eeacba9ea5355a8920cdd1f70e8cf7a7c15b9547b46da51473",
}

# modality -> (run_name, attempt_id, train_job, eval_job, selected_epoch)
GEMMA4_RUNS = {
    "audio_text": (
        "gemma4_harmonized_v1_gemma4_v1_prod_20260812T020449Z_cca3f4ae_daic_audio_text_r2",
        "20260812T031624Z-gemma4_daic_audio_text_seed1337-a6749b05-146c8805",
        "44518086", "44518087", 3,
    ),
    "audio_only": (
        "gemma4_harmonized_v1_gemma4_v1_prod_20260812T020449Z_cca3f4ae_daic_audio_only",
        "20260812T020449Z-gemma4_daic_audio_only_seed1337-cca3f4ae-8789edf2",
        "44517565", "44517566", 1,
    ),
    "text_only": (
        "gemma4_harmonized_v1_gemma4_v1_prod_20260812T020449Z_cca3f4ae_daic_text_only",
        "20260812T020449Z-gemma4_daic_text_only_seed1337-cca3f4ae-ed58a7a3",
        "44517484", "44517485", 1,
    ),
}
# modality -> (macro_f1, positive_f1, accuracy, precision, recall, confusion_matrix)
GEMMA4_QWEN = {
    "audio_text": (0.7631048387096775, 0.6875000000000001, 0.7872340425531915, 0.6111111111111112, 0.7857142857142857, [[26, 7], [3, 11]]),
    "audio_only": (0.4125, 0.0, 0.7021276595744681, 0.0, 0.0, [[33, 0], [14, 0]]),
    "text_only": (0.7552083333333333, 0.6666666666666666, 0.7872340425531915, 0.625, 0.7142857142857143, [[27, 6], [4, 10]]),
}
GEMMA4_EVIDENCE = "output_model/harmonized_v1_gemma4/<modality>/daic/<run>/fold_0/best_model/standalone_eval/"

# --------------------------------------------------------------------------- Gemma 4 DAIC fixed heads
# Post-hoc fixed classifiers on the final prompt-token hidden state of each completed Gemma 4 DAIC
# best_model (docs/GEMMA4_DAIC_FIXED_HEADS_IMPLEMENTATION_RUNBOOK.md). Hidden state: final layer,
# last-valid-prompt-token pooling, float32, dimension 3840, cache gemma4_hidden_cache.v1. Fit on the
# saved 107 official training subjects (audio modes: all packed30 chunks, inverse-chunk weights
# rescaled to mean one; text: one vector per subject, weight one). Test: official 47 subjects,
# mean depressed probability over chunks, fixed >= 0.5 threshold. LogReg: StandardScaler + balanced
# LogisticRegression C=1.0 liblinear max_iter=5000 seed 1337; XGBoost: binary:logistic 300 trees
# lr 0.03 depth 2 min_child_weight 5 subsample 0.8 colsample 0.25 alpha 1.0 lambda 10.0 hist n_jobs=1
# seed 1337 (scikit-learn 1.7.0, xgboost 2.1.4). Campaign gemma4-daic-fixed-heads-v1-799cc412;
# implementation merge 799cc41 (PRs #35, #45, #46); parent checkpoints are the three successful
# Gemma training attempts above; parent adapters hashed in run_config.yaml. Backends:
# gemma4_hidden_logreg_raw / gemma4_hidden_xgb_raw. One seed — observed differences, not variance.
# All values recomputed locally by verify-local from predictions_subject_level.csv (match metrics.json).
GEMMA4_HEADS_CAMPAIGN = {
    "campaign_id": "gemma4-daic-fixed-heads-v1-799cc412",
    "group_id": "gemma4-daic-fixed-heads-v1-799cc412",
    "source_sha": "799cc4125c7db7802a003d63aa2edc02ee7174cf",
    "github_issue": None,
    "github_pr": 46,
    "model": "google/gemma-4-12B-it",
    "revision": "707f0a3b8a3c7ad586ed01e27eafbad8a27dd0f7",
    "manifest_sha256": "72e2dd204b915ccba3ebf922f030531fe5678b3ea8c9c52b81b41242fe9dda17",
    "split_sha256": "441333e0c88845eeacba9ea5355a8920cdd1f70e8cf7a7c15b9547b46da51473",
}

# modality -> (run_name, attempt_id, extract_job, heads_job, supersedes_attempt_id)
GEMMA4_HEADS_RUNS = {
    "audio_text": (
        "gemma4_daic_audio_text_fixed_heads_20260812T155635Z_799cc412",
        "20260812T155659Z-gemma4_daic_audio_text_fixed_heads_20260812t155635z_799cc412-799cc412-5ad76728",
        "44538980", "44538981",
        "20260812T153613Z-gemma4_daic_audio_text_fixed_heads_20260812t153611z_dcfaa142-dcfaa142-a739f168",
    ),
    "audio_only": (
        "gemma4_daic_audio_only_fixed_heads_20260812T155635Z_799cc412",
        "20260812T155701Z-gemma4_daic_audio_only_fixed_heads_20260812t155635z_799cc412-799cc412-4fafe36d",
        "44538982", "44538983",
        "20260812T153615Z-gemma4_daic_audio_only_fixed_heads_20260812t153611z_dcfaa142-dcfaa142-f9cac152",
    ),
    "text_only": (
        "gemma4_daic_text_only_fixed_heads_20260812T155635Z_799cc412",
        "20260812T155703Z-gemma4_daic_text_only_fixed_heads_20260812t155635z_799cc412-799cc412-735f2c58",
        "44538984", "44538985",
        "20260812T153617Z-gemma4_daic_text_only_fixed_heads_20260812t153611z_dcfaa142-dcfaa142-066f6db6",
    ),
}

# modality -> variant -> (macro_f1, positive_f1, accuracy, precision, recall, confusion_matrix)
GEMMA4_HEADS = {
    "audio_text": {
        "logreg": (0.7898658718330849, 0.7272727272727273, 0.8085106382978723, 0.631578947368421, 0.8571428571428571, [[26, 7], [2, 12]]),
        "xgb": (0.7834101382488479, 0.7096774193548386, 0.8085106382978723, 0.6470588235294118, 0.7857142857142857, [[27, 6], [3, 11]]),
    },
    "audio_only": {
        "logreg": (0.6258420085731782, 0.43478260869565216, 0.723404255319149, 0.5555555555555556, 0.35714285714285715, [[29, 4], [9, 5]]),
        "xgb": (0.4125, 0.0, 0.7021276595744681, 0.0, 0.0, [[33, 0], [14, 0]]),
    },
    "text_only": {
        "logreg": (0.70625, 0.6000000000000001, 0.7446808510638298, 0.5625, 0.6428571428571429, [[26, 7], [5, 9]]),
        "xgb": (0.6948051948051948, 0.5714285714285714, 0.7446808510638298, 0.5714285714285714, 0.5714285714285714, [[27, 6], [6, 8]]),
    },
}
GEMMA4_HEADS_BACKEND = {"logreg": "gemma4_hidden_logreg_raw", "xgb": "gemma4_hidden_xgb_raw"}

# --------------------------------------------------------------------------- Native Gemma 4 harmonized campaign (verified)
# Teacher-forced (TF) and LogReg fold-means from the local native group reports
# (outputs/experiment_reports/gemma4_harmonized/native_tf_* and native_lr/*),
# seed 1337, macro-F1, binary-strict, harmonized view. DAIC rows reuse the DAIC
# official-test campaign (GEMMA4_QWEN / GEMMA4_HEADS). XGBoost uses the
# standardized Optuna-100 fold-mean: the runbook deliberately fits no fixed XGB
# head for Gemma ("every new Gemma XGBoost result comes from the standardized
# Optuna-100 protocol, never from a fixed head").
GEMMA_NATIVE_TF = {
    ("DAIC", "Audio + Text"): 0.7631048387096775,
    ("DAIC", "Audio only"): 0.4125,
    ("DAIC", "Text only"): 0.7552083333333333,
    ("CMDC", "Audio + Text"): 1.0,
    ("CMDC", "Audio only"): 0.784022,
    ("CMDC", "Text only"): 0.957439,
    ("Turkish", "Audio + Text"): 0.565776,
    ("Turkish", "Audio only"): 0.468559,
    ("Turkish", "Text only"): 0.678178,
    ("D3TEC", "Audio + Text"): 0.571399,
    ("D3TEC", "Audio only"): 0.405413,
    ("D3TEC", "Text only"): 0.550505,
    ("Androids Interview", "Audio + Text"): 0.860541,
    ("Androids Interview", "Audio only"): 0.806987,
    ("Androids Interview", "Text only"): 0.776201,
}
GEMMA_NATIVE_LR = {
    ("DAIC", "Audio + Text"): 0.7898658718330849,
    ("DAIC", "Audio only"): 0.6258420085731782,
    ("DAIC", "Text only"): 0.70625,
    ("CMDC", "Audio + Text"): 1.0,
    ("CMDC", "Audio only"): 0.984127,
    ("CMDC", "Text only"): 0.956334,
    ("Turkish", "Audio + Text"): 0.618816,
    ("Turkish", "Audio only"): 0.529039,
    ("Turkish", "Text only"): 0.558253,
    ("D3TEC", "Audio + Text"): 0.498159,
    ("D3TEC", "Audio only"): 0.522677,
    ("D3TEC", "Text only"): 0.522456,
    ("Androids Interview", "Audio + Text"): 0.857269,
    ("Androids Interview", "Audio only"): 0.796701,
    ("Androids Interview", "Text only"): 0.802669,
}
# Paired Optuna-100 fold-mean macro-F1 (same standardized protocol, seed 1337,
# harmonized view) for the Qwen/Gemma XGB comparison rows.
QWEN_OPTUNA = {
    ("DAIC", "Audio + Text"): 0.715726, ("DAIC", "Audio only"): 0.576127, ("DAIC", "Text only"): 0.725729,
    ("CMDC", "Audio + Text"): 0.925089, ("CMDC", "Audio only"): 0.951555, ("CMDC", "Text only"): 0.953900,
    ("Turkish", "Audio + Text"): 0.631141, ("Turkish", "Audio only"): 0.514507, ("Turkish", "Text only"): 0.589237,
    ("D3TEC", "Audio + Text"): 0.523398, ("D3TEC", "Audio only"): 0.519475, ("D3TEC", "Text only"): 0.634535,
    ("Androids Interview", "Audio + Text"): 0.854305, ("Androids Interview", "Audio only"): 0.810943, ("Androids Interview", "Text only"): 0.846597,
}
GEMMA_OPTUNA = {
    ("DAIC", "Audio + Text"): 0.815686, ("DAIC", "Audio only"): 0.725729, ("DAIC", "Text only"): 0.769608,
    ("CMDC", "Audio + Text"): 0.984127, ("CMDC", "Audio only"): 0.958369, ("CMDC", "Text only"): 0.924588,
    ("Turkish", "Audio + Text"): 0.681937, ("Turkish", "Audio only"): 0.567052, ("Turkish", "Text only"): 0.520832,
    ("D3TEC", "Audio + Text"): 0.583227, ("D3TEC", "Audio only"): 0.546075, ("D3TEC", "Text only"): 0.571942,
    ("Androids Interview", "Audio + Text"): 0.856547, ("Androids Interview", "Audio only"): 0.802024, ("Androids Interview", "Text only"): 0.748177,
}
# --------------------------------------------------------------------------- Merged (symmetric) comparison values
# Qwen merged TF/LogReg from the historical merged campaign (Merged Symmetric
# Summary); Gemma merged CV TF from training selection, merged final TF from
# the postprocess teacher-forced DAIC evaluation, and LogReg from the merged
# heads (verified local evidence). XGBoost = the standardized Optuna-100
# fold-mean for both backends.
MERGED_TF = {
    ("cv", "Audio + Text"): (0.6976, 0.761239),
    ("cv", "Audio only"): (0.69046, 0.386284),
    ("cv", "Text only"): (0.75408, 0.776333),
    ("final", "Audio + Text"): (0.7631, 0.7257294429708223),
    ("final", "Audio only"): (0.5332, 0.5190058479532164),
    ("final", "Text only"): (0.7756, 0.7755968169761273),
}
MERGED_LR = {
    ("cv", "Audio + Text"): (0.70036, 0.741441),
    ("cv", "Audio only"): (0.67834, 0.637276),
    ("cv", "Text only"): (0.73994, 0.726577),
    ("final", "Audio + Text"): (0.7432, 0.7834101382488479),
    ("final", "Audio only"): (0.6189, 0.42837837837837833),
    ("final", "Text only"): (0.7157, 0.70625),
}
MERGED_XGB = {
    ("cv", "Audio + Text"): (0.741211, 0.764446),
    ("cv", "Audio only"): (0.747404, 0.700808),
    ("cv", "Text only"): (0.760274, 0.765794),
    ("final", "Audio + Text"): (0.743235, 0.745671),
    ("final", "Audio only"): (0.472823, 0.555405),
    ("final", "Text only"): (0.755208, 0.763105),
}

# --------------------------------------------------------------------------- English-translated comparison values
# Qwen TF/LogReg from the harmonized EN campaign (EN Translation sheet); Gemma
# TF fold-means from the EN training attempts' teacher-forced evaluations,
# Gemma LogReg from the EN LR summary, and XGBoost from the EN Optuna-100
# group reports (all verified local evidence).
EN_TF = {
    ("D3TEC", "Audio + Text"): (0.6125, 0.564983),
    ("D3TEC", "Text only"): (0.5465, 0.493600),
    ("Androids Interview", "Audio + Text"): (0.8873, 0.876968),
    ("Androids Interview", "Text only"): (0.7921, 0.744654),
    ("CMDC", "Audio + Text"): (0.9856, 0.955419),
    ("CMDC", "Text only"): (0.9713, 0.957439),
    ("Turkish", "Audio + Text"): (0.6295, 0.669119),
    ("Turkish", "Text only"): (0.6641, 0.677400),
}
EN_LR = {
    ("D3TEC", "Audio + Text"): (0.5464, 0.532292),
    ("D3TEC", "Text only"): (0.5226, 0.596902),
    ("Androids Interview", "Audio + Text"): (0.8802, 0.902794),
    ("Androids Interview", "Text only"): (0.8298, 0.830160),
    ("CMDC", "Audio + Text"): (0.9856, 0.926348),
    ("CMDC", "Text only"): (0.9698, 0.969835),
    ("Turkish", "Audio + Text"): (0.6076, 0.599564),
    ("Turkish", "Text only"): (0.5944, 0.601435),
}
EN_XGB = {
    ("D3TEC", "Audio + Text"): (0.5982, 0.546749),
    ("D3TEC", "Text only"): (0.5289, 0.559946),
    ("Androids Interview", "Audio + Text"): (0.8727, 0.902794),
    ("Androids Interview", "Text only"): (0.8111, 0.808427),
    ("CMDC", "Audio + Text"): (0.9134, 0.968280),
    ("CMDC", "Text only"): (0.9698, 0.939796),
    ("Turkish", "Audio + Text"): (0.6287, 0.592804),
    ("Turkish", "Text only"): (0.6584, 0.629565),
}

GEMMA_NATIVE_EVIDENCE = {
    "tf": "outputs/experiment_reports/gemma4_harmonized/native_tf_{ds}_{mod}/group_report.json",
    "lr": "outputs/experiment_reports/gemma4_harmonized/native_lr/{ds}_{mod}.json",
    "optuna": "outputs/experiment_reports/optuna100_native/{ds}_{mod}_{backend}/group_report.json",
}

GEMMA4_HEADS_EVIDENCE = "output_model/harmonized_v1_gemma4_heads/<modality>/daic/<run>/fold_0/hidden_classifiers/<variant>/"


def build_gemma4(wb: Workbook) -> None:
    ws = wb.create_sheet("Gemma 4 DAIC")
    _widths(ws, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 60, "I": 46})
    campaign = GEMMA4_CAMPAIGN
    _title(ws, "Gemma 4 vs Qwen — DAIC official test (seed 1337, fold 0, strict teacher-forced)", 8)
    _note(
        ws, 2,
        f"Gemma 4: {campaign['model']} revision {campaign['revision'][:12]}…, BF16 LoRA (288 decoder "
        f"modules, rank 16/alpha 32), SDPA, gradient checkpointing, batch 1 x 32 accum x 4 GPUs. "
        f"Campaign {campaign['campaign_id']}; source {campaign['source_sha'][:8]} (clean main); PR "
        f"#{campaign['github_pr']}; same manifest/split as the Qwen harmonized campaign "
        f"({campaign['manifest_sha256'][:12]}…/{campaign['split_sha256'][:12]}…). Evaluation: "
        f"original_teacher_forced, harmonized_all_windows_full_coverage, subject mean-score, "
        f"headline/binary_strict, INVALID counts as wrong (INVALID=0). Qwen column = canonical "
        f"harmonized _r1 DAIC rows. One seed each — differences are observations, not variance "
        f"estimates. All values recomputed locally from predictions_subject_level.csv.",
        8, height=110,
    )
    _header_row(ws, 4, [
        "Modality", "Model", "Macro-F1", "Positive-F1", "Accuracy", "Precision", "Recall",
        "Run / attempt / jobs", "Local evidence",
    ])
    row = 5
    mod_keys = ["audio_text", "audio_only", "text_only"]
    mod_labels = ["Audio + Text", "Audio only", "Text only"]
    for mod_key, mod_label in zip(mod_keys, mod_labels):
        run, attempt, train_job, eval_job, epoch = GEMMA4_RUNS[mod_key]
        g_macro, g_pos, g_acc, g_prec, g_rec, g_cm = GEMMA4_QWEN[mod_key]
        q_macro = STANDALONE_QWEN[("DAIC", mod_label)]
        heads_run, heads_attempt, heads_extract_job, heads_job, superseded = GEMMA4_HEADS_RUNS[mod_key]

        def _modality_cell(r: int) -> None:
            ws.cell(r, 1, mod_label).font = BODY_FONT
            ws.cell(r, 1).fill = BODY
            ws.cell(r, 1).alignment = LEFT
            ws.cell(r, 1).border = BORDER

        _modality_cell(row)
        _body_cell(ws, row, 2, "Gemma 4 teacher-forced head")
        _body_cell(ws, row, 3, g_macro, fmt="0.0000")
        _body_cell(ws, row, 4, g_pos, fmt="0.0000")
        _body_cell(ws, row, 5, g_acc, fmt="0.0000")
        _body_cell(ws, row, 6, g_prec, fmt="0.0000")
        _body_cell(ws, row, 7, g_rec, fmt="0.0000")
        source = (
            f"run {run}; attempt {attempt}; train {train_job} eval {eval_job} "
            f"(COMPLETED 0:0); selected epoch {epoch if epoch is not None else 'n/a'}"
        )
        if mod_key == "audio_text":
            source += (
                "; first attempt 20260812T020449Z-gemma4_daic_audio_text_seed1337-cca3f4ae-5704f1f7 "
                "FAILED (train 44517567 CUDA OOM / eval 44517568 CANCELLED); superseded by this retry"
            )
        cell = ws.cell(row, 8, source)
        cell.font = SMALL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        cell = ws.cell(row, 9, GEMMA4_EVIDENCE.replace("<modality>", mod_key).replace("<run>", run) + "metrics_original_teacher_forced.json + predictions_subject_level.csv")
        cell.font = SMALL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        row += 1
        for variant, variant_label in (("logreg", "Gemma 4 LogReg raw hidden head"), ("xgb", "Gemma 4 XGBoost raw hidden head")):
            h_macro, h_pos, h_acc, h_prec, h_rec, h_cm = GEMMA4_HEADS[mod_key][variant]
            _modality_cell(row)
            _body_cell(ws, row, 2, variant_label)
            _body_cell(ws, row, 3, h_macro, fmt="0.0000")
            _body_cell(ws, row, 4, h_pos, fmt="0.0000")
            _body_cell(ws, row, 5, h_acc, fmt="0.0000")
            _body_cell(ws, row, 6, h_prec, fmt="0.0000")
            _body_cell(ws, row, 7, h_rec, fmt="0.0000")
            heads_source = (
                f"run {heads_run}; attempt {heads_attempt}; extract {heads_extract_job} heads "
                f"{heads_job} (COMPLETED 0:0); backend {GEMMA4_HEADS_BACKEND[variant]}; "
                f"supersedes {superseded} (extract FAILED 1:0 / heads CANCELLED, wrapper race fixed in PR #45)"
            )
            cell = ws.cell(row, 8, heads_source)
            cell.font = SMALL_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            cell = ws.cell(row, 9, GEMMA4_HEADS_EVIDENCE.replace("<modality>", mod_key).replace("<run>", heads_run).replace("<variant>", variant) + "metrics.json + predictions_subject_level.csv")
            cell.font = SMALL_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            row += 1
        _modality_cell(row)
        _body_cell(ws, row, 2, "Qwen harmonized _r1 reference")
        _body_cell(ws, row, 3, q_macro, fmt="0.0000")
        _body_cell(ws, row, 4, None)
        _body_cell(ws, row, 5, None)
        _body_cell(ws, row, 6, None)
        _body_cell(ws, row, 7, None)
        cell = ws.cell(row, 8, "harmonized_v1_…_daic_" + mod_key + "_r1 (campaign harmonized_v1_prod_20260809T171705Z_d1e8130b; retry registry retry_r1_jobs.tsv)")
        cell.font = SMALL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        cell = ws.cell(row, 9, "output_model/harmonized_v1/" + mod_key + "/daic/*/fold_0/best_model/standalone_eval(_r1)/metrics_original_teacher_forced.json")
        cell.font = SMALL_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        row += 2
    _note(ws, row, "Fixed heads: final prompt-token hidden state (float32, 3840) from each Gemma 4 "
                   "best_model; fit on the saved 107 official training subjects only (audio modes: all "
                   "packed30 chunks, inverse-chunk weights; text: one vector per subject); test = official "
                   "47 subjects, mean depressed probability over chunks at >= 0.5. LogReg: StandardScaler + "
                   "balanced liblinear C=1.0 max_iter=5000; XGBoost: 300 trees lr 0.03 depth 2 "
                   "min_child_weight 5 subsample 0.8 colsample 0.25 alpha 1.0 lambda 10.0 hist n_jobs=1; "
                   "both seed 1337, sklearn 1.7.0 / xgboost 2.1.4. One seed each — differences are "
                   "observations, not variance estimates. No winner is selected from test results.", 7, height=110)
    ws.freeze_panes = "A5"


def _fmt(v: float | None) -> str:
    return "" if v is None else f"{v:.4f}"


# ---------------------------------------------------------------------------
# DAIC official-development comparison block (Qwen + Gemma 4, teacher-forced)
# and head ablation block (all eighteen backbone/modality/head cells).
# DAIC official-development campaign (runbook
# docs/DAIC_OFFICIAL_DEV_QWEN_GEMMA_RUNBOOK.md). Every value was recomputed
# locally from synced predictions and cross-checked against the registry
# (tools/exp.py provenance). Official-test rows elsewhere in this workbook
# are untouched.
DAIC_OFFICIALDEV_CAMPAIGN = {
    "campaign_id": "prod_20260813T150000Z_22f85e6e",
    "group_id": "daic-officialdev-qwen-gemma-v1-22f85e6e473b",
    "source_sha": "22f85e6e473bf8f4e2d0d71cc9ed0ea1098877d7",
    "github_issue": 60,
    "github_pr": 52,
    "manifest_sha256": "72e2dd204b915ccba3ebf922f030531fe5678b3ea8c9c52b81b41242fe9dda17",
    "split_sha256": "441333e0c88845eeacba9ea5355a8920cdd1f70e8cf7a7c15b9547b46da51473",
}
# Literature values on the official DAIC development partition, treated as
# Macro-F1 per the researcher decision; citations preserved.
DAIC_OFFICIALDEV_LITERATURE = [
    ("DepresInstruct (Li et al., 2025)", "Qwen2-Audio-7B", "Audio only", 0.6667, 0.7714,
     "DAIC development", "C — reported as F1; shown as Macro-F1 per researcher decision",
     "https://doi.org/10.1016/j.inffus.2025.104077"),
    ("DepresInstruct (Li et al., 2025)", "Qwen2-7B", "Text only", 0.7273, 0.7429,
     "DAIC development", "C — reported as F1; shown as Macro-F1 per researcher decision",
     "https://doi.org/10.1016/j.inffus.2025.104077"),
    ("DepresInstruct (Li et al., 2025)", "Qwen2-Audio-7B", "Audio + Text", 0.7619, 0.8571,
     "DAIC development", "C — reported as F1; shown as Macro-F1 per researcher decision",
     "https://doi.org/10.1016/j.inffus.2025.104077"),
    ("IT HEARS — Qwen2-7B baseline", "Qwen2-7B", "Text only", 0.564, None,
     "DAIC development", "C — paper reports F1; baseline; no validation/accuracy reported",
     "https://arxiv.org/abs/2511.19877"),
    ("IT HEARS — Qwen2-Audio-7B", "Qwen2-Audio-7B", "Audio + Text", 0.72, None,
     "DAIC development", "C — paper reports F1; no validation/accuracy reported",
     "https://arxiv.org/abs/2511.19877"),
]

DAIC_OFFICIALDEV_TEACHER_FORCED: dict[tuple[str, str], dict[str, Any]] = {
    ("audio_only", "qwen"): {
        "macro_f1": 0.3859649122807018,
        "positive_f1": 0.0,
        "accuracy": 0.6285714285714286,
        "precision": 0.0,
        "recall": 0.0,
        "selected_epoch": 7,
        "attempt_id": "20260813T142847Z-daic_officialdev_qwen_audio_only_seed1337-22f85e6e-9c326e1d",
        "evaluation_id": "eval-7c1ef60108d44e8081fa66cd",
    },
    ("audio_text", "qwen"): {
        "macro_f1": 0.6258503401360545,
        "positive_f1": 0.4761904761904762,
        "accuracy": 0.6857142857142857,
        "precision": 0.5555555555555556,
        "recall": 0.4166666666666667,
        "selected_epoch": 2,
        "attempt_id": "20260813T142848Z-daic_officialdev_qwen_audio_text_seed1337-22f85e6e-16e99f52",
        "evaluation_id": "eval-c4dd0ff1892bd0e3a39cc9bb",
    },
    ("text_only", "qwen"): {
        "macro_f1": 0.7086031452358927,
        "positive_f1": 0.6086956521739131,
        "accuracy": 0.7428571428571429,
        "precision": 0.6363636363636364,
        "recall": 0.5833333333333334,
        "selected_epoch": 1,
        "attempt_id": "20260813T142848Z-daic_officialdev_qwen_text_only_seed1337-22f85e6e-713b6d93",
        "evaluation_id": "eval-6a466b499f152ea1b9c3c489",
    },
    ("audio_only", "gemma4"): {
        "macro_f1": 0.39655172413793105,
        "positive_f1": 0.0,
        "accuracy": 0.6571428571428571,
        "precision": 0.0,
        "recall": 0.0,
        "selected_epoch": 1,
        "attempt_id": "20260813T142848Z-daic_officialdev_gemma4_audio_only_seed1337-22f85e6e-bb8bbbed",
        "evaluation_id": "eval-6df29e41d91c3bd14a35b17a",
    },
    ("audio_text", "gemma4"): {
        "macro_f1": 0.7552447552447553,
        "positive_f1": 0.6923076923076924,
        "accuracy": 0.7714285714285715,
        "precision": 0.6428571428571429,
        "recall": 0.75,
        "selected_epoch": 2,
        "attempt_id": "20260813T142848Z-daic_officialdev_gemma4_audio_text_seed1337-22f85e6e-df9dad6b",
        "evaluation_id": "eval-0eb033332e85c3a84b70b52e",
    },
    ("text_only", "gemma4"): {
        "macro_f1": 0.7822222222222222,
        "positive_f1": 0.7199999999999999,
        "accuracy": 0.8,
        "precision": 0.6923076923076923,
        "recall": 0.75,
        "selected_epoch": 4,
        "attempt_id": "20260813T142849Z-daic_officialdev_gemma4_text_only_seed1337-22f85e6e-75450bbe",
        "evaluation_id": "eval-8a5d52df50d78e5dbeae4008",
    },
}

DAIC_OFFICIALDEV_HEADS: dict[tuple[str, str, str], dict[str, Any]] = {
    ("audio_only", "qwen", "logreg"): {
        "macro_f1": 0.5716783216783217,
        "positive_f1": 0.4615384615384615,
        "accuracy": 0.6,
        "precision": 0.42857142857142855,
        "recall": 0.5,
        "run_name": "daic_officialdev_qwen_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r4",
        "attempt_id": "20260813T154455Z-daic_officialdev_qwen_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r4-90c6cc39-269b5f37",
        "evaluation_id": "eval-e84ce4e0072e72d7eff086bb",
        "backend": "qwen_hidden_logreg_raw",
    },
    ("audio_only", "qwen", "xgb"): {
        "macro_f1": 0.4484848484848484,
        "positive_f1": 0.13333333333333333,
        "accuracy": 0.6285714285714286,
        "precision": 0.3333333333333333,
        "recall": 0.08333333333333333,
        "run_name": "daic_officialdev_qwen_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r4",
        "attempt_id": "20260813T154455Z-daic_officialdev_qwen_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r4-90c6cc39-269b5f37",
        "evaluation_id": "eval-125e3b8f5bcc860eb4369ceb",
        "backend": "qwen_hidden_xgb_raw",
    },
    ("audio_text", "qwen", "logreg"): {
        "macro_f1": 0.6577777777777778,
        "positive_f1": 0.5599999999999999,
        "accuracy": 0.6857142857142857,
        "precision": 0.5384615384615384,
        "recall": 0.5833333333333334,
        "run_name": "daic_officialdev_qwen_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5",
        "attempt_id": "20260813T163105Z-daic_officialdev_qwen_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5-90c6cc39-b7d64c2f",
        "evaluation_id": "eval-536fc321c7611b917b42fdf6",
        "backend": "qwen_hidden_logreg_raw",
    },
    ("audio_text", "qwen", "xgb"): {
        "macro_f1": 0.5304437564499485,
        "positive_f1": 0.3157894736842105,
        "accuracy": 0.6285714285714286,
        "precision": 0.42857142857142855,
        "recall": 0.25,
        "run_name": "daic_officialdev_qwen_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5",
        "attempt_id": "20260813T163105Z-daic_officialdev_qwen_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5-90c6cc39-b7d64c2f",
        "evaluation_id": "eval-c9a4f95ffa07bebb8aa2c104",
        "backend": "qwen_hidden_xgb_raw",
    },
    ("text_only", "qwen", "logreg"): {
        "macro_f1": 0.6938775510204083,
        "positive_f1": 0.5714285714285715,
        "accuracy": 0.7428571428571429,
        "precision": 0.6666666666666666,
        "recall": 0.5,
        "run_name": "daic_officialdev_qwen_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r3",
        "attempt_id": "20260813T153611Z-daic_officialdev_qwen_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r3-90c6cc39-df721ed6",
        "evaluation_id": "eval-225153901c0e8896c9900e4a",
        "backend": "qwen_hidden_logreg_raw",
    },
    ("text_only", "qwen", "xgb"): {
        "macro_f1": 0.6026831785345718,
        "positive_f1": 0.4210526315789474,
        "accuracy": 0.6857142857142857,
        "precision": 0.5714285714285714,
        "recall": 0.3333333333333333,
        "run_name": "daic_officialdev_qwen_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r3",
        "attempt_id": "20260813T153611Z-daic_officialdev_qwen_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r3-90c6cc39-df721ed6",
        "evaluation_id": "eval-28be66b87b3641742e847e9d",
        "backend": "qwen_hidden_xgb_raw",
    },
    ("audio_only", "gemma4", "logreg"): {
        "macro_f1": 0.6499999999999999,
        "positive_f1": 0.5,
        "accuracy": 0.7142857142857143,
        "precision": 0.625,
        "recall": 0.4166666666666667,
        "run_name": "daic_officialdev_gemma4_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_1327849b_r2",
        "attempt_id": "20260813T151846Z-daic_officialdev_gemma4_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_1327849b_r2-1327849b-640f4031",
        "evaluation_id": "eval-3e3a37d62c53729cf5db69b5",
        "backend": "gemma4_hidden_logreg_raw",
    },
    ("audio_only", "gemma4", "xgb"): {
        "macro_f1": 0.39655172413793105,
        "positive_f1": 0.0,
        "accuracy": 0.6571428571428571,
        "precision": 0.0,
        "recall": 0.0,
        "run_name": "daic_officialdev_gemma4_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_1327849b_r2",
        "attempt_id": "20260813T151846Z-daic_officialdev_gemma4_audio_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_1327849b_r2-1327849b-640f4031",
        "evaluation_id": "eval-ad6e2a2ce408aed6f96f39d3",
        "backend": "gemma4_hidden_xgb_raw",
    },
    ("audio_text", "gemma4", "logreg"): {
        "macro_f1": 0.6829710144927537,
        "positive_f1": 0.5833333333333334,
        "accuracy": 0.7142857142857143,
        "precision": 0.5833333333333334,
        "recall": 0.5833333333333334,
        "run_name": "daic_officialdev_gemma4_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r6",
        "attempt_id": "20260813T175117Z-daic_officialdev_gemma4_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r6-90c6cc39-438b3be6",
        "evaluation_id": "eval-70d39c1caadc8712d02ae7fb",
        "backend": "gemma4_hidden_logreg_raw",
    },
    ("audio_text", "gemma4", "xgb"): {
        "macro_f1": 0.6829710144927537,
        "positive_f1": 0.5833333333333334,
        "accuracy": 0.7142857142857143,
        "precision": 0.5833333333333334,
        "recall": 0.5833333333333334,
        "run_name": "daic_officialdev_gemma4_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r6",
        "attempt_id": "20260813T175117Z-daic_officialdev_gemma4_audio_text_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r6-90c6cc39-438b3be6",
        "evaluation_id": "eval-e480a1844b709554d1d22b7c",
        "backend": "gemma4_hidden_xgb_raw",
    },
    ("text_only", "gemma4", "logreg"): {
        "macro_f1": 0.7619047619047619,
        "positive_f1": 0.6666666666666666,
        "accuracy": 0.8,
        "precision": 0.7777777777777778,
        "recall": 0.5833333333333334,
        "run_name": "daic_officialdev_gemma4_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5",
        "attempt_id": "20260813T163105Z-daic_officialdev_gemma4_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5-90c6cc39-bee30555",
        "evaluation_id": "eval-846a6445ccd968563f454777",
        "backend": "gemma4_hidden_logreg_raw",
    },
    ("text_only", "gemma4", "xgb"): {
        "macro_f1": 0.6685606060606061,
        "positive_f1": 0.5454545454545454,
        "accuracy": 0.7142857142857143,
        "precision": 0.6,
        "recall": 0.5,
        "run_name": "daic_officialdev_gemma4_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5",
        "attempt_id": "20260813T163105Z-daic_officialdev_gemma4_text_only_fixed_heads_seed1337_prod_20260813t150000z_22f85e6e_90c6cc39_r5-90c6cc39-bee30555",
        "evaluation_id": "eval-f5173eec4fb4b8a5fb612e04",
        "backend": "gemma4_hidden_xgb_raw",
    },
}


TF_EVIDENCE = {
    ("audio_only", "qwen"): "output_model/harmonized_v1_officialdev/audio_only/daic/daic_officialdev_qwen_audio_only_seed1337_prod_20260813T150000Z_22f85e6e_22f85e6e/fold_0/best_model/standalone_eval/",
    ("audio_text", "qwen"): "output_model/harmonized_v1_officialdev/audio_text/daic/daic_officialdev_qwen_audio_text_seed1337_prod_20260813T150000Z_22f85e6e_22f85e6e/fold_0/best_model/standalone_eval/",
    ("text_only", "qwen"): "output_model/harmonized_v1_officialdev/text_only/daic/daic_officialdev_qwen_text_only_seed1337_prod_20260813T150000Z_22f85e6e_22f85e6e/fold_0/best_model/standalone_eval/",
    ("audio_only", "gemma4"): "output_model/harmonized_v1_gemma4_officialdev/audio_only/daic/daic_officialdev_gemma4_audio_only_seed1337_prod_20260813T150000Z_22f85e6e_22f85e6e/fold_0/best_model/standalone_eval/",
    ("audio_text", "gemma4"): "output_model/harmonized_v1_gemma4_officialdev/audio_text/daic/daic_officialdev_gemma4_audio_text_seed1337_prod_20260813T150000Z_22f85e6e_22f85e6e/fold_0/best_model/standalone_eval/",
    ("text_only", "gemma4"): "output_model/harmonized_v1_gemma4_officialdev/text_only/daic/daic_officialdev_gemma4_text_only_seed1337_prod_20260813T150000Z_22f85e6e_22f85e6e/fold_0/best_model/standalone_eval/",
}
HEAD_EVIDENCE = {
    ("audio_only", "qwen"): "output_model/harmonized_v1_officialdev_heads/audio_only/daic/",
    ("audio_text", "qwen"): "output_model/harmonized_v1_officialdev_heads/audio_text/daic/",
    ("text_only", "qwen"): "output_model/harmonized_v1_officialdev_heads/text_only/daic/",
    ("audio_only", "gemma4"): "output_model/harmonized_v1_gemma4_officialdev_heads/audio_only/daic/",
    ("audio_text", "gemma4"): "output_model/harmonized_v1_gemma4_officialdev_heads/audio_text/daic/",
    ("text_only", "gemma4"): "output_model/harmonized_v1_gemma4_officialdev_heads/text_only/daic/",
}
BACKBONE_LABELS = {"qwen": "Qwen", "gemma4": "Gemma 4 12B IT"}
MODALITY_SHORT = {"audio_only": "A", "audio_text": "A+T", "text_only": "T"}
HEAD_LABELS = {"logreg": "LogReg raw", "xgb": "XGBoost raw"}


def build_daic_officialdev_comparison(wb: Workbook, *, detailed: bool) -> None:
    """DAIC LLM Comparison — literature and the six official-development
    teacher-forced rows (35-subject official development partition)."""
    ws = wb.create_sheet("DAIC LLM Comparison")
    _widths(ws, {"A": 30, "B": 18, "C": 13, "D": 13, "E": 13, "F": 13, "G": 22, "H": 46, "I": 46})
    _title(ws, "DAIC-WOZ LLM literature comparison — official development partition (35 subjects)", 9)
    _section(ws, 2, "Cited literature (DAIC official development partition)", 9)
    _header_row(
        ws, 3,
        ["Study", "Backbone", "Modality", "Macro-F1", "Accuracy", "Partition", "Comparability", "Source", "Note"],
    )
    row = 4
    for study, backbone, modality, macro, acc, partition, comparability, source in DAIC_OFFICIALDEV_LITERATURE:
        ws.cell(row, 1, study)
        ws.cell(row, 2, backbone)
        ws.cell(row, 3, modality)
        ws.cell(row, 4, macro)
        ws.cell(row, 5, acc)
        ws.cell(row, 6, partition)
        ws.cell(row, 7, comparability)
        ws.cell(row, 8, source)
        row += 1
    _section(ws, row, "Ours — official development (35 subjects, seed 1337, fold 0, best_model)", 9)
    row += 1
    _header_row(
        ws, row,
        ["Backbone", "Modality", "Macro-F1", "Positive-F1", "Accuracy", "Selected epoch", "Support", "Evidence", "Verification"],
    )
    row += 1
    for (modality, backbone) in sorted(DAIC_OFFICIALDEV_TEACHER_FORCED):
        value = DAIC_OFFICIALDEV_TEACHER_FORCED[(modality, backbone)]
        evidence = TF_EVIDENCE[(modality, backbone)]
        ws.cell(row, 1, BACKBONE_LABELS[backbone])
        ws.cell(row, 2, MODALITY_LABELS[modality])
        ws.cell(row, 3, value["macro_f1"])
        ws.cell(row, 4, value["positive_f1"])
        ws.cell(row, 5, value["accuracy"])
        ws.cell(row, 6, value["selected_epoch"])
        ws.cell(row, 7, 35)
        ws.cell(row, 8, evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv")
        ws.cell(row, 9, "recomputed locally from predictions; attempt " + value["attempt_id"][:20] + "…")
        for col in (3, 4, 5):
            ws.cell(row, col).number_format = "0.0000"
        row += 1


def build_daic_officialdev_heads(wb: Workbook, *, detailed: bool) -> None:
    """DAIC Head Ablation — all eighteen official-development
    backbone/modality/head cells (teacher-forced, LogReg, XGBoost)."""
    ws = wb.create_sheet("DAIC Head Ablation")
    _widths(ws, {"A": 16, "B": 14, "C": 14, "D": 13, "E": 13, "F": 13, "G": 13, "H": 13, "I": 10, "J": 50, "K": 40})
    _title(ws, "DAIC-WOZ downstream-head ablation — official development partition (35 subjects)", 11)
    _header_row(
        ws, 2,
        ["Backbone", "Modality", "Head", "Macro-F1", "Positive-F1", "Accuracy", "Precision", "Recall", "Support", "Evidence", "Verification"],
    )
    row = 3
    for modality in ("audio_only", "audio_text", "text_only"):
        for backbone in ("qwen", "gemma4"):
            tf = DAIC_OFFICIALDEV_TEACHER_FORCED[(modality, backbone)]
            tf_evidence = TF_EVIDENCE[(modality, backbone)]
            ws.cell(row, 1, BACKBONE_LABELS[backbone])
            ws.cell(row, 2, MODALITY_LABELS[modality])
            ws.cell(row, 3, "Teacher-forced")
            for col, key in ((4, "macro_f1"), (5, "positive_f1"), (6, "accuracy"), (7, "precision"), (8, "recall")):
                ws.cell(row, col, tf[key])
            ws.cell(row, 9, 35)
            ws.cell(row, 10, tf_evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv")
            ws.cell(row, 11, "recomputed locally; attempt " + tf["attempt_id"][:20] + "…, epoch " + str(tf["selected_epoch"]))
            for col in (4, 5, 6, 7, 8):
                ws.cell(row, col).number_format = "0.0000"
            row += 1
            for variant in ("logreg", "xgb"):
                head = DAIC_OFFICIALDEV_HEADS[(modality, backbone, variant)]
                head_evidence = HEAD_EVIDENCE[(modality, backbone)] + head["run_name"] + "/fold_0/hidden_classifiers/" + variant + "/"
                ws.cell(row, 1, BACKBONE_LABELS[backbone])
                ws.cell(row, 2, MODALITY_LABELS[modality])
                ws.cell(row, 3, HEAD_LABELS[variant])
                for col, key in ((4, "macro_f1"), (5, "positive_f1"), (6, "accuracy"), (7, "precision"), (8, "recall")):
                    ws.cell(row, col, head[key])
                ws.cell(row, 9, 35)
                ws.cell(row, 10, head_evidence + "metrics.json + predictions_subject_level.csv")
                ws.cell(row, 11, "recomputed locally by verify-local; backend " + head["backend"])
                for col in (4, 5, 6, 7, 8):
                    ws.cell(row, col).number_format = "0.0000"
                row += 1


def build_gemma_vs_qwen(wb: Workbook) -> None:
    """Qwen vs Gemma — the unified comparison across the three main experiments
    (standalone native, symmetric merged, English-translated), for the three
    methods (teacher-forced, LogReg head, XGBoost). The XGBoost rows use the
    standardized search (100 trials by default, noted once); both backends are
    shown side by side with the Gemma-minus-Qwen delta.
    """
    ws = wb.create_sheet("Qwen vs Gemma")
    _widths(ws, {"A": 18, "B": 26, "C": 22, "D": 17, "E": 17, "F": 17, "G": 15})
    _title(ws, "Qwen vs Gemma 4 — macro-F1 (seed 1337)", 7)
    _note(
        ws, 2,
        "Three main experiments, both models, macro-F1 (binary-strict, best_model, harmonized view). "
        "XGBoost uses the standardized search of 100 trials (the default), seed 1337, for both models; "
        "the runbook fits no fixed XGB head for Gemma. Delta = Gemma minus Qwen. DAIC = official "
        "47-subject test; CMDC/Turkish = 5-fold mean (train_val); D3TEC/Androids = pooled 5-fold "
        "subject-level; merged CV = mean over the five datasets; merged Final = DAIC official test. "
        "Per-cell provenance: Provenance sheet.",
        7, height=110,
    )
    _header_row(ws, 4, ["Experiment", "Dataset", "Modality", "Method", "Qwen", "Gemma 4", "Δ (Gemma − Qwen)"])
    mod_keys = ["Audio + Text", "Audio only", "Text only"]
    row = 5

    _section(ws, row, "Standalone (native)", 7)
    row += 1
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish", "DAIC"):
        for mod_label in mod_keys:
            qwen_tf = STANDALONE_QWEN[(dataset, mod_label)]
            gemma_tf = GEMMA_NATIVE_TF[(dataset, mod_label)]
            cells = [("Teacher-forced", qwen_tf, gemma_tf)]
            logreg = STANDALONE_HEADS[(dataset, mod_label)][0]
            if logreg is not None:
                cells.append(("LogReg head", logreg, GEMMA_NATIVE_LR[(dataset, mod_label)]))
            cells.append(("XGBoost", QWEN_OPTUNA[(dataset, mod_label)], GEMMA_OPTUNA[(dataset, mod_label)]))
            for method, q, g in cells:
                _fill_cell(ws, row, "Standalone", dataset, mod_label, method, q, g)
                row += 1

    _section(ws, row, "Merged (symmetric)", 7)
    row += 1
    for stage, stage_label in (("cv", "CV (5-fold)"), ("final", "Final (DAIC test)")):
        for mod_label in mod_keys:
            for method, table in (("Teacher-forced", MERGED_TF), ("LogReg head", MERGED_LR), ("XGBoost", MERGED_XGB)):
                q, g = table[(stage, mod_label)]
                if g is None:
                    continue
                _fill_cell(ws, row, f"Merged — {stage_label}", "Merged", mod_label, method, q, g)
                row += 1

    _section(ws, row, "English (translated)", 7)
    row += 1
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish"):
        for mod_label in ("Audio + Text", "Text only"):
            for method, table in (("Teacher-forced", EN_TF), ("LogReg head", EN_LR), ("XGBoost", EN_XGB)):
                q, g = table[(dataset, mod_label)]
                _fill_cell(ws, row, "English", dataset, mod_label, method, q, g)
                row += 1

    _note(ws, row, "Teacher-forced = backbone classification without hidden-state heads. Hidden heads "
                   "classify the final prompt-token hidden state with the locked LogReg or the standardized "
                   "XGBoost implementation. See the Provenance sheet for every value's run, aggregation, and "
                   "local artifact.", 7, height=60)
    ws.freeze_panes = "A5"


def _fill_cell(ws, row: int, experiment: str, dataset: str, modality: str, method: str,
               qwen_value: float | None, gemma_value: float | None) -> None:
    ws.cell(row, 1, experiment).font = BODY_FONT
    ws.cell(row, 1).fill = BODY
    ws.cell(row, 1).alignment = LEFT
    ws.cell(row, 1).border = BORDER
    _body_cell(ws, row, 2, dataset)
    _body_cell(ws, row, 3, modality)
    _body_cell(ws, row, 4, method)
    _body_cell(ws, row, 5, qwen_value, fmt="0.0000")
    _body_cell(ws, row, 6, gemma_value, fmt="0.0000")
    if qwen_value is not None and gemma_value is not None:
        _delta_cell(ws, row, 7, gemma_value - qwen_value)
    else:
        _body_cell(ws, row, 7, None)


def build_native_vs_english(wb: Workbook) -> None:
    """Paired native-vs-English comparison for both current backends.

    Translation changes transcript-bearing inputs only, so the sheet contains
    Audio + Text and Text only. Audio-only would be the same native control,
    not a separate English experiment. XGBoost values use the standardized
    Optuna-100 protocol, matching the main comparison sheets.
    """
    ws = wb.create_sheet("Native vs EN")
    _widths(ws, {
        "A": 22, "B": 18, "C": 17,
        "D": 14, "E": 14, "F": 16,
        "G": 14, "H": 14, "I": 16, "J": 24,
    })
    _title(ws, "Native vs English-translated transcripts — macro-F1", 10)
    _note(
        ws, 2,
        "Paired native and English-translated transcript conditions for Qwen and Gemma 4. "
        "Delta = EN minus native; positive means translation improved macro-F1. Teacher-forced "
        "and LogReg use the harmonized best_model evidence; XGBoost uses the standardized "
        "100-trial Optuna search, seed 1337. D3TEC/Androids = pooled 5-fold subject-level; "
        "CMDC/Turkish = 5-fold mean (train_val). Audio-only is omitted because transcript "
        "translation does not create a distinct audio-only experiment. Per-cell evidence is in "
        "the Provenance sheet.",
        10, height=92,
    )
    _header_row(ws, 4, [
        "Dataset", "Modality", "Method",
        "Qwen native", "Qwen EN", "Qwen Δ (EN − native)",
        "Gemma native", "Gemma EN", "Gemma Δ (EN − native)",
        "Direction",
    ])
    row = 5
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish"):
        for modality in ("Audio + Text", "Text only"):
            native_lr = STANDALONE_HEADS[(dataset, modality)][0]
            comparisons = (
                (
                    "Teacher-forced",
                    STANDALONE_QWEN[(dataset, modality)],
                    EN_TF[(dataset, modality)][0],
                    GEMMA_NATIVE_TF[(dataset, modality)],
                    EN_TF[(dataset, modality)][1],
                ),
                (
                    "LogReg head",
                    native_lr,
                    EN_LR[(dataset, modality)][0],
                    GEMMA_NATIVE_LR[(dataset, modality)],
                    EN_LR[(dataset, modality)][1],
                ),
                (
                    "XGBoost",
                    QWEN_OPTUNA[(dataset, modality)],
                    EN_XGB[(dataset, modality)][0],
                    GEMMA_OPTUNA[(dataset, modality)],
                    EN_XGB[(dataset, modality)][1],
                ),
            )
            for method, q_native, q_en, g_native, g_en in comparisons:
                q_delta = q_en - q_native
                g_delta = g_en - g_native
                _body_cell(ws, row, 1, dataset)
                _body_cell(ws, row, 2, modality)
                _body_cell(ws, row, 3, method)
                _body_cell(ws, row, 4, q_native, fmt="0.0000")
                _body_cell(ws, row, 5, q_en, fmt="0.0000")
                _delta_cell(ws, row, 6, q_delta)
                _body_cell(ws, row, 7, g_native, fmt="0.0000")
                _body_cell(ws, row, 8, g_en, fmt="0.0000")
                _delta_cell(ws, row, 9, g_delta)
                if abs(q_delta) < 0.03 and abs(g_delta) < 0.03:
                    direction = "~tie for both"
                elif q_delta >= 0.03 and g_delta >= 0.03:
                    direction = "EN better for both"
                elif q_delta <= -0.03 and g_delta <= -0.03:
                    direction = "Native better for both"
                elif q_delta > g_delta:
                    direction = "EN helps Qwen more"
                else:
                    direction = "EN helps Gemma more"
                _body_cell(ws, row, 10, direction)
                row += 1
    _note(
        ws, row,
        "Direction uses |Δ| < 0.03 as a practical tie for the compact label only; the exact "
        "deltas remain visible. This is descriptive and is not a significance test.",
        10, height=42,
    )
    ws.freeze_panes = "A5"


# --------------------------------------------------------------------------- sheets
def build_summary(wb: Workbook, *, detailed: bool) -> None:
    """Compact headline: the standardized XGBoost macro-F1 for the three main
    experiments, both models, with the Gemma-minus-Qwen delta. The full
    teacher-forced / LogReg / XGBoost detail lives in the Qwen vs Gemma sheet.
    """
    ws = wb.create_sheet("Summary")
    _widths(ws, {"A": 22, "B": 26, "C": 14, "D": 14, "E": 14})
    _title(ws, "Depression Detection — XGBoost macro-F1 summary (Qwen vs Gemma)", 5)
    _note(
        ws, 2,
        "Standardized XGBoost (100-trial search, seed 1337, macro-F1) for the three main experiments, "
        "both models. Delta = Gemma minus Qwen. Full teacher-forced, LogReg, and XGBoost detail: "
        "'Qwen vs Gemma' sheet. Per-cell provenance: Provenance sheet.",
        5, height=70,
    )
    _header_row(ws, 4, ["Experiment", "Dataset / Stage", "Qwen", "Gemma 4", "Δ (Gemma − Qwen)"])
    row = 5
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish", "DAIC"):
        for mod_label in ("Audio + Text", "Audio only", "Text only"):
            q = QWEN_OPTUNA[(dataset, mod_label)]
            g = GEMMA_OPTUNA[(dataset, mod_label)]
            _summary_row(ws, row, "Standalone", f"{dataset} — {mod_label}", q, g)
            row += 1
    for stage, stage_label in (("cv", "CV (5-fold)"), ("final", "Final (DAIC test)")):
        for mod_label in ("Audio + Text", "Audio only", "Text only"):
            q, g = MERGED_XGB[(stage, mod_label)]
            _summary_row(ws, row, "Merged", f"{stage_label} — {mod_label}", q, g)
            row += 1
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish"):
        for mod_label in ("Audio + Text", "Text only"):
            q, g = EN_XGB[(dataset, mod_label)]
            _summary_row(ws, row, "English", f"{dataset} — {mod_label}", q, g)
            row += 1
    ws.freeze_panes = "A5"


def _summary_row(ws, row: int, experiment: str, label: str, qwen_value: float, gemma_value: float) -> None:
    ws.cell(row, 1, experiment).font = BODY_FONT
    ws.cell(row, 1).fill = BODY
    ws.cell(row, 1).alignment = LEFT
    ws.cell(row, 1).border = BORDER
    _body_cell(ws, row, 2, label)
    _body_cell(ws, row, 3, qwen_value, fmt="0.0000")
    _body_cell(ws, row, 4, gemma_value, fmt="0.0000")
    _delta_cell(ws, row, 5, gemma_value - qwen_value)


def build_merged_summary(wb: Workbook, *, detailed: bool) -> None:
    ws = wb.create_sheet("Merged Symmetric Summary")
    _widths(ws, {"A": 34, "B": 15, "C": 15, "D": 15, "E": 15})
    _title(ws, "Symmetric Merged — Macro-F1 Summary (clean, verified)", 5)
    _note(
        ws, 2,
        "Macro-F1 only; higher is better. Harmonized merged campaign run "
        "harmonized_v1_prod_20260809T171705Z_d1e8130b (final epochs 2/4/5 = rounded median of CV-selected epochs; "
        "daic_official_test_only). CV values = mean of the five per-fold holdout macro-F1 values. "
        "Final values = the protected DAIC official holdout (47 subjects). "
        "Optuna heads not run -> blank. All values recomputed 2026-08-10 from local postprocess summaries.",
        5, height=78,
    )
    methods = [("qwen", "Fine-tuned Qwen"), ("logreg", "LogReg head"), ("xgb_fixed", "XGBoost fixed")]
    if detailed:
        methods.append(("xgb_optuna", "XGBoost Optuna"))
    headers = ["Evaluation / Modality", *(label for _, label in methods)]

    def cv_stats(modality: str, method: str) -> tuple[float | None, float | None]:
        values = [v for (ds, m), v in MERGED_RUNS[modality]["cv"].items() if m == method]
        if not values:
            return None, None
        return sum(values) / len(values), min(values)

    def table(row: int, getter) -> None:
        _header_row(ws, row, headers)
        for offset, (modality, modality_label) in enumerate(zip(["audio_text", "audio_only", "text_only"], MODALITIES), start=1):
            current = row + offset
            ws.cell(current, 1, modality_label).font = BODY_FONT
            ws.cell(current, 1).fill = BODY
            ws.cell(current, 1).alignment = LEFT
            ws.cell(current, 1).border = BORDER
            for col, (method, _) in enumerate(methods, start=2):
                _body_cell(ws, current, col, getter(modality, method), fmt="0.0000")

    _section(ws, 6, "DAIC Official Holdout — Final Macro-F1", 5)
    table(7, lambda mod, method: MERGED_RUNS[mod]["official"].get(method))

    _section(ws, 11, "Five-Dataset CV — Pooled Mean Macro-F1", 5)
    table(12, lambda mod, method: cv_stats(mod, method)[0])

    _section(ws, 16, "Five-Dataset CV — Worst-Dataset Macro-F1", 5)
    table(17, lambda mod, method: cv_stats(mod, method)[1])

    _section(ws, 21, "Five-Dataset CV — Macro-F1 by Dataset and Modality", 5)
    _header_row(ws, 22, headers)
    row = 23
    for modality in ["audio_text", "audio_only", "text_only"]:
        for dataset in ["daic", "cmdc", "turkish", "d3tec", "androids_interview"]:
            label = f"{DATASET_LABELS[dataset]} — {MODALITY_LABELS[modality]}"
            ws.cell(row, 1, label).font = BODY_FONT
            ws.cell(row, 1).fill = BODY
            ws.cell(row, 1).alignment = LEFT
            ws.cell(row, 1).border = BORDER
            for col, (method, _) in enumerate(methods, start=2):
                _body_cell(ws, row, col, MERGED_RUNS[modality]["cv"].get((dataset, method)), fmt="0.0000")
            row += 1
    ws.freeze_panes = "A5"


def _optuna_report_values(family: str, dataset: str, modality: str, backend: str) -> tuple:
    """Look up qualified macro-F1 / positive-F1 for an Optuna-100 workbook cell."""
    backend_key = "qwen" if backend == "Qwen" else "gemma4"
    modality_key = {
        "Audio + Text": "audio_text",
        "Audio only": "audio_only",
        "Text only": "text_only",
    }[modality]
    if family == "Symmetric merged":
        report_root = PROJECT_ROOT / "outputs/experiment_reports/optuna100_merged"
        stage_key = "cv" if dataset == "CV (5-fold)" else "final"
        cell = f"{modality_key}_{stage_key}_{backend_key}"
        report_path = report_root / cell / "group_report.json"
        if not report_path.is_file():
            return None, None, "missing report"
        report = json.loads(report_path.read_text(encoding="utf-8"))
        return report.get("aggregate", {}).get("mean"), None, f"report:{report_path.relative_to(PROJECT_ROOT)}"
    if family == "DAIC official development":
        report_root = PROJECT_ROOT / "outputs/experiment_reports/optuna100_officialdev"
        cell = f"daic_{modality_key}_{backend_key}"
        report_path = report_root / cell / "group_report.json"
        if not report_path.is_file():
            return None, None, "missing report"
        report = json.loads(report_path.read_text(encoding="utf-8"))
        return report.get("aggregate", {}).get("mean"), None, f"report:{report_path.relative_to(PROJECT_ROOT)}"
    dataset_key = {
        "D3TEC": "d3tec", "Androids Interview": "androids_interview",
        "CMDC": "cmdc", "Turkish t17": "turkish", "DAIC-WOZ": "daic",
    }[dataset]
    if family == "Native":
        report_root = PROJECT_ROOT / "outputs/experiment_reports/optuna100_native"
        cell = f"{dataset_key}_{modality_key}_{backend_key}"
        stage_key = None
    elif family == "English":
        report_root = PROJECT_ROOT / "outputs/experiment_reports/optuna100_english"
        cell = f"{dataset_key}_{modality_key}_{backend_key}"
        stage_key = None
    else:
        report_root = PROJECT_ROOT / "outputs/experiment_reports/optuna100_english"
        cell = f"{dataset_key}_{modality_key}_{backend_key}"
        stage_key = None
    report_path = report_root / cell / "group_report.json"
    if not report_path.is_file():
        return None, None, "missing report"
    report = json.loads(report_path.read_text(encoding="utf-8"))
    mean = report.get("aggregate", {}).get("mean")
    return mean, None, f"report:{report_path.relative_to(PROJECT_ROOT)}"


def build_optuna100_summary(wb: Workbook) -> None:
    """Standardized Optuna-100 XGBoost comparison sheet.

    Cells are populated from the qualified Optuna-100 group reports (runbook
    Tasks 9-12). The fixed XGBoost rows in the historical sheets remain
    available and labelled; this sheet is the new standardized comparison
    and never mixes the two.
    """
    ws = wb.create_sheet("Optuna-100 XGB")
    _widths(ws, {"A": 26, "B": 18, "C": 18, "D": 16, "E": 14, "F": 14, "G": 60})
    _title(ws, "Standardized Optuna-100 XGBoost — macro-F1", 7)
    _note(
        ws, 2,
        "Protocol harmonized_optuna100_v1: exactly 100 TPE trials, seeds 1337/1337/1337, "
        "3 subject-grouped inner folds, threshold 0.5, sampling none, pooled inner subject-level "
        "macro-F1 objective (merged: unweighted mean of five per-dataset inner macro-F1). "
        "Prediction backends qwen_hidden_xgb_optuna100 / gemma4_hidden_xgb_optuna100 "
        "(+ _symmetric_merged for merged). Qwen and Gemma paired cells share manifests, splits, "
        "weights, view (original_teacher_forced, harmonized_all_windows_full_coverage), namespace "
        "headline/binary_strict, and seeds. Values are fold-mean macro-F1 from the qualified "
        "group reports; the merged 'Final (DAIC official test)' rows are the single DAIC test "
        "fold. Fixed historical XGB rows remain in their original sheets, labelled, and are not "
        "replaced.",
        7, height=90,
    )
    _header_row(ws, 4, ["Family / Dataset", "Modality", "Backend", "Aggregation", "Macro-F1", "Positive-F1", "Evidence"])
    families = [
        ("Native", ("D3TEC", "Androids Interview", "CMDC", "Turkish t17", "DAIC-WOZ")),
        ("English", ("D3TEC", "Androids Interview", "CMDC", "Turkish t17")),
        ("Symmetric merged", ("CV (5-fold)", "Final (DAIC official test)")),
        ("DAIC official development", ("Official dev",)),
    ]
    row = 5
    for family, datasets in families:
        _section(ws, row, family, 7)
        row += 1
        for dataset in datasets:
            modalities = (
                ("Audio + Text", "Audio only", "Text only")
                if family == "Native"
                else ("Audio + Text", "Text only")
                if family == "English"
                else ("Audio + Text", "Audio only", "Text only")
                if family == "Symmetric merged"
                else ("Audio + Text", "Audio only", "Text only")
            )
            for modality in modalities:
                for backend in ("Qwen", "Gemma 4"):
                    ws.cell(row, 1, f"{family} — {dataset}").font = BODY_FONT
                    ws.cell(row, 1).fill = BODY
                    ws.cell(row, 1).alignment = LEFT
                    ws.cell(row, 1).border = BORDER
                    _body_cell(ws, row, 2, modality)
                    _body_cell(ws, row, 3, backend)
                    _body_cell(ws, row, 4, "fold-mean")
                    macro_f1, positive_f1, evidence = _optuna_report_values(family, dataset, modality, backend)
                    _body_cell(ws, row, 5, round(macro_f1, 4) if macro_f1 is not None else None, fmt="0.0000")
                    _body_cell(ws, row, 6, round(positive_f1, 4) if positive_f1 is not None else None, fmt="0.0000")
                    _body_cell(ws, row, 7, evidence or "missing report")
                    row += 1
    ws.freeze_panes = "A5"


def build_harmonized_en(wb: Workbook) -> None:
    ws = wb.create_sheet("EN Translation")
    _widths(ws, {c: 18 for c in "ABCDEFGHIJKLMN"})
    _title(ws, "English Translation vs Native — Macro-F1 and positive-F1", 14)
    campaign = HARMONIZED_EN_CAMPAIGN
    _note(
        ws, 2,
        f"Standard English-translation comparison: the default recipe with English transcript "
        f"overlay; only transcript language differs. "
        f"Campaign {campaign['campaign_id']}; source {campaign['source_sha'][:8]} (clean main); Issue "
        f"#{campaign['github_issue']} / PR #{campaign['github_pr']}; recipe {campaign['recipe_id']}; "
        f"translation cache {campaign['translation_cache_root']}. Teacher-forced, binary-strict, "
        f"best_model, macro-F1 checkpoint selection, audio encoder frozen. Native macro-F1 reproduces "
        f"the verified native campaign values exactly. Audio-only cells reuse the shared "
        f"native control (no new training).",
        14, height=88,
    )
    _header_row(ws, 4, [
        "Dataset", "Modality", "Method", "Native macro-F1", "EN macro-F1", "Δ macro-F1",
        "Native pos-F1", "EN pos-F1", "Δ pos-F1", "Aggregation", "Shared control",
        "Run / attempt identity", "Hashes (cfg/manifest/split/cache)", "Evidence path",
    ])
    row = 5
    methods = [("qwen", "Qwen"), ("logreg", "LogReg"), ("xgb_fixed", "XGBoost fixed")]
    en_ds_key = {"D3TEC": "d3tec", "Androids Interview": "androids_interview", "CMDC": "cmdc", "Turkish": "turkish"}
    en_mod_key = {"Audio + Text": "audio_text", "Audio only": "audio_only", "Text only": "text_only"}
    for dataset in ["D3TEC", "Androids Interview", "CMDC", "Turkish"]:
        for modality in ["Audio + Text", "Audio only", "Text only"]:
            key = (dataset, modality)
            if key not in HARMONIZED_EN_QWEN:
                continue
            nm, np1, em, ep1, agg, shared = HARMONIZED_EN_QWEN[key]
            nl, nx, el, ex = HARMONIZED_EN_HEADS[key]
            run = f"harmonized_v1_en_{campaign['campaign_id']}_{en_ds_key[dataset]}_{en_mod_key[modality]}"
            hashes = (
                f"{campaign['config_sha256'][dataset].get(modality, '-')}/"
                f"{campaign['manifest_sha256'][dataset]}/{campaign['split_sha256'][dataset]}/"
                f"{campaign['translation_cache_sha256'][dataset]}"
            )
            evidence = HARMONIZED_EN_EVIDENCE[dataset].replace("<run>", run)
            for method, label in methods:
                if method == "qwen":
                    native_m, en_m, native_p, en_p = nm, em, np1, ep1
                else:
                    native_p = en_p = None
                    if method == "logreg":
                        native_m, en_m = nl, el
                    else:
                        native_m, en_m = nx, ex
                ws.cell(row, 1, dataset).font = BODY_FONT
                ws.cell(row, 1).fill = BODY
                ws.cell(row, 1).alignment = LEFT
                ws.cell(row, 1).border = BORDER
                _body_cell(ws, row, 2, modality)
                _body_cell(ws, row, 3, label)
                _body_cell(ws, row, 4, native_m, fmt="0.0000")
                _body_cell(ws, row, 5, en_m, fmt="0.0000")
                _delta_cell(ws, row, 6, en_m - native_m)
                _body_cell(ws, row, 7, native_p, fmt="0.0000")
                _body_cell(ws, row, 8, en_p, fmt="0.0000")
                _delta_cell(ws, row, 9, (en_p - native_p) if (en_p is not None and native_p is not None) else None)
                cell = ws.cell(row, 10, agg)
                cell.font = BODY_FONT
                cell.alignment = LEFT
                cell.border = BORDER
                cell.fill = BODY
                cell = ws.cell(row, 11, "shared native audio-only control; not separately trained" if shared else "—")
                cell.font = BODY_FONT
                cell.alignment = LEFT
                cell.border = BORDER
                cell.fill = BODY
                cell = ws.cell(row, 12, run)
                cell.font = SMALL_FONT
                cell.alignment = WRAP
                cell.border = BORDER
                cell = ws.cell(row, 13, hashes)
                cell.font = SMALL_FONT
                cell.alignment = WRAP
                cell.border = BORDER
                cell = ws.cell(row, 14, evidence)
                cell.font = SMALL_FONT
                cell.alignment = WRAP
                cell.border = BORDER
                row += 1
    ws.freeze_panes = "A5"


def build_en_vs_native(wb: Workbook) -> None:
    ws = wb.create_sheet("EN vs Native MacroF1")
    _widths(ws, {"A": 30, "B": 18, "C": 18, "D": 16, "E": 16})
    _title(ws, "English Translation vs Native — Macro-F1 by Head (clean, verified)", 5)
    campaign = HARMONIZED_EN_CAMPAIGN
    _note(
        ws, 2,
        f"Values are `native / EN`. Macro-F1 only; higher is better. Teacher-forced, binary-strict, "
        f"best_model, macro-F1 checkpoint selection, audio encoder frozen. EN: campaign "
        f"{campaign['campaign_id']} (source {campaign['source_sha'][:8]}, Issue #{campaign['github_issue']} / "
        f"PR #{campaign['github_pr']}); D3TEC/Androids pooled 5-fold subject-level, CMDC/Turkish 5-fold mean. "
        f"Native: verified native campaign values (reproduced exactly). Direction = Δ (EN − native); ~tie for "
        f"|Δ| < 0.03. Audio-only cells reuse the shared native control (no EN run). XGBoost Optuna omitted: not run.",
        5, height=88,
    )
    heads = [("qwen", "Fine-tuned Qwen head"), ("logreg", "LogReg head"), ("xgb_fixed", "XGBoost fixed head")]

    def native_value(dataset: str, modality: str, method: str) -> float | None:
        if method == "qwen":
            return HARMONIZED_EN_QWEN[(dataset, modality)][0]
        head_idx = {"logreg": 0, "xgb_fixed": 1}[method]
        return HARMONIZED_EN_HEADS[(dataset, modality)][head_idx]

    def en_value(dataset: str, modality: str, method: str) -> float | None:
        shared = HARMONIZED_EN_QWEN[(dataset, modality)][5]
        if method == "qwen":
            return HARMONIZED_EN_QWEN[(dataset, modality)][2]
        head_idx = {"logreg": 0, "xgb_fixed": 1}[method]
        en_head_idx = {"logreg": 2, "xgb_fixed": 3}[method]
        return HARMONIZED_EN_HEADS[(dataset, modality)][en_head_idx]

    row = 4
    for method, label in heads:
        _section(ws, row, label, 5)
        _header_row(ws, row + 1, ["Evaluation / Modality", "Native / EN Macro-F1", "Δ Macro-F1", "Direction"])
        r = row + 2
        for dataset in ["D3TEC", "Androids Interview", "CMDC", "Turkish"]:
            for modality in ["Audio + Text", "Audio only", "Text only"]:
                native = native_value(dataset, modality, method)
                en = en_value(dataset, modality, method)
                if native is None or en is None:
                    continue
                delta = en - native
                if HARMONIZED_EN_QWEN[(dataset, modality)][5]:
                    direction = "shared control (no EN run)"
                elif abs(delta) < 0.03:
                    direction = "~tie"
                else:
                    direction = "EN better" if delta > 0 else "native better"
                ws.cell(r, 1, f"{dataset} — {modality}").font = BODY_FONT
                ws.cell(r, 1).fill = BODY
                ws.cell(r, 1).alignment = LEFT
                ws.cell(r, 1).border = BORDER
                _body_cell(ws, r, 2, f"{native:.4f} / {en:.4f}")
                _delta_cell(ws, r, 3, delta)
                dir_cell = ws.cell(r, 4, direction)
                dir_cell.font = BODY_FONT
                dir_cell.alignment = CENTER
                dir_cell.border = BORDER
                dir_cell.fill = _delta_fill(delta) or BODY
                r += 1
        row = r + 1
    ws.freeze_panes = "A5"


def build_merged_vs_standalone(wb: Workbook) -> None:
    ws = wb.create_sheet("Merged vs Standalone MacroF1")
    _widths(ws, {"A": 30, "B": 18, "C": 18, "D": 12, "E": 16})
    _title(ws, "Symmetric Merged vs Standalone — Macro-F1 by Head (clean, verified)", 5)
    _note(
        ws, 2,
        "Values are `standalone / merged`. Macro-F1 only; higher is better. Teacher-forced, binary-strict. "
        "Merged: DAIC = protected official holdout of the harmonized merged run "
        "harmonized_v1_prod_20260809T171705Z_d1e8130b (final epochs 2/4/5); other datasets = the same run's "
        "five-fold holdout mean. Standalone: Summary sheet (harmonized campaign). Direction = Δ (merged − "
        "standalone); ~tie for |Δ| < 0.03. XGBoost Optuna omitted: not run.",
        5, height=78,
    )
    heads = [("qwen", "Fine-tuned Qwen head"), ("logreg", "LogReg head"), ("xgb_fixed", "XGBoost fixed head")]

    def merged_value(dataset: str, modality: str, method: str) -> float:
        mod_key = {"Audio + Text": "audio_text", "Audio only": "audio_only", "Text only": "text_only"}[modality]
        ds_key = next(k for k, v in DATASET_LABELS.items() if v == dataset)
        if dataset == "DAIC":
            return MERGED_RUNS[mod_key]["official"].get(method)
        return MERGED_RUNS[mod_key]["cv"].get((ds_key, method))

    row = 4
    for method, label in heads:
        _section(ws, row, label, 5)
        _header_row(ws, row + 1, ["Evaluation / Modality", "Standalone / Merged Macro-F1", "Δ Macro-F1", "Direction"])
        r = row + 2
        for dataset in DATASETS:
            for modality in MODALITIES:
                if method == "qwen":
                    standalone = STANDALONE_QWEN[(dataset, modality)]
                else:
                    head_idx = {"logreg": 0, "xgb_fixed": 1}[method]
                    standalone = STANDALONE_HEADS[(dataset, modality)][head_idx]
                merged = merged_value(dataset, modality, method)
                if standalone is None or merged is None:
                    continue
                delta = merged - standalone
                if abs(delta) < 0.03:
                    direction = "~tie"
                else:
                    direction = "merged better" if delta > 0 else "standalone better"
                ws.cell(r, 1, f"{dataset} — {modality}").font = BODY_FONT
                ws.cell(r, 1).fill = BODY
                ws.cell(r, 1).alignment = LEFT
                ws.cell(r, 1).border = BORDER
                _body_cell(ws, r, 2, f"{standalone:.4f} / {merged:.4f}")
                _delta_cell(ws, r, 3, delta)
                dir_cell = ws.cell(r, 4, direction)
                dir_cell.font = BODY_FONT
                dir_cell.alignment = CENTER
                dir_cell.border = BORDER
                dir_cell.fill = _delta_fill(delta) or BODY
                r += 1
        row = r + 1
    ws.freeze_panes = "A4"


def build_provenance(wb: Workbook, *, detailed: bool) -> None:
    ws = wb.create_sheet("Provenance")
    widths = {"A": 13, "B": 11, "C": 14, "D": 12, "E": 13, "F": 52, "G": 40, "H": 30, "I": 40}
    _widths(ws, widths)
    _title(ws, "Provenance — every headline number maps to a run, aggregation, eval view, and artifact", 9)
    _header_row(
        ws, 2,
        ["Experiment", "Dataset", "Modality", "Method", "Macro-F1", "Source run / checkpoint",
         "Aggregation / eval view", "Local artifact", "Verification"],
    )
    row = 3

    def put(exp, dataset, modality, method, value, source, agg, artifact, verified):
        nonlocal row
        values = [exp, dataset, modality, method, value, source, agg, artifact, verified]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row, col, v)
            cell.font = SMALL_FONT
            cell.alignment = WRAP if col in (6, 7, 8, 9) else LEFT
            if isinstance(v, float):
                cell.number_format = "0.0000"
        row += 1

    for dataset in DATASETS:
        run, agg, artifact = STANDALONE_QWEN_SOURCE[dataset]
        for modality in MODALITIES:
            put("Standalone", dataset, modality, "Fine-tuned Qwen",
                STANDALONE_QWEN[(dataset, modality)], run, agg, artifact,
                "recomputed from local artifact")

    campaign = HARMONIZED_EN_CAMPAIGN
    en_ds_key = {"D3TEC": "d3tec", "Androids Interview": "androids_interview", "CMDC": "cmdc", "Turkish": "turkish"}
    en_mod_key = {"Audio + Text": "audio_text", "Audio only": "audio_only", "Text only": "text_only"}
    en_run_prefix = f"harmonized_v1_en_{campaign['campaign_id']}"
    for dataset, modality in HARMONIZED_EN_QWEN:
        nm, np1, em, ep1, agg, shared = HARMONIZED_EN_QWEN[(dataset, modality)]
        run = f"{en_run_prefix}_{en_ds_key[dataset]}_{en_mod_key[modality]}"
        hashes = (
            f"cfg {campaign['config_sha256'][dataset].get(modality, '-')} / "
            f"manifest {campaign['manifest_sha256'][dataset]} / "
            f"split {campaign['split_sha256'][dataset]} / "
            f"cache {campaign['translation_cache_sha256'][dataset]}"
        )
        evidence = HARMONIZED_EN_EVIDENCE[dataset].replace("<run>", run)
        if shared:
            source = f"shared native audio-only control (native harmonized run; not separately trained); EN cell reuses {run}"
        else:
            source = f"campaign {campaign['campaign_id']}, run {run}, folds 0-4, source {campaign['source_sha'][:8]}, Issue #{campaign['github_issue']} / PR #{campaign['github_pr']}"
        put("EN Translation", dataset, modality, "Fine-tuned Qwen", em, source,
            f"{agg}, teacher-forced, binary-strict, harmonized_all_windows_full_coverage",
            evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
            "recomputed from locally synced predictions; native macro-F1 reproduces the verified native value")
        put("EN Translation", dataset, modality, "Qwen positive-F1", ep1, source,
            f"{agg}, teacher-forced, binary-strict, harmonized_all_windows_full_coverage",
            evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
            "recomputed from locally synced predictions")
    for dataset, modality in HARMONIZED_EN_HEADS:
        nl, nx, el, ex = HARMONIZED_EN_HEADS[(dataset, modality)]
        run = f"{en_run_prefix}_{en_ds_key[dataset]}_{en_mod_key[modality]}"
        source = f"campaign {campaign['campaign_id']}, run {run}, folds 0-4, source {campaign['source_sha'][:8]}"
        evidence = f"outputs/hidden_classifiers/harmonized_v1_en/{en_ds_key[dataset]}/{run}/fold_<n>/variant_summary.json"
        for value, method in ((el, "LogReg head (EN)"), (ex, "XGBoost fixed (EN)")):
            put("EN heads", dataset, modality, method, value, source,
                "5-fold mean, subject-level macro-F1",
                evidence,
                "recomputed from locally synced variant_summary.json")
    for dataset, modality in HARMONIZED_EN_HEADS:
        nl, nx, el, ex = HARMONIZED_EN_HEADS[(dataset, modality)]
        run = f"{en_run_prefix}_{en_ds_key[dataset]}_{en_mod_key[modality]}"
        if modality == "Audio only":
            source = f"shared native audio-only control (native harmonized heads; not separately trained); EN cell reuses {run}"
        else:
            source = f"campaign {campaign['campaign_id']}, run {run}, folds 0-4, source {campaign['source_sha'][:8]}"
        for value, method in ((nl, "LogReg head (native)"), (nx, "XGBoost fixed (native)")):
            put("EN heads", dataset, modality, method, value, source,
                "5-fold mean, subject-level macro-F1 (STANDALONE_HEADS convention)",
                "outputs/hidden_classifiers/<dataset>/ (native campaign; verified 2026-08-10)",
                "matched to audited hidden-classifier outputs")
    for dataset in DATASETS:
        for modality in MODALITIES:
            logreg, xgb, optuna, os_ = STANDALONE_HEADS[(dataset, modality)]
            for value, method in ((logreg, "LogReg head"), (xgb, "XGBoost fixed")):
                if value is None:
                    continue
                put("Standalone heads", dataset, modality, method, value,
                    "hidden-state heads on frozen Qwen checkpoints (see docs)",
                    "pooled subject-level, 5-fold",
                    f"outputs/hidden_classifiers/{dataset.lower()}/ + docs/D3TEC_HIDDEN_CLASSIFIER_REPORT_2026-07-29.md / "
                    "reports/androids_hidden_classifier_*.md / outputs/daic_coverage_heads/ (MN5)",
                    "matched to audited hidden-classifier outputs")
            if detailed:
                for value, method in ((optuna, "XGBoost Optuna"), (os_, "XGBoost Subject OS (3-seed mean)")):
                    if value is None:
                        continue
                    if method.startswith("XGBoost Subject OS"):
                        put("Standalone heads", dataset, modality, method, value,
                            "Turkish subject-oversampling experiment, ratio 0.75, seeds 1337/2024/7",
                            "3-seed mean, 5-fold pooled",
                            "outputs/turkish_oversampling_* + docs/TURKISH_SUBJECT_OVERSAMPLING_REPORT_2026-07-25.md",
                            "mean of the three seed rows (internally consistent)")
                    else:
                        put("Standalone heads", dataset, modality, method, value,
                            "hidden-state heads on frozen Qwen checkpoints (see docs)",
                            "pooled subject-level, 5-fold",
                            f"outputs/hidden_classifiers/{dataset.lower()}/ + docs/D3TEC_HIDDEN_CLASSIFIER_REPORT_2026-07-29.md / "
                            "reports/androids_hidden_classifier_*.md / outputs/daic_coverage_heads/ (MN5)",
                            "matched to audited hidden-classifier outputs")

    merged_methods = ["qwen", "logreg", "xgb_fixed"] + (["xgb_optuna"] if detailed else [])

    for modality in ["audio_text", "audio_only", "text_only"]:
        run = MERGED_RUNS[modality]["run_id"]
        for method in merged_methods:
            value = MERGED_RUNS[modality]["official"].get(method)
            if value is None:
                put("Merged DAIC official", "DAIC", MODALITY_LABELS[modality], METHOD_LABELS[method], None,
                    run, "protected official holdout (47 subjects), full coverage",
                    f"outputs/symmetric_merged/{modality}/{run}/final/fold_0/",
                    "optuna not run on retrain checkpoints")
            else:
                put("Merged DAIC official", "DAIC", MODALITY_LABELS[modality], METHOD_LABELS[method], value,
                    run, "protected official holdout (47 subjects), full coverage",
                    f"outputs/symmetric_merged/{modality}/{run}/final/fold_0/",
                    "recomputed from prediction files")
        for (ds_key, method), value in sorted(MERGED_RUNS[modality]["cv"].items()):
            put("Merged CV", DATASET_LABELS[ds_key], MODALITY_LABELS[modality], METHOD_LABELS[method], value,
                run, "pooled subject-level 5-fold CV",
                f"outputs/symmetric_merged/{modality}/{run}/cv/fold_*/",
                "recomputed from prediction files")

    packed30_rows = [
        ("Packed30 v1 Audio + Text", "Qwen TF", 0.545,
         "daic_participant_p30_audio_text_s1337_05b52c6b (commit 3caa208)",
         "per-chunk mean teacher-forced score margin, official 47-subject test",
         "output_model/experiments/daic_participant_packed30/audio_text/<run>/fold_0/best_model/standalone_eval/"),
        ("Packed30 v1 Audio only", "Qwen TF", 0.468,
         "daic_participant_p30_audio_only_s1337_05b52c6b (commit 3caa208)",
         "per-chunk mean teacher-forced score margin, official 47-subject test",
         "output_model/experiments/daic_participant_packed30/audio_only/<run>/fold_0/best_model/standalone_eval/"),
        ("Joint-K4 Audio + Text", "Qwen TF", 0.7857,
         "daic_participant_p30_jointk4_audio_text_s1337_e3b0f1c3 (commit e3b0f1c; jobs 44365895/44365896/44365897/44365898, audit 44369723)",
         "per-bundle mean teacher-forced score margin, balanced K=4 coverage, official 47-subject test",
         "output_model/experiments/daic_participant_packed30_jointk4/audio_text/<run>/fold_0/best_model/standalone_eval/"),
        ("Joint-K4 Audio only", "Qwen TF", 0.4444,
         "daic_participant_p30_jointk4_audio_only_s1337_e3b0f1c3 (commit e3b0f1c; jobs 44365889/44365891/44365892/44365893, audit 44369722)",
         "per-bundle mean teacher-forced score margin, balanced K=4 coverage, official 47-subject test",
         "output_model/experiments/daic_participant_packed30_jointk4/audio_only/<run>/fold_0/best_model/standalone_eval/"),
    ]
    for condition, method, value, source, aggregation, artifact in packed30_rows:
        put("DAIC packed30 family", "DAIC", condition, method, value, source, aggregation,
            artifact + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
            "recomputed from subject rows (matches metrics JSON); audits 44369722/44369723 PASSED on GPFS + local evidence audit")

    packed30_head_rows = [
        ("Canonical Audio + Text", "LogReg raw", 0.7647, "outputs/daic_coverage_heads/daic_coverage_heads_20260805_04f2e19/audio_text/classical/c2/logreg_raw/"),
        ("Canonical Audio + Text", "XGBoost raw", 0.6429, "outputs/daic_coverage_heads/daic_coverage_heads_20260805_04f2e19/audio_text/classical/c2/xgb_raw/"),
        ("Canonical Audio only", "LogReg raw", 0.5714, "outputs/daic_coverage_heads/daic_coverage_heads_20260805_04f2e19/audio_only/classical/c2/logreg_raw/"),
        ("Canonical Audio only", "XGBoost raw", 0.3636, "outputs/daic_coverage_heads/daic_coverage_heads_20260805_04f2e19/audio_only/classical/c2/xgb_raw/"),
        ("Joint-K4 Audio + Text", "LogReg raw", 0.7333, "hidden_classifiers/audio_text/logreg_raw/"),
        ("Joint-K4 Audio + Text", "XGBoost raw", 0.7692, "hidden_classifiers/audio_text/xgb_raw/"),
        ("Joint-K4 Audio only", "LogReg raw", 0.6471, "hidden_classifiers/audio_only/logreg_raw/"),
        ("Joint-K4 Audio only", "XGBoost raw", 0.2000, "hidden_classifiers/audio_only/xgb_raw/"),
    ]
    for condition, method, value, artifact in packed30_head_rows:
        if condition.startswith("Canonical"):
            source = ("daic_coverage_heads_20260805_04f2e19 canonical checkpoints "
                      "(audio_text=daic_main_k4_control_20260804_f26dd45, audio_only=daic_replicates_20ep_s1337_daic_audio_only_selposf1_tf)")
            aggregation = "mean depressed probability >= 0.5 per subject, 47 test subjects, complete-coverage (c2_balanced) view"
        else:
            source = "daic_participant_p30_jointk4_<mod>_s1337_e3b0f1c3 selected-epoch head fit (107 vectors, weights 1.0)"
            aggregation = "mean depressed probability >= 0.5 per subject, 47 test subjects, 617 bundles"
        put("DAIC packed30 family", "DAIC", condition, method, value, source, aggregation,
            artifact, "recomputed from predictions_subject_level.jsonl (metrics match)")

    build_gemma4_provenance(ws, put)
    build_gemma_native_provenance(ws, put)
    build_optuna100_provenance(ws, put)
    build_en_merged_gemma_provenance(ws, put)
    build_daic_officialdev_provenance(ws, put)

    ws.freeze_panes = "A3"


def build_optuna100_provenance(ws, put) -> None:
    """Provenance for the standardized Optuna-100 XGBoost comparison sheet
    (both Qwen and Gemma), linking every displayed fold-mean to its group
    report and the registry attempt group."""
    backend_label = {"qwen": "Qwen", "gemma4": "Gemma 4"}
    run_ids = {
        "native": "optuna100_native_20260815T0330Z_99efc52",
        "english": "optuna100_english_20260815T2300Z_a955cdd",
        "merged": "optuna100_merged_20260816T2000Z_9efd6e5",
        "officialdev": "optuna100_officialdev_20260815T0330Z_99efc52",
    }
    ds_key = {"D3TEC": "d3tec", "Androids Interview": "androids_interview", "CMDC": "cmdc",
              "Turkish": "turkish", "DAIC": "daic"}
    mod_key = {"Audio + Text": "audio_text", "Audio only": "audio_only", "Text only": "text_only"}
    cells = []
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish", "DAIC"):
        for modality in MODALITIES:
            for backend in ("qwen", "gemma4"):
                cells.append(("native", dataset, modality, backend,
                              f"outputs/experiment_reports/optuna100_native/{ds_key[dataset]}_{mod_key[modality]}_{backend}/group_report.json"))
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish"):
        for modality in ("Audio + Text", "Text only"):
            for backend in ("qwen", "gemma4"):
                cells.append(("english", dataset, modality, backend,
                              f"outputs/experiment_reports/optuna100_english/{ds_key[dataset]}_{mod_key[modality]}_{backend}/group_report.json"))
    for modality in MODALITIES:
        for stage, stage_label in (("cv", "CV (5-fold)"), ("final", "Final (DAIC official test)")):
            for backend in ("qwen", "gemma4"):
                cells.append(("merged", stage_label, modality, backend,
                              f"outputs/experiment_reports/optuna100_merged/{mod_key[modality]}_{stage}_{backend}/group_report.json"))
    for modality in MODALITIES:
        for backend in ("qwen", "gemma4"):
            cells.append(("officialdev", "DAIC", modality, backend,
                          f"outputs/experiment_reports/optuna100_officialdev/daic_{mod_key[modality]}_{backend}/group_report.json"))
    agg = ("Optuna-100 fold-mean, 100 TPE trials seed 1337, 3 subject-grouped inner folds, "
           "macro-F1 objective, harmonized_all_windows_full_coverage, headline/binary-strict")
    for family, dataset, modality, backend, report in cells:
        report_path = PROJECT_ROOT / report
        if not report_path.is_file():
            continue
        value = json.loads(report_path.read_text(encoding="utf-8")).get("aggregate", {}).get("mean")
        source = (
            f"campaign {run_ids[family]}, {backend_label[backend]} backend, 100 trials/fold, "
            f"all attempts REPORTABLE in registry (group {family}-optuna100-{run_ids[family]})"
        )
        put(f"Optuna100 {family}", dataset, modality,
            f"XGBoost Optuna-100 ({backend_label[backend]})", value, source, agg,
            report, "group report OK; attempts REPORTABLE")


def build_en_merged_gemma_provenance(ws, put) -> None:
    """Provenance for the English-translated and merged Gemma cells (teacher-
    forced and LogReg) in the Qwen vs Gemma sheet. The XGBoost cells are
    covered by the Optuna100 provenance rows."""
    ds_key = {"D3TEC": "d3tec", "Androids Interview": "androids_interview", "CMDC": "cmdc", "Turkish": "turkish"}
    mod_key = {"Audio + Text": "audio_text", "Text only": "text_only"}
    en_campaign = "gemma4_harmonized_v1_en_gemma4_en_prod_20260815T1300Z_a955cdd"
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish"):
        for modality, mk in mod_key.items():
            run = f"{en_campaign}_{ds_key[dataset]}_{mk}"
            source = f"campaign {en_campaign}, group gemma4-harmonized-v1-en-{en_campaign}, 5 folds REPORTABLE"
            agg = ("English-translated, pooled 5-fold subject-level (D3TEC/Androids) or 5-fold mean "
                   "(CMDC/Turkish), teacher-forced, binary-strict, harmonized_all_windows_full_coverage")
            put("Gemma EN", dataset, modality, "Fine-tuned Gemma (teacher-forced)",
                EN_TF[(dataset, modality)][1], source, agg,
                f"output_model/harmonized_v1_en_gemma4/{mk}/{ds_key[dataset]}/{run}/fold_*/best_model/standalone_eval",
                "recomputed from local predictions (registry REPORTABLE)")
            lr_agg = agg.replace("teacher-forced", "LogReg raw hidden head")
            put("Gemma EN", dataset, modality, "Gemma LogReg raw hidden head",
                EN_LR[(dataset, modality)][1], source, lr_agg,
                f"outputs/experiment_reports/gemma4_harmonized/english_lr/{ds_key[dataset]}_{mk}.json",
                "recomputed from predictions_subject_level.csv (matches metrics.json)")
    merged_campaign = "gemma4_merged_v1_prod_20260816T0000Z_d4ff33e"
    mmod_key = {"Audio + Text": "audio_text", "Audio only": "audio_only", "Text only": "text_only"}
    merged_postprocess_jobs = {"audio_text": "44684476", "audio_only": "44684479", "text_only": "44684482"}
    for stage, stage_label in (("cv", "CV (5-fold)"), ("final", "Final (DAIC test)")):
        for modality, mk in mmod_key.items():
            q, g = MERGED_TF[(stage, modality)]
            if g is None:
                continue
            if stage == "final":
                put("Gemma merged", stage_label, modality, "Teacher-forced",
                    g, (f"campaign {merged_campaign}, postprocess job {merged_postprocess_jobs[mk]}, "
                        "source bfc13b4f8b177547a11a9abad526115b712d32ef"),
                    ("DAIC official test, 47 subjects, teacher-forced, binary-strict, subject mean-score, "
                     "harmonized_all_windows_full_coverage"),
                    (f"outputs/symmetric_merged/gemma4/harmonized_v1/{mk}/{merged_campaign}/final/fold_0/"
                     "gemma4/daic/metrics_original_teacher_forced.json + predictions_subject_level.csv"),
                    "recomputed locally from 47 subject predictions; zero invalid subjects")
            else:
                put("Gemma merged", stage_label, modality, "Teacher-forced",
                    g, f"campaign {merged_campaign}, merged training selection (mean_dataset_macro_f1)",
                    f"{stage_label}, teacher-forced, mean over five datasets",
                    f"output_model/symmetric_merged/gemma4/harmonized_v1/{mk}/{merged_campaign}/{stage}/fold_0/logs/training_history.json",
                    "from local training_history selected epoch")
            q, g = MERGED_LR[(stage, modality)]
            put("Gemma merged", stage_label, modality, "LogReg head",
                g, f"campaign {merged_campaign}, merged heads",
                f"{stage_label}, LogReg raw hidden head, mean over five datasets",
                f"outputs/symmetric_merged/gemma4/harmonized_v1/{mk}/{merged_campaign}/{stage}/fold_0/heads/logreg/metrics_by_dataset.json",
                "recomputed from predictions (matches metrics_by_dataset.json)")


def build_daic_officialdev_provenance(ws, put) -> None:
    """Provenance rows for the DAIC official-development campaign: six
    teacher-forced evaluations and twelve fixed-head evaluations on the 35
    official development subjects."""
    campaign = DAIC_OFFICIALDEV_CAMPAIGN
    base_source = (
        f"campaign {campaign['campaign_id']}, group {campaign['group_id']}, "
        f"source {campaign['source_sha'][:8]} (clean main), Issue #{campaign['github_issue']} "
        f"/ PR #{campaign['github_pr']}, manifest {campaign['manifest_sha256'][:12]}… "
        f"/ split {campaign['split_sha256'][:12]}…"
    )
    tf_agg = (
        "official development partition (35 subjects), subject mean-score, "
        "teacher-forced, binary-strict, harmonized_all_windows_full_coverage, "
        "daic_official_train_inner_split_dev_evaluation"
    )
    for modality in ("audio_only", "audio_text", "text_only"):
        for backbone in ("qwen", "gemma4"):
            value = DAIC_OFFICIALDEV_TEACHER_FORCED[(modality, backbone)]
            evidence = TF_EVIDENCE[(modality, backbone)]
            method = "Qwen teacher-forced" if backbone == "qwen" else "Gemma 4 teacher-forced"
            source = (
                f"{base_source}, attempt {value['attempt_id']}, eval {value['evaluation_id']}, "
                f"selected epoch {value['selected_epoch']}, best_model"
            )
            put("DAIC official development", "DAIC", MODALITY_LABELS[modality], method + " macro-F1",
                value["macro_f1"], source, tf_agg,
                evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
                "recomputed locally from synced predictions; REPORTABLE")
            put("DAIC official development", "DAIC", MODALITY_LABELS[modality], method + " positive-F1",
                value["positive_f1"], source, tf_agg,
                evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
                "recomputed locally from synced predictions; REPORTABLE")
            put("DAIC official development", "DAIC", MODALITY_LABELS[modality], method + " accuracy",
                value["accuracy"], source, tf_agg,
                evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
                "recomputed locally from synced predictions; REPORTABLE")
    heads_agg = (
        "official development partition (35 subjects), subject mean probability >= 0.5, "
        "binary-strict, harmonized_all_windows_full_coverage, "
        "daic_official_train_inner_split_dev_evaluation"
    )
    for modality in ("audio_only", "audio_text", "text_only"):
        for backbone in ("qwen", "gemma4"):
            for variant, variant_label in (
                ("logreg", "LogReg raw hidden head"),
                ("xgb", "XGBoost raw hidden head"),
            ):
                head = DAIC_OFFICIALDEV_HEADS[(modality, backbone, variant)]
                method = ("Qwen " if backbone == "qwen" else "Gemma 4 ") + variant_label
                head_evidence = HEAD_EVIDENCE[(modality, backbone)] + head["run_name"] + "/fold_0/hidden_classifiers/" + variant + "/"
                source = (
                    f"{base_source}, attempt {head['attempt_id']}, eval {head['evaluation_id']}, "
                    f"backend {head['backend']}, parent best_model"
                )
                for metric_key, metric_label in (
                    ("macro_f1", "macro-F1"),
                    ("positive_f1", "positive-F1"),
                    ("accuracy", "accuracy"),
                    ("precision", "precision"),
                    ("recall", "recall"),
                ):
                    put("DAIC Head Ablation", "DAIC", MODALITY_LABELS[modality],
                        f"{method} {metric_label}", head[metric_key], source, heads_agg,
                        head_evidence + "metrics.json + predictions_subject_level.csv",
                        "recomputed locally by verify-local; REPORTABLE")


def build_gemma_native_provenance(ws, put) -> None:
    """Provenance for the native Gemma 4 harmonized campaign cells added to the
    Summary and Gemma vs Qwen sheets (teacher-forced, LogReg, and Optuna-100
    XGBoost). DAIC rows are covered by build_gemma4_provenance; here we add the
    non-DAIC cells with their verified local report paths.
    """
    campaign_id = "gemma4_v1_prod_20260814T2030Z_1ab337d2_r2"
    group_id = "gemma4-harmonized-v1-gemma4_v1_prod_20260814T2030Z_1ab337d2_r2"
    ds_key = {"D3TEC": "d3tec", "Androids Interview": "androids_interview", "CMDC": "cmdc", "Turkish": "turkish"}
    mod_key = {"Audio + Text": "audio_text", "Audio only": "audio_only", "Text only": "text_only"}
    for dataset in ("D3TEC", "Androids Interview", "CMDC", "Turkish"):
        for modality in MODALITIES:
            m = mod_key[modality]
            if dataset == "CMDC" or dataset == "Turkish":
                agg = "5-fold mean (train_val protocol), teacher-forced, binary-strict, harmonized_all_windows_full_coverage"
            else:
                agg = "pooled 5-fold subject-level, teacher-forced, binary-strict, harmonized_all_windows_full_coverage"
            source = f"campaign {campaign_id}, group {group_id}, seed 1337, 5 folds REPORTABLE (registry); model google/gemma-4-12B-it rev 707f0a3b8a3c7ad586ed01e27eafbad8a27dd0f7"
            tf_report = GEMMA_NATIVE_EVIDENCE["tf"].replace("{ds}", ds_key[dataset]).replace("{mod}", m)
            put("Gemma native", dataset, modality, "Fine-tuned Gemma (teacher-forced)",
                GEMMA_NATIVE_TF[(dataset, modality)], source, agg, tf_report,
                "recomputed from local predictions (matches metrics JSON); group report OK")
            lr_report = GEMMA_NATIVE_EVIDENCE["lr"].replace("{ds}", ds_key[dataset]).replace("{mod}", m)
            lr_agg = agg.replace("teacher-forced", "LogReg raw hidden head")
            put("Gemma native", dataset, modality, "Gemma LogReg raw hidden head",
                GEMMA_NATIVE_LR[(dataset, modality)], source, lr_agg, lr_report,
                "recomputed from predictions_subject_level.csv (matches metrics.json)")
            optuna_report = GEMMA_NATIVE_EVIDENCE["optuna"].replace("{ds}", ds_key[dataset]).replace("{mod}", m).replace("{backend}", "gemma4")
            optuna_agg = "Optuna-100 fold-mean, 100 trials seed 1337, 3 inner folds, subject-level macro-F1 objective"
            put("Gemma native", dataset, modality, "Gemma XGBoost Optuna-100 raw hidden head",
                GEMMA_OPTUNA[(dataset, modality)], source, optuna_agg, optuna_report,
                "group report OK; attempt REPORTABLE in registry")


def build_gemma4_provenance(ws, put) -> None:
    campaign = GEMMA4_CAMPAIGN
    for mod_key, mod_label in zip(
        ["audio_text", "audio_only", "text_only"],
        ["Audio + Text", "Audio only", "Text only"],
    ):
        run, attempt, train_job, eval_job, epoch = GEMMA4_RUNS[mod_key]
        g_macro, g_pos, g_acc, g_prec, g_rec, g_cm = GEMMA4_QWEN[mod_key]
        source = (
            f"campaign {campaign['campaign_id']}, run {run}, attempt {attempt}, "
            f"source {campaign['source_sha'][:8]} (clean main), PR #{campaign['github_pr']}, "
            f"jobs {train_job}/{eval_job}, selected epoch {epoch if epoch is not None else 'n/a'}, "
            f"model {campaign['model']} rev {campaign['revision'][:12]}…, "
            f"cfg/manifest/split {campaign['manifest_sha256'][:12]}…/{campaign['split_sha256'][:12]}…"
        )
        if mod_key == "audio_text":
            source += (
                "; first attempt …-cca3f4ae-5704f1f7 FAILED (train 44517567, eval 44517568 "
                "CANCELLED), superseded by retry …-a6749b05-146c8805"
            )
        agg = "official 47-subject test, subject mean-score, teacher-forced, binary-strict, harmonized_all_windows_full_coverage"
        evidence = GEMMA4_EVIDENCE.replace("<modality>", mod_key).replace("<run>", run)
        put("Gemma 4 DAIC", "DAIC", mod_label, "Gemma 4 macro-F1", g_macro, source, agg,
            evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
            "recomputed locally from predictions (matches metrics JSON); Slurm COMPLETED 0:0")
        put("Gemma 4 DAIC", "DAIC", mod_label, "Gemma 4 positive-F1", g_pos, source, agg,
            evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
            "recomputed locally from predictions (matches metrics JSON)")
        put("Gemma 4 DAIC", "DAIC", mod_label, "Gemma 4 accuracy", g_acc, source, agg,
            evidence + "metrics_original_teacher_forced.json + predictions_subject_level.csv",
            "recomputed locally from predictions (matches metrics JSON)")
        put("Gemma 4 DAIC", "DAIC", mod_label, "Gemma 4 confusion matrix", str(g_cm), source, agg,
            evidence + "confusion_matrix.json",
            "recomputed locally from predictions (matches metrics JSON)")
        put("Gemma 4 DAIC", "DAIC", mod_label, "Gemma 4 invalid count", 0, source, agg,
            evidence + "metrics_original_teacher_forced.json",
            "invalid_subjects == 0 in metrics JSON and recomputation")
        heads_campaign = GEMMA4_HEADS_CAMPAIGN
        heads_run, heads_attempt, heads_extract_job, heads_job, superseded = GEMMA4_HEADS_RUNS[mod_key]
        heads_agg = (
            "official 47-subject test, subject mean probability >= 0.5, binary-strict, "
            "harmonized_all_windows_full_coverage, daic_official_train_fit_locked_test_evaluation"
        )
        heads_base = (
            f"campaign {heads_campaign['campaign_id']}, run {heads_run}, attempt {heads_attempt}, "
            f"source {heads_campaign['source_sha'][:8]} (clean main), PR #{heads_campaign['github_pr']}, "
            f"jobs {heads_extract_job}/{heads_job} (COMPLETED 0:0), parent attempt {attempt}, "
            f"parent adapter {GEMMA4_CAMPAIGN['manifest_sha256'][:12]}…/… split "
            f"{heads_campaign['split_sha256'][:12]}…, model {heads_campaign['model']} rev "
            f"{heads_campaign['revision'][:12]}…; supersedes {superseded} "
            f"(extract FAILED 1:0, heads CANCELLED, wrapper race fixed in PR #45)"
        )
        heads_evidence = GEMMA4_HEADS_EVIDENCE.replace("<modality>", mod_key).replace("<run>", heads_run)
        for variant, variant_label in (("logreg", "Gemma 4 LogReg raw hidden head"), ("xgb", "Gemma 4 XGBoost raw hidden head")):
            h_macro, h_pos, h_acc, h_prec, h_rec, h_cm = GEMMA4_HEADS[mod_key][variant]
            backend = GEMMA4_HEADS_BACKEND[variant]
            v_source = heads_base + f", backend {backend}"
            v_evidence = heads_evidence.replace("<variant>", variant)
            put("Gemma 4 DAIC", "DAIC", mod_label, f"{variant_label} macro-F1", h_macro, v_source, heads_agg,
                v_evidence + "metrics.json + predictions_subject_level.csv",
                "recomputed locally by verify-local (matches metrics JSON); REPORTABLE")
            put("Gemma 4 DAIC", "DAIC", mod_label, f"{variant_label} positive-F1", h_pos, v_source, heads_agg,
                v_evidence + "metrics.json + predictions_subject_level.csv",
                "recomputed locally by verify-local (matches metrics JSON)")
            put("Gemma 4 DAIC", "DAIC", mod_label, f"{variant_label} accuracy", h_acc, v_source, heads_agg,
                v_evidence + "metrics.json + predictions_subject_level.csv",
                "recomputed locally by verify-local (matches metrics JSON)")
            put("Gemma 4 DAIC", "DAIC", mod_label, f"{variant_label} precision", h_prec, v_source, heads_agg,
                v_evidence + "metrics.json + predictions_subject_level.csv",
                "recomputed locally by verify-local (matches metrics JSON)")
            put("Gemma 4 DAIC", "DAIC", mod_label, f"{variant_label} recall", h_rec, v_source, heads_agg,
                v_evidence + "metrics.json + predictions_subject_level.csv",
                "recomputed locally by verify-local (matches metrics JSON)")
            put("Gemma 4 DAIC", "DAIC", mod_label, f"{variant_label} confusion matrix", str(h_cm), v_source, heads_agg,
                v_evidence + "metrics.json",
                "recomputed locally by verify-local (matches metrics JSON)")


# --------------------------------------------------------------------------- data
# DAIC participant-packed30 family (runtime participant-only chunks, seed 1337):
#   packed30_v1  = one chunk per prompt, all chunks subject-normalized (commit 3caa208)
#   packed30_jointk4 = joint K=4 random bundle per subject/epoch, balanced-cover
#                     (commit e3b0f1c, run daic_participant_p30_jointk4_<mod>_s1337_e3b0f1c3)
# Values are strict teacher-forced positive-F1 on the official 47-subject test
# (mean per-subject margin; INVALID counts as wrong; INVALID=0). Verified
# locally by recomputation from predictions_subject_level.csv.
PACKED30_QWEN: dict[tuple[str, str], float] = {
    ("DAIC", "Packed30 v1 Audio + Text"): 0.545,
    ("DAIC", "Packed30 v1 Audio only"): 0.468,
    ("DAIC", "Joint-K4 Audio + Text"): 0.7857,
    ("DAIC", "Joint-K4 Audio only"): 0.4444,
}
PACKED30_HEAD_POSF1: dict[tuple[str, str, str], float] = {
    ("DAIC", "Packed30 v1 Audio + Text", "LogReg"): 0.741,
    ("DAIC", "Packed30 v1 Audio + Text", "XGB"): 0.583,
    ("DAIC", "Packed30 v1 Audio only", "LogReg"): 0.333,
    ("DAIC", "Packed30 v1 Audio only", "XGB"): 0.435,
    ("DAIC", "Joint-K4 Audio + Text", "LogReg"): 0.7333,
    ("DAIC", "Joint-K4 Audio + Text", "XGB"): 0.7692,
    ("DAIC", "Joint-K4 Audio only", "LogReg"): 0.6471,
    ("DAIC", "Joint-K4 Audio only", "XGB"): 0.2000,
}
PACKED30_SOURCE = {
    "Packed30 v1 Audio + Text": ("daic_participant_p30_audio_text_s1337_05b52c6b (commit 3caa208)",
                                 "output_model/experiments/daic_participant_packed30/audio_text/"),
    "Packed30 v1 Audio only": ("daic_participant_p30_audio_only_s1337_05b52c6b (commit 3caa208)",
                               "output_model/experiments/daic_participant_packed30/audio_only/"),
    "Joint-K4 Audio + Text": ("daic_participant_p30_jointk4_audio_text_s1337_e3b0f1c3 (commit e3b0f1c)",
                              "output_model/experiments/daic_participant_packed30_jointk4/audio_text/"),
    "Joint-K4 Audio only": ("daic_participant_p30_jointk4_audio_only_s1337_e3b0f1c3 (commit e3b0f1c)",
                            "output_model/experiments/daic_participant_packed30_jointk4/audio_only/"),
}

def build_packed30(wb: Workbook) -> None:
    ws = wb.create_sheet("DAIC Packed30 Family")
    _widths(ws, {"A": 34, "B": 12, "C": 14, "D": 14, "E": 12, "F": 52, "G": 46})
    _title(ws, "DAIC runtime participant-packed30: canonical vs v1 vs joint-K4 (official 47-subject test, seed 1337, strict teacher-forced)", 7)
    _note(ws, 2, "Qwen verdict = mean teacher-forced score margin per subject (INVALID counts as wrong; INVALID=0). "
                 "Heads = logreg_raw / xgb_raw on hidden features, mean depressed probability >= 0.5 per subject. "
                 "Canonical heads = daic_coverage_heads_20260805_04f2e19 complete-coverage (c2_balanced) view. "
                 "Selected epochs: joint-K4 audio-only 5, audio+text 3; v1 audio-only 3, audio+text 2.", 7)
    _header_row(ws, 3, ["Condition", "Method", "Positive-F1", "Macro-F1", "AUROC (heads)", "Canonical joint-K4 (pos / macro / AUROC)", "Source run", "Local artifact"])
    row = 4
    canonical = {
        "Audio + Text": {"qwen": (0.800, 0.841, None), "logreg": (0.7647, 0.8157, 0.9177), "xgb": (0.6429, 0.7457, 0.8831)},
        "Audio only": {"qwen": (0.522, 0.683, None), "logreg": (0.5714, 0.6586, 0.8225), "xgb": (0.3636, 0.5846, 0.7814)},
    }
    conditions = [
        ("Packed30 v1 Audio + Text", "Audio + Text", {"qwen_pos": 0.545, "qwen_macro": 0.703, "logreg": (0.741, 0.818, 0.846), "xgb": (0.583, 0.720, 0.908)}),
        ("Packed30 v1 Audio only", "Audio only", {"qwen_pos": 0.468, "qwen_macro": 0.468, "logreg": (0.333, 0.510, 0.617), "xgb": (0.435, 0.626, 0.602)}),
        ("Joint-K4 Audio + Text", "Audio + Text", {"qwen_pos": 0.7857, "qwen_macro": 0.8474, "logreg": (0.7333, 0.8042, 0.9156), "xgb": (0.7692, 0.8405, 0.9134)}),
        ("Joint-K4 Audio only", "Audio only", {"qwen_pos": 0.4444, "qwen_macro": 0.5498, "logreg": (0.6471, 0.7235, 0.8074), "xgb": (0.2000, 0.4919, 0.7814)}),
    ]
    for condition, modality_key, values in conditions:
        source, artifact_root = PACKED30_SOURCE[condition]
        canon = canonical[modality_key]
        for method, value in (
            ("Qwen TF", (values["qwen_pos"], values["qwen_macro"], None)),
            ("LogReg raw", values["logreg"]),
            ("XGBoost raw", values["xgb"]),
        ):
            pos, macro, auroc = value
            canon_key = "qwen" if method == "Qwen TF" else "logreg" if method == "LogReg raw" else "xgb"
            canon_values = canon[canon_key]
            _body_cell(ws, row, 1, condition)
            _body_cell(ws, row, 2, method)
            _body_cell(ws, row, 3, pos, fmt="0.0000")
            _body_cell(ws, row, 4, macro, fmt="0.0000")
            _body_cell(ws, row, 5, auroc, fmt="0.0000")
            _body_cell(ws, row, 6, " / ".join("n/a" if v is None else f"{v:0.3f}" for v in canon_values))
            _body_cell(ws, row, 7, source)
            suffix = "logreg_raw/metrics.json" if method == "LogReg raw" else "xgb_raw/metrics.json" if method == "XGBoost raw" else "standalone_eval/metrics_original_teacher_forced.json"
            _body_cell(ws, row, 8, f"{artifact_root}<run>/fold_0/best_model/{suffix} (recomputed from predictions)")
            row += 1
        row += 1
    _note(ws, row, "Canonical = preprocessed joint-K4 (Qwen: 2026-08-05 coverage validation; heads: "
                   "daic_coverage_heads_20260805_04f2e19 c2_balanced, verified locally by recomputation). "
                   "Provenance sheet holds the full mapping (commit, job IDs incl. resubmits, eval view, aggregation, "
                   "audits 44369722/44369723 GPFS PASSED + local evidence audit PASSED).", 7)
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------- main
# --------------------------------------------------------------------------- validation
def validate_selected_results(selected_results: Path, cell_values: dict[tuple[str, str], float | None]) -> tuple[bool, list[str]]:
    """Cross-check an explicit selected-results export against workbook headline cells.

    The workbook stays script-only: values are compared, never overwritten. Missing
    records are listed as legacy-unmigrated; nothing is invented and nothing is zeroed.
    """
    payload = json.loads(Path(selected_results).read_text(encoding="utf-8"))
    selections = payload.get("selections", [])
    legacy_unmigrated: list[str] = []
    mismatches: list[str] = []
    checked = 0
    for selection in selections:
        cell = str(selection.get("cell"))
        dataset_label, modality_label = cell.split("|", 1)
        key = (dataset_label.strip(), modality_label.strip())
        status = selection.get("status")
        if status != "selected":
            legacy_unmigrated.append(f"{cell}: {status} ({selection.get('reason', 'no record')})")
            continue
        expected = cell_values.get(key)
        if expected is None:
            legacy_unmigrated.append(f"{cell}: no workbook cell for selected record")
            continue
        value = selection.get("value")
        if value is None:
            mismatches.append(f"{cell}: selected record has null value")
            continue
        if expected is not None and abs(float(value) - float(expected)) > 1e-6 * max(1.0, abs(float(expected))):
            mismatches.append(
                f"{cell}: registry {float(value):.6f} differs from workbook {float(expected):.6f}"
            )
        checked += 1
    report = [
        *mismatches,
        *[f"legacy-unmigrated: {entry}" for entry in legacy_unmigrated],
    ]
    return (not mismatches, report)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--detailed", action="store_true",
                        help="include XGBoost Optuna and Subject-OS columns/rows; write the detailed workbook")
    parser.add_argument("--validate-selected", default=None,
                        help="cross-check an explicit selected-results JSON against workbook cells; never rewrites cells")
    args = parser.parse_args()
    detailed = args.detailed

    if args.validate_selected is not None:
        cell_values: dict[tuple[str, str], float | None] = {
            (dataset_label, modality_label): value
            for (dataset_label, modality_label), value in STANDALONE_QWEN.items()
        }
        # Gemma 4 DAIC macro-F1 headline cells (Gemma 4 DAIC sheet).
        for mod_key, mod_label in zip(
            ["audio_text", "audio_only", "text_only"],
            ["Audio + Text", "Audio only", "Text only"],
        ):
            cell_values[("Gemma 4 DAIC", mod_label)] = GEMMA4_QWEN[mod_key][0]
            for variant in ("logreg", "xgb"):
                variant_label = (
                    "Gemma 4 LogReg raw hidden head"
                    if variant == "logreg"
                    else "Gemma 4 XGBoost raw hidden head"
                )
                cell_values[("Gemma 4 DAIC", f"{mod_label} — {variant_label}")] = (
                    GEMMA4_HEADS[mod_key][variant][0]
                )
        # DAIC official-development cells (DAIC LLM Comparison and DAIC Head
        # Ablation sheets): six teacher-forced and twelve fixed-head macro-F1.
        for modality in ("audio_only", "audio_text", "text_only"):
            for backbone in ("qwen", "gemma4"):
                cell_values[
                    ("DAIC official development", f"{MODALITY_LABELS[modality]} — {BACKBONE_LABELS[backbone]}")
                ] = DAIC_OFFICIALDEV_TEACHER_FORCED[(modality, backbone)]["macro_f1"]
                for variant in ("logreg", "xgb"):
                    variant_label = (
                        "Qwen LogReg raw hidden head"
                        if backbone == "qwen" and variant == "logreg"
                        else "Qwen XGBoost raw hidden head"
                        if backbone == "qwen"
                        else "Gemma 4 LogReg raw hidden head"
                        if variant == "logreg"
                        else "Gemma 4 XGBoost raw hidden head"
                    )
                    cell_values[("DAIC Head Ablation", f"{MODALITY_LABELS[modality]} — {variant_label}")] = (
                        DAIC_OFFICIALDEV_HEADS[(modality, backbone, variant)]["macro_f1"]
                    )
        ok, report = validate_selected_results(Path(args.validate_selected), cell_values)
        print("\n".join(report))
        print(f"checked={sum(1 for s in json.loads(Path(args.validate_selected).read_text())['selections'] if s['status'] == 'selected')} "
              f"mismatches={sum(1 for line in report if 'differs from' in line)} "
              f"legacy_unmigrated={sum(1 for line in report if line.startswith('legacy-unmigrated:'))}")
        return 0 if ok else 1

    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, detailed=detailed)
    build_gemma_vs_qwen(wb)
    build_native_vs_english(wb)
    build_packed30(wb)
    build_provenance(wb, detailed=detailed)
    out = OUT_DETAILED if detailed else OUT
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    sys.exit(main())
