#!/usr/bin/env python3
"""Update the tracked Androids rows, highlights, notes, and positive-F1 chart."""

from __future__ import annotations

import argparse
import copy
import json
import os
import tempfile
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill


MODALITIES = ("audio_only", "audio_text", "text_only")
MODALITY_LABELS = {
    "audio_only": "Audio only",
    "audio_text": "Audio + Text",
    "text_only": "Text only",
}
HEADS = ("logreg_raw", "xgb_raw", "xgb_optuna_150t_d6")
HEAD_LABELS = {
    "logreg_raw": "Logistic Regression",
    "xgb_raw": "XGBoost (fixed raw)",
    "xgb_optuna_150t_d6": "XGBoost Optuna (150t d6, seed1337)",
}
ANDROID_ROWS = {"audio_only": 18, "audio_text": 19, "text_only": 20}
ANDROID_GRAPH_ROWS = {"audio_only": 5, "audio_text": 6, "text_only": 7}
BASELINE_F1 = 128.0 / 180.0
ORANGE = "FFFFF4E5"
GREEN = "00E2F0D9"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--acceptance", required=True, type=Path)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("depression_results_combined_with_posf1_graphs.xlsx"),
    )
    parser.add_argument("--output", type=Path)
    return parser.parse_args()


def _load_acceptance(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("status") != "passed" or payload.get("mode") != "production":
        raise ValueError("Workbook update requires a passed production Androids acceptance audit.")
    pooled = payload.get("pooled_results")
    expected = {f"{modality}/{head}" for modality in MODALITIES for head in HEADS}
    if not isinstance(pooled, dict) or set(pooled) != expected:
        raise ValueError("Acceptance audit does not contain exactly the nine Androids pooled head results.")
    if payload.get("counts", {}).get("pooled_subjects_per_result") != 116:
        raise ValueError("Acceptance audit does not cover the required 116 pooled subjects.")
    return payload


def _metrics(payload: dict[str, Any], modality: str, head: str) -> dict[str, Any]:
    result = payload["pooled_results"][f"{modality}/{head}"]
    metrics = result["metrics"]
    required = (
        "accuracy",
        "positive_f1",
        "precision",
        "recall",
        "macro_f1",
        "negative_f1",
        "auroc",
        "confusion_matrix",
    )
    if any(key not in metrics for key in required):
        raise ValueError(f"Missing Androids metric in {modality}/{head}.")
    return metrics


def _copy_row_style(ws: Any, source_row: int, target_row: int, max_col: int) -> None:
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, max_col + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)


def _set_fill(cell: Any, color: str) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=color)


def _reset_fills(ws: Any, rows: range, columns: range, color: str = ORANGE) -> None:
    for row in rows:
        for column in columns:
            _set_fill(ws.cell(row, column), color)


def _highlight_maxima(ws: Any, rows: list[int], columns: list[int]) -> None:
    for column in columns:
        numeric = [
            (row, float(ws.cell(row, column).value))
            for row in rows
            if isinstance(ws.cell(row, column).value, (int, float))
        ]
        if not numeric:
            continue
        maximum = max(value for _, value in numeric)
        for row, value in numeric:
            if value == maximum:
                _set_fill(ws.cell(row, column), GREEN)


def _confusion(metrics: dict[str, Any]) -> str:
    return json.dumps(metrics["confusion_matrix"], separators=(",", ":"))


def _update_summary(ws: Any, payload: dict[str, Any]) -> None:
    for modality, row in ANDROID_ROWS.items():
        ws.cell(row, 3).value = _metrics(payload, modality, "logreg_raw")["macro_f1"]
        ws.cell(row, 4).value = _metrics(payload, modality, "xgb_raw")["macro_f1"]
        ws.cell(row, 5).value = _metrics(payload, modality, "xgb_optuna_150t_d6")["macro_f1"]
        ws.cell(row, 6).value = "N/A"
    _reset_fills(ws, range(18, 21), range(2, 6))
    _highlight_maxima(ws, list(ANDROID_ROWS.values()), list(range(2, 6)))
    ws.cell(2, 1).value = (
        "Macro-F1 only; higher is better. Paper scores from different evaluation sets are excluded. "
        "Blank cells mean that a comparable classifier result is unavailable. Subject-OS values are "
        "three-seed means; the matched Qwen pilot did not clear the Stage-5 gate. Androids Interview "
        "includes the fine-tuned Qwen scores and the three raw hidden-state heads."
    )
    ws.cell(22, 1).value = (
        "Reading note: DAIC and CMDC were tuned using positive F1; Turkish and D3TEC were selected using "
        "Macro-F1. D3TEC audio checkpoints and hidden heads use the response-normalized hierarchy; "
        "Subject OS is outside this experiment. Androids Interview hidden heads use arithmetic probability "
        "means from window to turn to subject; fixed XGBoost is table-only in the positive-F1 graph."
    )
    ws.auto_filter.ref = "A5:F20"


def _update_xgboost_head(ws: Any, payload: dict[str, Any]) -> None:
    start = 38
    for offset, (modality, head) in enumerate(
        ( (modality, head) for modality in MODALITIES for head in HEADS )
    ):
        row = start + offset
        _copy_row_style(ws, 37, row, 10)
        metrics = _metrics(payload, modality, head)
        values = [
            "androids_interview",
            modality,
            HEAD_LABELS[head],
            metrics["accuracy"],
            metrics["positive_f1"],
            metrics["recall"],
            metrics["precision"],
            metrics["macro_f1"],
            metrics["auroc"],
            _confusion(metrics),
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row, column).value = value
    _reset_fills(ws, range(38, 47), range(1, 11))
    _highlight_maxima(ws, list(range(38, 47)), list(range(4, 10)))
    ws.tables["Table1"].ref = "A1:J46"
    ws.auto_filter.ref = "A1:J46"


def _update_combined(ws: Any, payload: dict[str, Any]) -> None:
    start = 56
    for offset, (modality, head) in enumerate(
        ( (modality, head) for modality in MODALITIES for head in HEADS )
    ):
        row = start + offset
        _copy_row_style(ws, 55, row, 13)
        metrics = _metrics(payload, modality, head)
        values = [
            "Androids Interview",
            "Hidden-state head",
            MODALITY_LABELS[modality],
            HEAD_LABELS[head],
            metrics["accuracy"],
            metrics["positive_f1"],
            metrics["precision"],
            metrics["recall"],
            metrics["macro_f1"],
            metrics["negative_f1"],
            metrics["auroc"],
            BASELINE_F1,
            _confusion(metrics),
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row, column).value = value
    _reset_fills(ws, range(53, 65), range(1, 14))
    _highlight_maxima(ws, list(range(53, 65)), list(range(5, 12)))
    ws.tables["Table_1"].ref = "A4:M64"
    ws.auto_filter.ref = "A4:M64"
    ws.cell(2, 1).value = (
        "Comparable in-repository results only; paper scores from different evaluation sets were removed. "
        "Green marks the highest numeric value within each dataset. D3TEC audio LLMs and hidden heads use "
        "the response-normalized hierarchy; each of 27 responses has one subject vote. Androids Interview "
        "includes fine-tuned Qwen scores and raw hidden-state heads; hidden-head probabilities are averaged "
        "window to turn to subject."
    )


def _update_graph(ws: Any, payload: dict[str, Any]) -> None:
    for modality, row in ANDROID_GRAPH_ROWS.items():
        ws.cell(row, 2).value = _metrics(payload, modality, "logreg_raw")["positive_f1"]
        ws.cell(row, 3).value = _metrics(payload, modality, "xgb_raw")["positive_f1"]
        ws.cell(row, 4).value = _metrics(payload, modality, "xgb_optuna_150t_d6")["positive_f1"]
    ws.cell(2, 1).value = (
        "Androids positive-F1 comparison; fixed XGBoost remains table-only. "
        "All three modalities use generic labels."
    )
    old_chart = ws._charts[0] if ws._charts else None
    ws._charts = []
    chart = BarChart()
    chart.type = "col"
    chart.style = old_chart.style if old_chart is not None else None
    chart.grouping = "clustered"
    chart.overlap = 0
    chart.title = "Androids Interview Positive F1"
    chart.y_axis.title = "Positive F1"
    chart.x_axis.title = "Modality"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.legend.position = "r"
    chart.height = old_chart.height if old_chart is not None else 7.5
    chart.width = old_chart.width if old_chart is not None else 15
    chart.add_data(Reference(ws, min_col=2, max_col=4, min_row=4, max_row=7), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=7))
    ws.add_chart(chart, "A10")


def _validate_workbook(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        if archive.testzip() is not None:
            raise ValueError("The updated workbook ZIP is corrupt.")
    wb = load_workbook(path, data_only=False)
    xgb = wb["XGBoost Head"]
    combined = wb["Combined Results"]
    graph = wb["Androids PosF1 Graph"]
    if xgb.tables["Table1"].ref != "A1:J46" or xgb.auto_filter.ref != "A1:J46":
        raise ValueError("XGBoost Head table/filter reference is incorrect.")
    if combined.tables["Table_1"].ref != "A4:M64" or combined.auto_filter.ref != "A4:M64":
        raise ValueError("Combined Results table/filter reference is incorrect.")
    if len(graph._charts) != 1 or len(graph._charts[0].ser) != 3:
        raise ValueError("Androids graph must contain exactly three series.")
    expected_series = ("B", "C", "D")
    for series, column in zip(graph._charts[0].ser, expected_series):
        if series.val is None or series.val.numRef is None:
            raise ValueError("Androids graph series has no numeric range.")
        if series.val.numRef.f != f"'Androids PosF1 Graph'!${column}$5:${column}$7":
            raise ValueError(f"Androids graph range is incorrect for column {column}.")
    for row in range(56, 65):
        if combined.cell(row, 3).value not in MODALITY_LABELS.values():
            raise ValueError("Androids Combined Results modality labels are not generic.")
        if combined.cell(row, 12).value != BASELINE_F1:
            raise ValueError("Androids all-positive baseline F1 is not 128/180.")
    forbidden = ("full_turn", "full-turn", "transcript-technique")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value.lower() for token in forbidden):
                    raise ValueError(f"Forbidden Androids wording remains in {ws.title}!{cell.coordinate}.")


def main() -> None:
    args = _parse_args()
    payload = _load_acceptance(args.acceptance)
    source = args.workbook.resolve()
    destination = (args.output or args.workbook).resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    if destination != source:
        destination.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source)
    _update_summary(wb["Summary"], payload)
    _update_xgboost_head(wb["XGBoost Head"], payload)
    _update_combined(wb["Combined Results"], payload)
    _update_graph(wb["Androids PosF1 Graph"], payload)

    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=f".{destination.name}.", suffix=".tmp", dir=destination.parent, delete=False
    ) as handle:
        temporary = Path(handle.name)
    try:
        wb.save(temporary)
        _validate_workbook(temporary)
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)
    _validate_workbook(destination)
    print(json.dumps({"status": "updated", "workbook": str(destination)}, indent=2))


if __name__ == "__main__":
    main()
