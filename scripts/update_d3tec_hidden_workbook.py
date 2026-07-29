#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import json
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


HEADS = (
    ("logreg_raw", "Logistic Regression"),
    ("xgb_raw", "XGBoost (fixed raw)"),
    (
        "xgb_optuna_raw_t150_d6_seed1337_inner1337",
        "XGBoost Optuna (150t d6, seed1337)",
    ),
)
MODALITIES = {
    "audio_text": "Audio + Text",
    "audio_only": "Audio only",
    "text_only": "Text only",
}
GREEN = PatternFill("solid", fgColor="E2F0D9")


def _copy_row_style(sheet, source_row: int, target_row: int, max_col: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_col + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy.copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy.copy(source.alignment)


def _headline_results(audit: dict[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    result = {}
    for row in audit["results"]:
        key = (row["modality"], row["variant"])
        if row["variant"] in {item[0] for item in HEADS}:
            result[key] = row["pooled_metrics"]
    expected = {(modality, variant) for modality in MODALITIES for variant, _ in HEADS}
    if set(result) != expected:
        raise ValueError(f"Acceptance audit is missing headline results: {sorted(expected-set(result))}.")
    return result


def _clear_and_highlight(sheet, row_groups: list[list[int]], columns: range) -> None:
    for rows in row_groups:
        for column in columns:
            numeric = [
                (row, float(sheet.cell(row, column).value))
                for row in rows
                if isinstance(sheet.cell(row, column).value, (int, float))
            ]
            for row in rows:
                cell = sheet.cell(row, column)
                if cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb in {
                    "00E2F0D9",
                    "E2F0D9",
                }:
                    cell.fill = PatternFill()
            if numeric:
                best = max(value for _, value in numeric)
                for row, value in numeric:
                    if value == best:
                        sheet.cell(row, column).fill = copy.copy(GREEN)


def update_workbook(workbook_path: Path, audit_path: Path) -> None:
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    if audit.get("status") != "passed":
        raise ValueError("Workbook updates require a passed acceptance audit.")
    headline = _headline_results(audit)
    workbook = load_workbook(workbook_path)

    summary = workbook["Summary"]
    summary_rows = {"audio_text": 15, "audio_only": 16, "text_only": 17}
    for modality, row in summary_rows.items():
        summary.cell(row, 3, headline[(modality, "logreg_raw")]["macro_f1"])
        summary.cell(row, 4, headline[(modality, "xgb_raw")]["macro_f1"])
        summary.cell(
            row,
            5,
            headline[(modality, "xgb_optuna_raw_t150_d6_seed1337_inner1337")][
                "macro_f1"
            ],
        )
        summary.cell(row, 6, "N/A")
    summary["A19"] = (
        "Reading note: DAIC and CMDC were tuned using positive F1; Turkish and "
        "D3TEC were selected using Macro-F1. D3TEC audio checkpoints and hidden "
        "heads use response-normalized segments with equal response votes; "
        "Subject OS is outside this experiment."
    )
    if summary._charts:
        chart = summary._charts[0]
        for series in chart.ser:
            if series.val and series.val.numRef:
                formula = series.val.numRef.f
                series.val.numRef.f = formula.rsplit("$", 1)[0] + "$17"
            if series.cat and series.cat.numRef:
                formula = series.cat.numRef.f
                series.cat.numRef.f = formula.rsplit("$", 1)[0] + "$17"
            if series.cat and series.cat.strRef:
                formula = series.cat.strRef.f
                series.cat.strRef.f = formula.rsplit("$", 1)[0] + "$17"

    xgb = workbook["XGBoost Head"]
    for row in range(29, 38):
        _copy_row_style(xgb, 29, row, 10)
    xgb_row = 29
    for modality in MODALITIES:
        for variant, label in HEADS:
            metrics = headline[(modality, variant)]
            values = (
                "d3tec",
                modality,
                label,
                metrics["accuracy"],
                metrics["positive_f1"],
                metrics["recall"],
                metrics["precision"],
                metrics["macro_f1"],
                metrics["auroc"],
                json.dumps(metrics["confusion_matrix"]),
            )
            for column, value in enumerate(values, 1):
                xgb.cell(xgb_row, column, value)
            xgb_row += 1
    xgb.tables["Table1"].ref = "A1:J37"
    _clear_and_highlight(
        xgb,
        [list(range(2, 11)), list(range(11, 20)), list(range(20, 29)), list(range(29, 38))],
        range(4, 10),
    )

    combined = workbook["Combined Results"]
    combined["A2"] = (
        "Comparable in-repository results only; paper scores from different "
        "evaluation sets were removed. Green marks the highest numeric value "
        "within each dataset. D3TEC audio LLMs and hidden heads use the "
        "response-normalized hierarchy; each of 27 responses has one subject vote."
    )
    for row in range(44, 53):
        _copy_row_style(combined, 44, row, 13)
    combined_row = 44
    all_positive = float(audit["all_positive_baseline"]["positive_f1"])
    for modality, modality_label in MODALITIES.items():
        for variant, label in HEADS:
            metrics = headline[(modality, variant)]
            values = (
                "D3TEC (Spanish)",
                "Hidden-state head",
                modality_label,
                label,
                metrics["accuracy"],
                metrics["positive_f1"],
                metrics["precision"],
                metrics["recall"],
                metrics["macro_f1"],
                metrics["negative_f1"],
                metrics["auroc"],
                all_positive,
                json.dumps(metrics["confusion_matrix"]),
            )
            for column, value in enumerate(values, 1):
                combined.cell(combined_row, column, value)
            combined_row += 1
    combined.tables["Table_1"].ref = "A4:M52"
    _clear_and_highlight(
        combined,
        [
            list(range(5, 17)),
            list(range(17, 29)),
            list(range(29, 41)),
            list(range(41, 53)),
        ],
        range(5, 12),
    )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(workbook_path)

    # Reopen through both the OOXML reader and ZIP parser; also assert the
    # topology that downstream readers and charts depend on.
    checked = load_workbook(workbook_path, data_only=False)
    if checked["XGBoost Head"].tables["Table1"].ref != "A1:J37":
        raise AssertionError("XGBoost Head table reference was not extended.")
    if checked["Combined Results"].tables["Table_1"].ref != "A4:M52":
        raise AssertionError("Combined Results table reference was not extended.")
    if len(checked["Summary"]._charts) != 1:
        raise AssertionError("Summary chart was lost or duplicated.")
    with zipfile.ZipFile(workbook_path) as archive:
        bad = archive.testzip()
        if bad is not None:
            raise ValueError(f"Corrupt XLSX ZIP member: {bad}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Populate D3TEC hidden-head workbook rows.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("depression_results_combined_with_posf1_graphs.xlsx"),
    )
    parser.add_argument("--audit", type=Path, required=True)
    args = parser.parse_args()
    update_workbook(args.workbook, args.audit)
    print(f"Updated and validated {args.workbook}")


if __name__ == "__main__":
    main()
